"""
Risk Karnesi Modülü
Sürekli Envanter Analizi - Kapsamlı Risk Puanlama ve Excel Rapor Üretimi

ChatGPT Standardı:
- RAW_PUAN: 0-100 (şiddet skoru)
- KATKI_PUAN = RAW_PUAN/100 * MAX_KATKI[risk]
- Toplam mağaza puanı = SUM(KATKI_PUAN) normalize edilmiş 0-100
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
from datetime import datetime

# ==================== CHATGPT STANDARDI: MAX KATKI PUANLARI ====================
MAX_KATKI = {
    'ic_hirsizlik': 200,      # En kritik risk
    'acik_orani': 30,
    'sayim_disiplini': 15,
    'kronik_acik': 15,
    'kronik_fire': 10,
    'yuksek_sayim': 5,
    'ayni_sayim': 3,
    'tam_sayili': 2
}

# Toplam max = 280, normalize edilecek (0-100 arası)
TOPLAM_MAX_KATKI = sum(MAX_KATKI.values())  # 280

# ==================== EXCEL STİLLERİ ====================
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
KRITIK_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
RISKLI_FILL = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
DIKKAT_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
TEMIZ_FILL = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


# ==================== YARDIMCI FONKSİYONLAR ====================

def get_numeric_col(df, col_names, default=0):
    """DataFrame'den numeric kolon değeri al"""
    for col in col_names if isinstance(col_names, list) else [col_names]:
        if col in df.columns:
            return pd.to_numeric(df[col], errors='coerce').fillna(default)
    return pd.Series([default] * len(df), index=df.index)


def get_seviye(puan):
    """Puana göre risk seviyesi döndür"""
    if puan >= 60:
        return 'KRİTİK', KRITIK_FILL, '🔴'
    elif puan >= 40:
        return 'RİSKLİ', RISKLI_FILL, '🟠'
    elif puan >= 20:
        return 'DİKKAT', DIKKAT_FILL, '🟡'
    return 'TEMİZ', TEMIZ_FILL, '🟢'


# ==================== 1. AÇIK ORANI PUANI (RAW: 0-100) ====================

def hesapla_acik_orani_raw(magaza_acik_oran, bolge_acik_oran):
    """
    Açık Oranı RAW Puanı (0-100)

    Katsayı = |Mağaza Açık Oranı| / |Bölge Açık Oranı|

    Lineer interpolasyon:
    - Katsayı <= 0.5 → 0 puan
    - Katsayı = 1.0 → 20 puan
    - Katsayı = 1.5 → 45 puan
    - Katsayı = 2.0 → 70 puan
    - Katsayı >= 2.5 → 100 puan
    """
    if bolge_acik_oran == 0:
        bolge_acik_oran = 0.001

    katsayi = abs(magaza_acik_oran) / abs(bolge_acik_oran)

    # Lineer interpolasyon noktaları
    noktalar = [(0.0, 0), (0.5, 0), (1.0, 20), (1.5, 45), (2.0, 70), (2.5, 100)]

    if katsayi <= 0.5:
        raw_puan = 0
    elif katsayi >= 2.5:
        raw_puan = 100
    else:
        for i in range(len(noktalar) - 1):
            x1, y1 = noktalar[i]
            x2, y2 = noktalar[i + 1]
            if x1 <= katsayi <= x2:
                raw_puan = y1 + (y2 - y1) * (katsayi - x1) / (x2 - x1)
                break

    katki_puan = round(raw_puan / 100 * MAX_KATKI['acik_orani'], 2)

    return {
        'raw_puan': round(raw_puan, 2),
        'katki_puan': katki_puan,
        'katsayi': round(katsayi, 4),
        'magaza_oran': round(magaza_acik_oran * 100, 2),
        'bolge_oran': round(bolge_acik_oran * 100, 2)
    }


# ==================== 2. İÇ HIRSIZLIK PUANI (RAW: 0-100) ====================

def hesapla_ic_hirsizlik_raw(df_magaza):
    """
    İç Hırsızlık RAW Puanı (0-100)

    Filtre:
    - iptal_satir_tutari <= -200 TL
    - fark_miktari < 0
    - iptal_satir_miktari != 0

    Risk Seviyeleri:
    - ÇOK BÜYÜK RİSK: (fark_miktari - iptal_satir_miktari) == 0
    - YÜKSEK RİSK: |iptal| ∈ [0.5×|fark|, 2×|fark|]

    RAW = MIN(100, supheli_satir×2 + cok_buyuk×8)
    """
    if df_magaza is None or df_magaza.empty:
        return {'raw_puan': 0, 'katki_puan': 0, 'supheli_satir': 0, 'cok_buyuk_risk': 0, 'detay_df': pd.DataFrame()}

    # Kolonları al
    iptal_tutar = get_numeric_col(df_magaza, ['iptal_satir_tutari', 'İptal Satır Tutarı'])
    iptal_miktar = get_numeric_col(df_magaza, ['iptal_satir_miktari', 'İptal Satır Miktarı'])
    fark_miktar = get_numeric_col(df_magaza, ['fark_miktari', 'Fark Miktarı'])

    # Temel filtre: iptal <= -200, fark < 0, iptal != 0
    temel_mask = (iptal_tutar <= -200) & (fark_miktar < 0) & (iptal_miktar != 0)

    if not temel_mask.any():
        return {'raw_puan': 0, 'katki_puan': 0, 'supheli_satir': 0, 'cok_buyuk_risk': 0, 'detay_df': pd.DataFrame()}

    supheli_df = df_magaza[temel_mask].copy()

    # Risk seviyesi
    fark_abs = supheli_df['fark_miktari'].abs() if 'fark_miktari' in supheli_df.columns else get_numeric_col(supheli_df, ['fark_miktari', 'Fark Miktarı']).abs()
    iptal_abs = get_numeric_col(supheli_df, ['iptal_satir_miktari', 'İptal Satır Miktarı']).abs()
    fark_val = get_numeric_col(supheli_df, ['fark_miktari', 'Fark Miktarı'])
    iptal_val = get_numeric_col(supheli_df, ['iptal_satir_miktari', 'İptal Satır Miktarı'])

    # ÇOK BÜYÜK RİSK: fark - iptal ≈ 0
    cok_buyuk_mask = (fark_val - iptal_val).abs() < 0.01

    supheli_df['risk_seviyesi'] = 'YÜKSEK RİSK'
    supheli_df.loc[cok_buyuk_mask, 'risk_seviyesi'] = 'ÇOK BÜYÜK RİSK'

    supheli_satir = len(supheli_df)
    cok_buyuk_risk = cok_buyuk_mask.sum()

    # RAW puan
    ham_puan = (supheli_satir * 2) + (cok_buyuk_risk * 8)
    raw_puan = min(100, ham_puan)
    katki_puan = round(raw_puan / 100 * MAX_KATKI['ic_hirsizlik'], 2)

    return {
        'raw_puan': round(raw_puan, 2),
        'katki_puan': katki_puan,
        'supheli_satir': supheli_satir,
        'cok_buyuk_risk': cok_buyuk_risk,
        'detay_df': supheli_df
    }


# ==================== 3. YÜKSEK SAYIM PUANI (RAW: 0-100) ====================

def hesapla_yuksek_sayim_raw(df_magaza):
    """
    Yüksek Sayım RAW Puanı (0-100)

    Kural: sayim_miktari >= 50
    RAW = MIN(100, urun_sayisi × 5)
    """
    if df_magaza is None or df_magaza.empty:
        return {'raw_puan': 0, 'katki_puan': 0, 'urun_sayisi': 0, 'max_sayim': 0, 'detay_df': pd.DataFrame()}

    sayim = get_numeric_col(df_magaza, ['sayim_miktari', 'Sayım Miktarı'])

    yuksek_mask = sayim >= 50

    if not yuksek_mask.any():
        return {'raw_puan': 0, 'katki_puan': 0, 'urun_sayisi': 0, 'max_sayim': 0, 'detay_df': pd.DataFrame()}

    yuksek_df = df_magaza[yuksek_mask].copy()
    yuksek_df['sayim_miktari_calc'] = sayim[yuksek_mask]

    urun_sayisi = len(yuksek_df)
    max_sayim = sayim[yuksek_mask].max()

    raw_puan = min(100, urun_sayisi * 5)
    katki_puan = round(raw_puan / 100 * MAX_KATKI['yuksek_sayim'], 2)

    return {
        'raw_puan': round(raw_puan, 2),
        'katki_puan': katki_puan,
        'urun_sayisi': urun_sayisi,
        'max_sayim': round(max_sayim, 2),
        'detay_df': yuksek_df
    }


# ==================== 4. KRONİK AÇIK PUANI (RAW: 0-100) ====================

def hesapla_kronik_acik_raw(df_magaza):
    """
    Kronik Açık RAW Puanı (0-100)

    Kural: Ardışık 2 envanterde fark_tutari < -500 TL (delta bazlı)
    RAW = MIN(100, kronik_urun × 10)

    Not: Seri verisi yoksa 0 döner
    """
    KRONIK_ESIK = -500

    if df_magaza is None or df_magaza.empty:
        return {'raw_puan': 0, 'katki_puan': 0, 'urun_sayisi': 0, 'toplam_tutar': 0, 'seri_var': False, 'detay_df': pd.DataFrame()}

    # Gerekli kolonlar
    if 'malzeme_kodu' not in df_magaza.columns or 'envanter_sayisi' not in df_magaza.columns:
        return {'raw_puan': 0, 'katki_puan': 0, 'urun_sayisi': 0, 'toplam_tutar': 0, 'seri_var': False, 'detay_df': pd.DataFrame()}

    df = df_magaza.copy()
    df['envanter_sayisi'] = pd.to_numeric(df['envanter_sayisi'], errors='coerce')
    df['fark_tutari'] = get_numeric_col(df, ['fark_tutari', 'Fark Tutarı'])
    df = df.dropna(subset=['envanter_sayisi'])

    if df.empty:
        return {'raw_puan': 0, 'katki_puan': 0, 'urun_sayisi': 0, 'toplam_tutar': 0, 'seri_var': False, 'detay_df': pd.DataFrame()}

    # Seri var mı kontrol (en az 2 farklı envanter_sayisi)
    unique_env = df['envanter_sayisi'].nunique()
    if unique_env < 2:
        return {'raw_puan': 0, 'katki_puan': 0, 'urun_sayisi': 0, 'toplam_tutar': 0, 'seri_var': False, 'detay_df': pd.DataFrame()}

    df = df.sort_values(['malzeme_kodu', 'envanter_sayisi'])

    kronik_urunler = []

    for malzeme, grup in df.groupby('malzeme_kodu'):
        if len(grup) < 2:
            continue

        grup = grup.sort_values('envanter_sayisi').reset_index(drop=True)

        # Delta hesapla
        grup['prev_fark'] = grup['fark_tutari'].shift(1)
        grup['delta'] = grup['fark_tutari'] - grup['prev_fark'].fillna(0)
        grup['prev_delta'] = grup['delta'].shift(1)
        grup['prev_env'] = grup['envanter_sayisi'].shift(1)

        # Ardışık kontrol
        ardisik_mask = (
            grup['prev_delta'].notna() &
            (grup['prev_delta'] < KRONIK_ESIK) &
            (grup['delta'] < KRONIK_ESIK) &
            (grup['envanter_sayisi'] == grup['prev_env'] + 1)
        )

        if ardisik_mask.any():
            hit = grup[ardisik_mask].iloc[0]
            kronik_urunler.append({
                'malzeme_kodu': malzeme,
                'malzeme_tanimi': hit.get('malzeme_tanimi', ''),
                'onceki_env': int(hit['prev_env']),
                'sonraki_env': int(hit['envanter_sayisi']),
                'onceki_delta': hit['prev_delta'],
                'sonraki_delta': hit['delta'],
                'toplam': hit['prev_delta'] + hit['delta']
            })

    urun_sayisi = len(kronik_urunler)
    toplam_tutar = sum(u['toplam'] for u in kronik_urunler) if kronik_urunler else 0

    raw_puan = min(100, urun_sayisi * 10)
    katki_puan = round(raw_puan / 100 * MAX_KATKI['kronik_acik'], 2)

    detay_df = pd.DataFrame(kronik_urunler) if kronik_urunler else pd.DataFrame()

    return {
        'raw_puan': round(raw_puan, 2),
        'katki_puan': katki_puan,
        'urun_sayisi': urun_sayisi,
        'toplam_tutar': round(toplam_tutar, 2),
        'seri_var': True,
        'detay_df': detay_df
    }


# ==================== 5. KRONİK FİRE PUANI (RAW: 0-100) ====================

def hesapla_kronik_fire_raw(df_magaza):
    """
    Kronik Fire RAW Puanı (0-100)

    Kural: Ardışık 2 envanterde fire_tutari < -500 TL (delta bazlı)
    RAW = MIN(100, kronik_urun × 10)
    """
    KRONIK_ESIK = -500

    if df_magaza is None or df_magaza.empty:
        return {'raw_puan': 0, 'katki_puan': 0, 'urun_sayisi': 0, 'toplam_tutar': 0, 'seri_var': False, 'detay_df': pd.DataFrame()}

    if 'malzeme_kodu' not in df_magaza.columns or 'envanter_sayisi' not in df_magaza.columns:
        return {'raw_puan': 0, 'katki_puan': 0, 'urun_sayisi': 0, 'toplam_tutar': 0, 'seri_var': False, 'detay_df': pd.DataFrame()}

    df = df_magaza.copy()
    df['envanter_sayisi'] = pd.to_numeric(df['envanter_sayisi'], errors='coerce')
    df['fire_tutari'] = get_numeric_col(df, ['fire_tutari', 'Fire Tutarı'])
    df = df.dropna(subset=['envanter_sayisi'])

    if df.empty:
        return {'raw_puan': 0, 'katki_puan': 0, 'urun_sayisi': 0, 'toplam_tutar': 0, 'seri_var': False, 'detay_df': pd.DataFrame()}

    unique_env = df['envanter_sayisi'].nunique()
    if unique_env < 2:
        return {'raw_puan': 0, 'katki_puan': 0, 'urun_sayisi': 0, 'toplam_tutar': 0, 'seri_var': False, 'detay_df': pd.DataFrame()}

    df = df.sort_values(['malzeme_kodu', 'envanter_sayisi'])

    kronik_urunler = []

    for malzeme, grup in df.groupby('malzeme_kodu'):
        if len(grup) < 2:
            continue

        grup = grup.sort_values('envanter_sayisi').reset_index(drop=True)

        grup['prev_fire'] = grup['fire_tutari'].shift(1)
        grup['delta'] = grup['fire_tutari'] - grup['prev_fire'].fillna(0)
        grup['prev_delta'] = grup['delta'].shift(1)
        grup['prev_env'] = grup['envanter_sayisi'].shift(1)

        ardisik_mask = (
            grup['prev_delta'].notna() &
            (grup['prev_delta'] < KRONIK_ESIK) &
            (grup['delta'] < KRONIK_ESIK) &
            (grup['envanter_sayisi'] == grup['prev_env'] + 1)
        )

        if ardisik_mask.any():
            hit = grup[ardisik_mask].iloc[0]
            kronik_urunler.append({
                'malzeme_kodu': malzeme,
                'malzeme_tanimi': hit.get('malzeme_tanimi', ''),
                'onceki_env': int(hit['prev_env']),
                'sonraki_env': int(hit['envanter_sayisi']),
                'onceki_delta': hit['prev_delta'],
                'sonraki_delta': hit['delta'],
                'toplam': hit['prev_delta'] + hit['delta']
            })

    urun_sayisi = len(kronik_urunler)
    toplam_tutar = sum(u['toplam'] for u in kronik_urunler) if kronik_urunler else 0

    raw_puan = min(100, urun_sayisi * 10)
    katki_puan = round(raw_puan / 100 * MAX_KATKI['kronik_fire'], 2)

    detay_df = pd.DataFrame(kronik_urunler) if kronik_urunler else pd.DataFrame()

    return {
        'raw_puan': round(raw_puan, 2),
        'katki_puan': katki_puan,
        'urun_sayisi': urun_sayisi,
        'toplam_tutar': round(toplam_tutar, 2),
        'seri_var': True,
        'detay_df': detay_df
    }


# ==================== 6. TAM SAYILI SAYIM PUANI (RAW: 0-100) ====================

def hesapla_tam_sayili_raw(df_magaza):
    """
    Tam Sayılı Sayım RAW Puanı (0-100)

    Kapsam: Meyve/Sebz ve Et-Tavuk (Ekmek hariç - paketli)
    Kural: sayim_miktari % 1 == 0 (tam sayı) ve istisna değil

    İstisnalar: PAKET, FİLE, ADET, BÜTÜN, TABAK, 500G, 1KG vb.

    RAW = MIN(100, urun_sayisi × 4)
    """
    if df_magaza is None or df_magaza.empty:
        return {'raw_puan': 0, 'katki_puan': 0, 'urun_sayisi': 0, 'buyuk_risk': 0, 'detay_df': pd.DataFrame()}

    # Depolama koşulu filtresi
    depolama_col = None
    for col in ['depolama_kosulu', 'Depolama Koşulu']:
        if col in df_magaza.columns:
            depolama_col = col
            break

    if depolama_col is None:
        return {'raw_puan': 0, 'katki_puan': 0, 'urun_sayisi': 0, 'buyuk_risk': 0, 'detay_df': pd.DataFrame()}

    # Sadece Meyve/Sebz ve Et-Tavuk (Ekmek hariç)
    kapsam_mask = df_magaza[depolama_col].str.contains('Meyve|Sebz|Et|Tavuk', case=False, na=False)
    kapsam_df = df_magaza[kapsam_mask].copy()

    if kapsam_df.empty:
        return {'raw_puan': 0, 'katki_puan': 0, 'urun_sayisi': 0, 'buyuk_risk': 0, 'detay_df': pd.DataFrame()}

    sayim = get_numeric_col(kapsam_df, ['sayim_miktari', 'Sayım Miktarı'])

    # Tam sayı kontrolü (sayim > 0 ve tam sayı)
    tam_sayi_mask = (sayim > 0) & (sayim % 1 == 0)

    # İstisna kontrolü
    malzeme_col = None
    for col in ['malzeme_tanimi', 'Malzeme Tanımı']:
        if col in kapsam_df.columns:
            malzeme_col = col
            break

    if malzeme_col:
        istisna_pattern = r'PAKET|FİLE|ADET|BÜTÜN|TABAK|500G|1KG|250G|300G|380G|200G'
        istisna_mask = kapsam_df[malzeme_col].str.contains(istisna_pattern, case=False, na=False)
        tam_sayi_mask = tam_sayi_mask & (~istisna_mask)

    supheli_df = kapsam_df[tam_sayi_mask].copy()

    if supheli_df.empty:
        return {'raw_puan': 0, 'katki_puan': 0, 'urun_sayisi': 0, 'buyuk_risk': 0, 'detay_df': pd.DataFrame()}

    urun_sayisi = len(supheli_df)

    # Büyük risk: fark = 0
    fark = get_numeric_col(supheli_df, ['fark_tutari', 'Fark Tutarı'])
    buyuk_risk = (fark == 0).sum()

    raw_puan = min(100, urun_sayisi * 4)
    katki_puan = round(raw_puan / 100 * MAX_KATKI['tam_sayili'], 2)

    return {
        'raw_puan': round(raw_puan, 2),
        'katki_puan': katki_puan,
        'urun_sayisi': urun_sayisi,
        'buyuk_risk': buyuk_risk,
        'detay_df': supheli_df
    }


# ==================== 7. AYNI SAYIM PUANI (RAW: 0-100) ====================

def hesapla_ayni_sayim_raw(df_magaza):
    """
    Aynı Sayım RAW Puanı (0-100)

    Kural: Aynı ürün için ardışık 2+ envanterde aynı sayım miktarı
    RAW = MIN(100, urun_sayisi × 5)

    Not: Seri verisi yoksa 0 döner
    """
    if df_magaza is None or df_magaza.empty:
        return {'raw_puan': 0, 'katki_puan': 0, 'urun_sayisi': 0, 'seri_var': False, 'detay_df': pd.DataFrame()}

    if 'malzeme_kodu' not in df_magaza.columns or 'envanter_sayisi' not in df_magaza.columns:
        return {'raw_puan': 0, 'katki_puan': 0, 'urun_sayisi': 0, 'seri_var': False, 'detay_df': pd.DataFrame()}

    df = df_magaza.copy()
    df['sayim'] = get_numeric_col(df, ['sayim_miktari', 'Sayım Miktarı'])
    df['envanter_sayisi'] = pd.to_numeric(df['envanter_sayisi'], errors='coerce')
    df = df.dropna(subset=['envanter_sayisi'])

    if df.empty:
        return {'raw_puan': 0, 'katki_puan': 0, 'urun_sayisi': 0, 'seri_var': False, 'detay_df': pd.DataFrame()}

    unique_env = df['envanter_sayisi'].nunique()
    if unique_env < 2:
        return {'raw_puan': 0, 'katki_puan': 0, 'urun_sayisi': 0, 'seri_var': False, 'detay_df': pd.DataFrame()}

    df = df.sort_values(['malzeme_kodu', 'envanter_sayisi'])

    ayni_urunler = []

    for malzeme, grup in df.groupby('malzeme_kodu'):
        if len(grup) < 2:
            continue

        grup = grup.sort_values('envanter_sayisi').reset_index(drop=True)
        sayimlar = grup['sayim'].tolist()

        ardisik_ayni = 0
        max_ardisik = 0

        for i in range(1, len(sayimlar)):
            if sayimlar[i] == sayimlar[i-1] and sayimlar[i] > 0:
                ardisik_ayni += 1
                max_ardisik = max(max_ardisik, ardisik_ayni + 1)
            else:
                ardisik_ayni = 0

        if max_ardisik >= 2:
            ayni_urunler.append({
                'malzeme_kodu': malzeme,
                'malzeme_tanimi': grup.iloc[0].get('malzeme_tanimi', ''),
                'depolama_kosulu': grup.iloc[0].get('depolama_kosulu', ''),
                'seri': ' → '.join([f"{int(e)}:{s:.0f}" for e, s in zip(grup['envanter_sayisi'], grup['sayim'])]),
                'ardisik_adet': max_ardisik,
                'fark_tutari': get_numeric_col(grup, ['fark_tutari', 'Fark Tutarı']).sum()
            })

    urun_sayisi = len(ayni_urunler)

    raw_puan = min(100, urun_sayisi * 5)
    katki_puan = round(raw_puan / 100 * MAX_KATKI['ayni_sayim'], 2)

    detay_df = pd.DataFrame(ayni_urunler) if ayni_urunler else pd.DataFrame()

    return {
        'raw_puan': round(raw_puan, 2),
        'katki_puan': katki_puan,
        'urun_sayisi': urun_sayisi,
        'seri_var': True,
        'detay_df': detay_df
    }


# ==================== 8. SAYIM DİSİPLİNİ PUANI (HAFTA BAĞIMSIZ - ORAN BAZLI) ====================

def hesapla_sayim_disiplini_raw(df_magaza):
    """
    Sayım Disiplini RAW Puanı (0-100) - ChatGPT Standardı

    HAFTA BAĞIMSIZ - ORAN BAZLI

    Kapsam: Meyve/Sebz, Et-Tavuk, Ekmek

    Formül:
    - beklenen = min(4, max(envanter_sayisi_scope)) (en az 1)
    - R0 = envanter==0 / toplam_scope (sıfır oranı)
    - Rmiss = 0<envanter<beklenen / toplam_scope (eksik oranı)
    - hic_yapmadi = (R0 == 1) → tüm ürünler sıfır
    - RAW = min(100, 100 * (0.65*R0 + 0.35*Rmiss) + (hic_yapmadi ? 15 : 0))
    """
    if df_magaza is None or df_magaza.empty:
        return {
            'raw_puan': 0, 'katki_puan': 0, 'toplam_urun': 0,
            'sifir_urun': 0, 'eksik_urun': 0, 'tam_urun': 0,
            'sifir_oran': 0, 'eksik_oran': 0, 'hic_yapmadi': False,
            'beklenen': 1, 'detay_sifir_df': pd.DataFrame(), 'detay_eksik_df': pd.DataFrame()
        }

    # Depolama koşulu filtresi
    depolama_col = None
    for col in ['depolama_kosulu', 'Depolama Koşulu']:
        if col in df_magaza.columns:
            depolama_col = col
            break

    if depolama_col is None:
        return {
            'raw_puan': 0, 'katki_puan': 0, 'toplam_urun': 0,
            'sifir_urun': 0, 'eksik_urun': 0, 'tam_urun': 0,
            'sifir_oran': 0, 'eksik_oran': 0, 'hic_yapmadi': False,
            'beklenen': 1, 'detay_sifir_df': pd.DataFrame(), 'detay_eksik_df': pd.DataFrame()
        }

    # Kapsam: Meyve/Sebz, Et-Tavuk, Ekmek
    kapsam_mask = df_magaza[depolama_col].str.contains('Meyve|Sebz|Et|Tavuk|Ekmek', case=False, na=False)
    kapsam_df = df_magaza[kapsam_mask].copy()

    if kapsam_df.empty:
        return {
            'raw_puan': 0, 'katki_puan': 0, 'toplam_urun': 0,
            'sifir_urun': 0, 'eksik_urun': 0, 'tam_urun': 0,
            'sifir_oran': 0, 'eksik_oran': 0, 'hic_yapmadi': False,
            'beklenen': 1, 'detay_sifir_df': pd.DataFrame(), 'detay_eksik_df': pd.DataFrame()
        }

    # Envanter sayısı
    env = get_numeric_col(kapsam_df, ['envanter_sayisi', 'Envanter Sayisi']).astype(int)
    kapsam_df['envanter'] = env

    toplam_urun = len(kapsam_df)

    # Beklenen = min(4, max(envanter_sayisi_scope)) - en az 1
    beklenen = max(1, min(4, env.max()))

    # Kategorilere ayır
    sifir_mask = kapsam_df['envanter'] == 0
    eksik_mask = (kapsam_df['envanter'] > 0) & (kapsam_df['envanter'] < beklenen)
    tam_mask = kapsam_df['envanter'] >= beklenen

    sifir_urun = sifir_mask.sum()
    eksik_urun = eksik_mask.sum()
    tam_urun = tam_mask.sum()

    # Oranlar
    R0 = sifir_urun / toplam_urun if toplam_urun > 0 else 0  # Sıfır oranı
    Rmiss = eksik_urun / toplam_urun if toplam_urun > 0 else 0  # Eksik oranı

    # Hiç yapmadı = tüm ürünler sıfır
    hic_yapmadi = (R0 == 1) and (toplam_urun > 0)

    # RAW puan formülü (ChatGPT standardı)
    raw_puan = min(100, 100 * (0.65 * R0 + 0.35 * Rmiss) + (15 if hic_yapmadi else 0))
    katki_puan = round(raw_puan / 100 * MAX_KATKI['sayim_disiplini'], 2)

    # Detay dataframe'leri
    detay_sifir_df = kapsam_df[sifir_mask].copy() if sifir_urun > 0 else pd.DataFrame()
    detay_eksik_df = kapsam_df[eksik_mask].copy() if eksik_urun > 0 else pd.DataFrame()

    return {
        'raw_puan': round(raw_puan, 2),
        'katki_puan': katki_puan,
        'toplam_urun': toplam_urun,
        'sifir_urun': sifir_urun,
        'eksik_urun': eksik_urun,
        'tam_urun': tam_urun,
        'sifir_oran': round(R0, 4),
        'eksik_oran': round(Rmiss, 4),
        'hic_yapmadi': hic_yapmadi,
        'beklenen': beklenen,
        'detay_sifir_df': detay_sifir_df,
        'detay_eksik_df': detay_eksik_df
    }


# ==================== MAĞAZA TOPLAM RİSK KARNESİ ====================

def hesapla_magaza_risk_karnesi(df_magaza, bolge_acik_oran):
    """
    Mağaza için tüm risk kategorilerini hesaplar.

    Returns:
        dict: Tüm risk kategorileri, RAW puanlar, KATKI puanlar ve toplam
    """
    # Mağaza açık oranı hesapla
    satis = get_numeric_col(df_magaza, ['satis_hasilati', 'Satış Hasılatı']).sum()
    fark = get_numeric_col(df_magaza, ['fark_tutari', 'Fark Tutarı']).sum()
    fire = get_numeric_col(df_magaza, ['fire_tutari', 'Fire Tutarı']).sum()
    toplam_acik = fark + fire
    magaza_acik_oran = abs(toplam_acik / satis) if satis > 0 else 0

    # 8 risk kategorisi hesapla
    r_acik = hesapla_acik_orani_raw(magaza_acik_oran, bolge_acik_oran)
    r_ic = hesapla_ic_hirsizlik_raw(df_magaza)
    r_yuksek = hesapla_yuksek_sayim_raw(df_magaza)
    r_kronik_acik = hesapla_kronik_acik_raw(df_magaza)
    r_kronik_fire = hesapla_kronik_fire_raw(df_magaza)
    r_tam = hesapla_tam_sayili_raw(df_magaza)
    r_ayni = hesapla_ayni_sayim_raw(df_magaza)
    r_disiplin = hesapla_sayim_disiplini_raw(df_magaza)

    # Toplam katkı puanı
    toplam_katki = (
        r_ic['katki_puan'] +
        r_acik['katki_puan'] +
        r_disiplin['katki_puan'] +
        r_kronik_acik['katki_puan'] +
        r_kronik_fire['katki_puan'] +
        r_yuksek['katki_puan'] +
        r_ayni['katki_puan'] +
        r_tam['katki_puan']
    )

    # Normalize (0-100 arası)
    toplam_puan = round((toplam_katki / TOPLAM_MAX_KATKI) * 100, 2)
    toplam_puan = min(100, toplam_puan)  # Cap at 100

    # Seviye belirleme
    seviye, fill, emoji = get_seviye(toplam_puan)

    # Top 3 risk (RAW puana göre)
    risk_puanlari = [
        ('İç Hırsızlık', r_ic['raw_puan'], r_ic['katki_puan']),
        ('Açık Oranı', r_acik['raw_puan'], r_acik['katki_puan']),
        ('Sayım Disiplini', r_disiplin['raw_puan'], r_disiplin['katki_puan']),
        ('Kronik Açık', r_kronik_acik['raw_puan'], r_kronik_acik['katki_puan']),
        ('Kronik Fire', r_kronik_fire['raw_puan'], r_kronik_fire['katki_puan']),
        ('Yüksek Sayım', r_yuksek['raw_puan'], r_yuksek['katki_puan']),
        ('Aynı Sayım', r_ayni['raw_puan'], r_ayni['katki_puan']),
        ('Tam Sayılı', r_tam['raw_puan'], r_tam['katki_puan'])
    ]
    top3 = sorted(risk_puanlari, key=lambda x: x[1], reverse=True)[:3]
    top3_str = ' | '.join([f"{r[0]}:{r[1]:.0f}" for r in top3 if r[1] > 0])

    # Teşhis cümlesi
    teshis_parts = []
    if r_ic['supheli_satir'] > 0:
        teshis_parts.append(f"İç hırsızlık şüphesi: {r_ic['supheli_satir']} satır")
    if r_acik['raw_puan'] > 50:
        teshis_parts.append(f"Yüksek açık oranı (katsayı: {r_acik['katsayi']:.1f}x)")
    if r_disiplin['hic_yapmadi']:
        teshis_parts.append("HİÇ SAYIM YAPMAMIŞ!")
    elif r_disiplin['sifir_oran'] > 0.3:
        teshis_parts.append(f"Sayım disiplini zayıf (%{r_disiplin['sifir_oran']*100:.0f} sıfır)")
    if r_yuksek['urun_sayisi'] > 0:
        teshis_parts.append(f"Yüksek sayım: {r_yuksek['urun_sayisi']} ürün")
    if r_tam['urun_sayisi'] > 0:
        teshis_parts.append(f"Tam sayılı şüphe: {r_tam['urun_sayisi']} ürün")

    teshis = ' | '.join(teshis_parts) if teshis_parts else 'Normal'

    return {
        'toplam_puan': toplam_puan,
        'toplam_katki': round(toplam_katki, 2),
        'seviye': seviye,
        'emoji': emoji,
        'top3_str': top3_str,
        'teshis': teshis,
        'toplam_satis': satis,
        'toplam_acik': toplam_acik,
        'acik_oran': magaza_acik_oran,
        'risk_acik_orani': r_acik,
        'risk_ic_hirsizlik': r_ic,
        'risk_yuksek_sayim': r_yuksek,
        'risk_kronik_acik': r_kronik_acik,
        'risk_kronik_fire': r_kronik_fire,
        'risk_tam_sayili': r_tam,
        'risk_ayni_sayim': r_ayni,
        'risk_sayim_disiplini': r_disiplin
    }


# ==================== VEKTÖREL HESAPLAMA (TÜM MAĞAZALAR) ====================

def hesapla_tum_magazalar_risk(df_bolge):
    """
    Bölgedeki tüm mağazalar için risk karnesi hesapla (vektörel).

    Returns:
        pd.DataFrame: Mağaza bazlı özet
        dict: Mağaza detayları
    """
    if df_bolge is None or df_bolge.empty:
        return pd.DataFrame(), {}

    # Bölge açık oranı hesapla
    bolge_satis = get_numeric_col(df_bolge, ['satis_hasilati', 'Satış Hasılatı']).sum()
    bolge_fark = get_numeric_col(df_bolge, ['fark_tutari', 'Fark Tutarı']).sum()
    bolge_fire = get_numeric_col(df_bolge, ['fire_tutari', 'Fire Tutarı']).sum()
    bolge_acik = bolge_fark + bolge_fire
    bolge_acik_oran = abs(bolge_acik / bolge_satis) if bolge_satis > 0 else 0

    # Mağaza kodu kontrolü
    mag_col = None
    for col in ['magaza_kodu', 'Mağaza Kodu']:
        if col in df_bolge.columns:
            mag_col = col
            break

    if mag_col is None:
        return pd.DataFrame(), {}

    magazalar = df_bolge[mag_col].unique()

    sonuclar = []
    detaylar = {}

    for mag_kodu in magazalar:
        df_mag = df_bolge[df_bolge[mag_col] == mag_kodu]

        # Mağaza bilgileri
        mag_adi = df_mag['magaza_tanim'].iloc[0] if 'magaza_tanim' in df_mag.columns else ''
        sm = df_mag['satis_muduru'].iloc[0] if 'satis_muduru' in df_mag.columns else ''
        bs = df_mag['bolge_sorumlusu'].iloc[0] if 'bolge_sorumlusu' in df_mag.columns else ''

        # Risk karnesi hesapla
        karne = hesapla_magaza_risk_karnesi(df_mag, bolge_acik_oran)

        sonuclar.append({
            'magaza_kodu': mag_kodu,
            'magaza_adi': mag_adi,
            'satis_muduru': sm,
            'bolge_sorumlusu': bs,
            'satis': karne['toplam_satis'],
            'toplam_acik': karne['toplam_acik'],
            'acik_oran': round(karne['acik_oran'] * 100, 2),
            'toplam_puan': karne['toplam_puan'],
            'seviye': karne['seviye'],
            'emoji': karne['emoji'],
            'top3': karne['top3_str'],
            'teshis': karne['teshis'],
            # RAW puanlar
            'raw_ic_hirsizlik': karne['risk_ic_hirsizlik']['raw_puan'],
            'raw_acik_orani': karne['risk_acik_orani']['raw_puan'],
            'raw_sayim_disiplini': karne['risk_sayim_disiplini']['raw_puan'],
            'raw_kronik_acik': karne['risk_kronik_acik']['raw_puan'],
            'raw_kronik_fire': karne['risk_kronik_fire']['raw_puan'],
            'raw_yuksek_sayim': karne['risk_yuksek_sayim']['raw_puan'],
            'raw_ayni_sayim': karne['risk_ayni_sayim']['raw_puan'],
            'raw_tam_sayili': karne['risk_tam_sayili']['raw_puan'],
            # Katkı puanlar
            'katki_ic_hirsizlik': karne['risk_ic_hirsizlik']['katki_puan'],
            'katki_acik_orani': karne['risk_acik_orani']['katki_puan'],
            'katki_sayim_disiplini': karne['risk_sayim_disiplini']['katki_puan'],
            'katki_kronik_acik': karne['risk_kronik_acik']['katki_puan'],
            'katki_kronik_fire': karne['risk_kronik_fire']['katki_puan'],
            'katki_yuksek_sayim': karne['risk_yuksek_sayim']['katki_puan'],
            'katki_ayni_sayim': karne['risk_ayni_sayim']['katki_puan'],
            'katki_tam_sayili': karne['risk_tam_sayili']['katki_puan'],
            # Detay sayılar
            'ic_supheli_satir': karne['risk_ic_hirsizlik']['supheli_satir'],
            'ic_cok_buyuk': karne['risk_ic_hirsizlik']['cok_buyuk_risk'],
            'yuksek_sayim_urun': karne['risk_yuksek_sayim']['urun_sayisi'],
            'tam_sayili_urun': karne['risk_tam_sayili']['urun_sayisi'],
            'disiplin_sifir': karne['risk_sayim_disiplini']['sifir_urun'],
            'disiplin_eksik': karne['risk_sayim_disiplini']['eksik_urun'],
            'disiplin_toplam': karne['risk_sayim_disiplini']['toplam_urun'],
            'disiplin_hic_yapmadi': karne['risk_sayim_disiplini']['hic_yapmadi']
        })

        detaylar[mag_kodu] = karne

    ozet_df = pd.DataFrame(sonuclar)
    ozet_df = ozet_df.sort_values('toplam_puan', ascending=False).reset_index(drop=True)

    return ozet_df, detaylar


# ==================== EXCEL ÜRETME FONKSİYONLARI ====================

def _style_header(ws, row, col_count):
    """Header satırını stillendir"""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER


def _auto_column_width(ws):
    """Kolon genişliklerini otomatik ayarla"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(50, max(10, max_length + 2))
        ws.column_dimensions[column_letter].width = adjusted_width


def uret_bolge_risk_karnesi_excel(df_bolge, bolge_adi, donem):
    """
    Bölge bazlı Risk Karnesi Excel dosyası üretir.

    Sekmeler:
    - 00_ÖZET: Tüm mağazalar özet
    - 01_İÇ_HIRSIZLIK: Detay
    - 02_AÇIK_ORANI: Detay
    - 03_SAYIM_DİSİPLİNİ: Detay
    - 04_KRONİK_AÇIK: Detay
    - 05_KRONİK_FİRE: Detay
    - 06_YÜKSEK_SAYIM: Detay
    - 07_AYNI_SAYIM: Detay
    - 08_TAM_SAYILI: Detay
    """
    # Hesapla
    ozet_df, detaylar = hesapla_tum_magazalar_risk(df_bolge)

    if ozet_df.empty:
        return None

    wb = Workbook()

    # ==================== 00_ÖZET ====================
    ws = wb.active
    ws.title = "00_ÖZET"

    # Başlık
    ws['A1'] = f"RİSK KARNESİ - {bolge_adi}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Dönem: {donem} | Oluşturma: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A2'].font = Font(italic=True, size=10)

    # Özet tablosu
    ozet_kolonlar = [
        'magaza_kodu', 'magaza_adi', 'satis_muduru', 'bolge_sorumlusu',
        'satis', 'toplam_acik', 'acik_oran', 'toplam_puan', 'seviye',
        'top3', 'teshis',
        'raw_ic_hirsizlik', 'raw_acik_orani', 'raw_sayim_disiplini',
        'raw_kronik_acik', 'raw_kronik_fire', 'raw_yuksek_sayim',
        'raw_ayni_sayim', 'raw_tam_sayili',
        'katki_ic_hirsizlik', 'katki_acik_orani', 'katki_sayim_disiplini',
        'katki_kronik_acik', 'katki_kronik_fire', 'katki_yuksek_sayim',
        'katki_ayni_sayim', 'katki_tam_sayili'
    ]

    header_row = 4
    for col, kolon in enumerate(ozet_kolonlar, 1):
        ws.cell(row=header_row, column=col, value=kolon.upper().replace('_', ' '))
    _style_header(ws, header_row, len(ozet_kolonlar))

    for row_idx, row_data in ozet_df[ozet_kolonlar].iterrows():
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=header_row + 1 + row_idx, column=col, value=value)
            cell.border = THIN_BORDER
            # Seviye kolonunu renklendir
            if ozet_kolonlar[col-1] == 'seviye':
                seviye_str = str(value)
                if 'KRİTİK' in seviye_str:
                    cell.fill = KRITIK_FILL
                elif 'RİSKLİ' in seviye_str:
                    cell.fill = RISKLI_FILL
                elif 'DİKKAT' in seviye_str:
                    cell.fill = DIKKAT_FILL
                else:
                    cell.fill = TEMIZ_FILL

    _auto_column_width(ws)

    # ==================== 01_İÇ_HIRSIZLIK ====================
    ws_ic = wb.create_sheet("01_İÇ_HIRSIZLIK")
    ws_ic['A1'] = "İÇ HIRSIZLIK RİSKİ - DETAY"
    ws_ic['A1'].font = Font(bold=True, size=14)

    ic_rows = []
    for mag_kodu, karne in detaylar.items():
        detay_df = karne['risk_ic_hirsizlik']['detay_df']
        if not detay_df.empty:
            for _, row in detay_df.head(50).iterrows():  # İlk 50 satır
                ic_rows.append({
                    'Mağaza Kodu': mag_kodu,
                    'Mağaza Adı': ozet_df[ozet_df['magaza_kodu']==mag_kodu]['magaza_adi'].iloc[0] if len(ozet_df[ozet_df['magaza_kodu']==mag_kodu]) > 0 else '',
                    'SM': ozet_df[ozet_df['magaza_kodu']==mag_kodu]['satis_muduru'].iloc[0] if len(ozet_df[ozet_df['magaza_kodu']==mag_kodu]) > 0 else '',
                    'Malzeme Kodu': row.get('malzeme_kodu', ''),
                    'Malzeme Tanımı': row.get('malzeme_tanimi', ''),
                    'İptal Tutarı': row.get('iptal_satir_tutari', 0),
                    'Fark Miktarı': row.get('fark_miktari', 0),
                    'Risk Seviyesi': row.get('risk_seviyesi', '')
                })

    if ic_rows:
        ic_df = pd.DataFrame(ic_rows)
        for col, header in enumerate(ic_df.columns, 1):
            ws_ic.cell(row=3, column=col, value=header)
        _style_header(ws_ic, 3, len(ic_df.columns))

        for row_idx, row_data in ic_df.iterrows():
            for col, value in enumerate(row_data, 1):
                ws_ic.cell(row=4 + row_idx, column=col, value=value).border = THIN_BORDER
        _auto_column_width(ws_ic)
    else:
        ws_ic['A3'] = "Şüpheli satır bulunamadı."

    # ==================== DİĞER SEKMELER (Benzer yapıda) ====================

    # 02_AÇIK_ORANI
    ws_acik = wb.create_sheet("02_AÇIK_ORANI")
    ws_acik['A1'] = "AÇIK ORANI RİSKİ - DETAY"
    ws_acik['A1'].font = Font(bold=True, size=14)

    acik_rows = []
    for _, row in ozet_df.iterrows():
        acik_rows.append({
            'Mağaza Kodu': row['magaza_kodu'],
            'Mağaza Adı': row['magaza_adi'],
            'SM': row['satis_muduru'],
            'BS': row['bolge_sorumlusu'],
            'Satış': row['satis'],
            'Toplam Açık': row['toplam_acik'],
            'Açık Oranı %': row['acik_oran'],
            'RAW Puan': row['raw_acik_orani'],
            'Katkı Puan': row['katki_acik_orani']
        })

    acik_df = pd.DataFrame(acik_rows)
    for col, header in enumerate(acik_df.columns, 1):
        ws_acik.cell(row=3, column=col, value=header)
    _style_header(ws_acik, 3, len(acik_df.columns))

    for row_idx, row_data in acik_df.iterrows():
        for col, value in enumerate(row_data, 1):
            ws_acik.cell(row=4 + row_idx, column=col, value=value).border = THIN_BORDER
    _auto_column_width(ws_acik)

    # 03_SAYIM_DİSİPLİNİ
    ws_dis = wb.create_sheet("03_SAYIM_DİSİPLİNİ")
    ws_dis['A1'] = "SAYIM DİSİPLİNİ RİSKİ - DETAY"
    ws_dis['A1'].font = Font(bold=True, size=14)

    dis_rows = []
    for _, row in ozet_df.iterrows():
        dis_rows.append({
            'Mağaza Kodu': row['magaza_kodu'],
            'Mağaza Adı': row['magaza_adi'],
            'SM': row['satis_muduru'],
            'Toplam Ürün': row['disiplin_toplam'],
            'Sıfır Ürün': row['disiplin_sifir'],
            'Eksik Ürün': row['disiplin_eksik'],
            'Hiç Yapmadı': 'EVET' if row['disiplin_hic_yapmadi'] else 'HAYIR',
            'RAW Puan': row['raw_sayim_disiplini'],
            'Katkı Puan': row['katki_sayim_disiplini']
        })

    dis_df = pd.DataFrame(dis_rows)
    for col, header in enumerate(dis_df.columns, 1):
        ws_dis.cell(row=3, column=col, value=header)
    _style_header(ws_dis, 3, len(dis_df.columns))

    for row_idx, row_data in dis_df.iterrows():
        for col, value in enumerate(row_data, 1):
            cell = ws_dis.cell(row=4 + row_idx, column=col, value=value)
            cell.border = THIN_BORDER
            if dis_df.columns[col-1] == 'Hiç Yapmadı' and value == 'EVET':
                cell.fill = KRITIK_FILL
    _auto_column_width(ws_dis)

    # 04-08 Diğer sekmeler (basitleştirilmiş)
    for i, (sekme, risk_key, baslik) in enumerate([
        ("04_KRONİK_AÇIK", "kronik_acik", "KRONİK AÇIK"),
        ("05_KRONİK_FİRE", "kronik_fire", "KRONİK FİRE"),
        ("06_YÜKSEK_SAYIM", "yuksek_sayim", "YÜKSEK SAYIM"),
        ("07_AYNI_SAYIM", "ayni_sayim", "AYNI SAYIM"),
        ("08_TAM_SAYILI", "tam_sayili", "TAM SAYILI")
    ], 4):
        ws_x = wb.create_sheet(sekme)
        ws_x['A1'] = f"{baslik} RİSKİ - DETAY"
        ws_x['A1'].font = Font(bold=True, size=14)

        rows = []
        for _, row in ozet_df.iterrows():
            rows.append({
                'Mağaza Kodu': row['magaza_kodu'],
                'Mağaza Adı': row['magaza_adi'],
                'SM': row['satis_muduru'],
                'RAW Puan': row[f'raw_{risk_key}'],
                'Katkı Puan': row[f'katki_{risk_key}']
            })

        temp_df = pd.DataFrame(rows)
        for col, header in enumerate(temp_df.columns, 1):
            ws_x.cell(row=3, column=col, value=header)
        _style_header(ws_x, 3, len(temp_df.columns))

        for row_idx, row_data in temp_df.iterrows():
            for col, value in enumerate(row_data, 1):
                ws_x.cell(row=4 + row_idx, column=col, value=value).border = THIN_BORDER
        _auto_column_width(ws_x)

    # Kaydet
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def uret_magaza_risk_raporu_excel(df_magaza, magaza_bilgi, bolge_acik_oran):
    """
    Tek mağaza için detaylı Risk Raporu Excel dosyası üretir.
    """
    karne = hesapla_magaza_risk_karnesi(df_magaza, bolge_acik_oran)

    wb = Workbook()
    ws = wb.active
    ws.title = "00_ÖZET"

    # Başlık
    ws['A1'] = f"MAĞAZA RİSK KARNESİ"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A2'] = f"{magaza_bilgi.get('kodu', '')} - {magaza_bilgi.get('adi', '')}"
    ws['A2'].font = Font(bold=True, size=14)

    # Bilgiler
    ws['A4'] = "Bölge:"
    ws['B4'] = magaza_bilgi.get('bolge', '')
    ws['A5'] = "SM:"
    ws['B5'] = magaza_bilgi.get('sm', '')
    ws['A6'] = "BS:"
    ws['B6'] = magaza_bilgi.get('bs', '')
    ws['A7'] = "Dönem:"
    ws['B7'] = magaza_bilgi.get('donem', '')

    # KPI
    ws['D4'] = "Satış:"
    ws['E4'] = f"₺{karne['toplam_satis']:,.0f}"
    ws['D5'] = "Toplam Açık:"
    ws['E5'] = f"₺{karne['toplam_acik']:,.0f}"
    ws['D6'] = "Açık Oranı:"
    ws['E6'] = f"%{karne['acik_oran']*100:.2f}"

    # Risk Puanı
    ws['A9'] = "TOPLAM RİSK PUANI:"
    ws['A9'].font = Font(bold=True, size=12)
    ws['B9'] = karne['toplam_puan']
    ws['B9'].font = Font(bold=True, size=14)
    ws['C9'] = karne['seviye']
    seviye, fill, _ = get_seviye(karne['toplam_puan'])
    ws['C9'].fill = fill

    # Top 3
    ws['A11'] = "En Yüksek 3 Risk:"
    ws['B11'] = karne['top3_str']

    # Teşhis
    ws['A12'] = "Teşhis:"
    ws['B12'] = karne['teshis']

    # Risk Kırılımı Tablosu
    ws['A14'] = "RİSK KIRILIMI"
    ws['A14'].font = Font(bold=True, size=12)

    headers = ['Risk Tipi', 'RAW Puan (0-100)', 'Max Katkı', 'Katkı Puan']
    for col, h in enumerate(headers, 1):
        ws.cell(row=15, column=col, value=h)
    _style_header(ws, 15, len(headers))

    risk_items = [
        ('İç Hırsızlık', karne['risk_ic_hirsizlik']['raw_puan'], MAX_KATKI['ic_hirsizlik'], karne['risk_ic_hirsizlik']['katki_puan']),
        ('Açık Oranı', karne['risk_acik_orani']['raw_puan'], MAX_KATKI['acik_orani'], karne['risk_acik_orani']['katki_puan']),
        ('Sayım Disiplini', karne['risk_sayim_disiplini']['raw_puan'], MAX_KATKI['sayim_disiplini'], karne['risk_sayim_disiplini']['katki_puan']),
        ('Kronik Açık', karne['risk_kronik_acik']['raw_puan'], MAX_KATKI['kronik_acik'], karne['risk_kronik_acik']['katki_puan']),
        ('Kronik Fire', karne['risk_kronik_fire']['raw_puan'], MAX_KATKI['kronik_fire'], karne['risk_kronik_fire']['katki_puan']),
        ('Yüksek Sayım', karne['risk_yuksek_sayim']['raw_puan'], MAX_KATKI['yuksek_sayim'], karne['risk_yuksek_sayim']['katki_puan']),
        ('Aynı Sayım', karne['risk_ayni_sayim']['raw_puan'], MAX_KATKI['ayni_sayim'], karne['risk_ayni_sayim']['katki_puan']),
        ('Tam Sayılı', karne['risk_tam_sayili']['raw_puan'], MAX_KATKI['tam_sayili'], karne['risk_tam_sayili']['katki_puan']),
    ]

    for row_idx, item in enumerate(risk_items, 16):
        for col, val in enumerate(item, 1):
            ws.cell(row=row_idx, column=col, value=val).border = THIN_BORDER

    # Toplam
    ws.cell(row=24, column=1, value="TOPLAM").font = Font(bold=True)
    ws.cell(row=24, column=3, value=TOPLAM_MAX_KATKI)
    ws.cell(row=24, column=4, value=karne['toplam_katki'])

    _auto_column_width(ws)

    # Kaydet
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
