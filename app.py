import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import zipfile
import json
import os
from supabase import create_client, Client

# Mobil uyumlu sayfa ayarı
st.set_page_config(page_title="Envanter Risk Analizi", layout="wide", page_icon="📊")

# ==================== CONFIG YÜKLEME ====================
def load_risk_weights():
    """Risk ağırlıklarını config dosyasından yükle"""
    config_path = os.path.join(os.path.dirname(__file__), 'weights.json')
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except:
        # Varsayılan değerler
        return {
            "risk_weights": {
                "toplam_oran": {"high": {"threshold": 2.0, "points": 40}, "medium": {"threshold": 1.5, "points": 25}, "low": {"threshold": 1.0, "points": 15}},
                "ic_hirsizlik": {"high": {"threshold": 50, "points": 30}, "medium": {"threshold": 30, "points": 20}, "low": {"threshold": 15, "points": 10}},
                "sigara": {"high": {"threshold": 5, "points": 35}, "low": {"threshold": 0, "points": 20}},
                "kronik": {"high": {"threshold": 100, "points": 15}, "low": {"threshold": 50, "points": 10}},
                "fire_manipulasyon": {"high": {"threshold": 10, "points": 20}, "low": {"threshold": 5, "points": 10}},
                "kasa_10tl": {"high": {"threshold": 20, "points": 15}, "low": {"threshold": 10, "points": 10}}
            },
            "risk_levels": {"kritik": 60, "riskli": 40, "dikkat": 20},
            "max_risk_score": 100
        }

RISK_CONFIG = load_risk_weights()

# ==================== GOOGLE SHEETS İPTAL VERİSİ (KAMERA ENTEGRASYONU) ====================
IPTAL_SHEETS_ID = '1F4Th-xZ2n0jDyayy5vayIN2j-EGUzqw5Akd8mXQVh4o'
IPTAL_SHEET_NAME = 'IptalVerisi'

@st.cache_data(ttl=300)  # 5 dakika cache
def get_iptal_verisi_from_sheets():
    """Google Sheets'ten iptal verisini çeker (public sheet gerekli) - CACHE YOK"""
    try:
        csv_url = f'https://docs.google.com/spreadsheets/d/{IPTAL_SHEETS_ID}/gviz/tq?tqx=out:csv&sheet={IPTAL_SHEET_NAME}'
        df = pd.read_csv(csv_url, encoding='utf-8')
        df.columns = df.columns.str.strip()
        return df
    except Exception as e:
        return pd.DataFrame()


def get_iptal_timestamps_for_magaza(magaza_kodu, malzeme_kodlari):
    """Belirli mağaza ve ürünler için iptal timestamp bilgilerini döner"""
    df_iptal = get_iptal_verisi_from_sheets()
    
    if df_iptal.empty:
        return {}
    
    # Sabit sütun isimleri - doğrudan kullan
    col_magaza = 'Mağaza - Anahtar'
    col_malzeme = 'Malzeme - Anahtar'
    col_tarih = 'Tarih - Anahtar'  # Tarih boş, Tarih - Anahtar dolu
    col_saat = 'Fiş Saati'
    col_miktar = 'Miktar'
    col_islem_no = 'İşlem Numarası'
    
    # Sütunlar yoksa index ile dene
    cols = df_iptal.columns.tolist()
    if col_magaza not in cols and len(cols) > 7:
        col_magaza = cols[7]
    if col_malzeme not in cols and len(cols) > 17:
        col_malzeme = cols[17]
    if col_tarih not in cols and len(cols) > 3:
        col_tarih = cols[3]
    if col_saat not in cols and len(cols) > 31:
        col_saat = cols[31]
    if col_islem_no not in cols and len(cols) > 36:
        col_islem_no = cols[36]
    
    # Mağaza ve Malzeme kodlarını temizle
    def clean_code(x):
        return str(x).strip().replace('.0', '')
    
    df_iptal[col_magaza] = df_iptal[col_magaza].apply(clean_code)
    df_iptal[col_malzeme] = df_iptal[col_malzeme].apply(clean_code)
    
    # Mağaza filtrele
    magaza_str = clean_code(magaza_kodu)
    df_mag = df_iptal[df_iptal[col_magaza] == magaza_str]
    
    if df_mag.empty:
        return {}
    
    # Malzeme kodlarını temizle
    malzeme_set = set(clean_code(m) for m in malzeme_kodlari)
    
    result = {}
    
    for _, row in df_mag.iterrows():
        malzeme = clean_code(row[col_malzeme])
        
        if malzeme not in malzeme_set:
            continue
        
        tarih = row.get(col_tarih, '')
        saat = row.get(col_saat, '')
        miktar = row.get(col_miktar, 0)
        islem_no = row.get(col_islem_no, '')
        
        if malzeme not in result:
            result[malzeme] = []
        
        result[malzeme].append({
            'tarih': tarih,
            'saat': saat,
            'miktar': miktar,
            'islem_no': islem_no
        })
    
    return result


def enrich_internal_theft_with_camera(internal_df, magaza_kodu, envanter_tarihi, full_df=None):
    """
    İç hırsızlık tablosuna kamera kontrol bilgisi ekler
    Eğer ürünün kendisi için iptal yoksa, aynı kategorideki 100+ TL ürünlerde iptal arar
    
    full_df: Tüm envanter verisi (kategori araması için gerekli)
    """
    if internal_df.empty:
        return internal_df
    
    df = internal_df.copy()
    
    # Envanter tarihini datetime'a çevir
    if isinstance(envanter_tarihi, str):
        try:
            envanter_tarihi = datetime.strptime(envanter_tarihi, '%Y-%m-%d')
        except:
            try:
                envanter_tarihi = datetime.strptime(envanter_tarihi, '%d.%m.%Y')
            except:
                envanter_tarihi = datetime.now()
    elif hasattr(envanter_tarihi, 'to_pydatetime'):
        envanter_tarihi = envanter_tarihi.to_pydatetime()
    
    # 15 gün öncesi (kamera erişim limiti)
    kamera_limit = envanter_tarihi - timedelta(days=15)
    
    # Malzeme kodlarını al
    malzeme_kodlari = df['Malzeme Kodu'].astype(str).tolist()
    
    # Kategori bilgisini al (Mal Grubu Tanımı)
    kategori_col = None
    for col in ['Mal Grubu Tanımı', 'Ürün Grubu', 'Ana Grup']:
        if col in df.columns:
            kategori_col = col
            break
    
    # Kategorideki tüm 100+ TL ürünleri bul (alternatif arama için)
    kategori_urunleri = {}
    if kategori_col and full_df is not None:
        for _, row in df.iterrows():
            kategori = row.get(kategori_col, '')
            if kategori and kategori not in kategori_urunleri:
                # Bu kategorideki 100+ TL ürünleri bul
                if kategori_col in full_df.columns and 'Satış Fiyatı' in full_df.columns:
                    kat_mask = (full_df[kategori_col] == kategori) & (full_df['Satış Fiyatı'] >= 100)
                    kat_urunler = full_df.loc[kat_mask, 'Malzeme Kodu'].astype(str).unique().tolist()
                    kategori_urunleri[kategori] = kat_urunler
    
    # Tüm kategori ürünlerinin iptal verilerini çek
    tum_kategori_kodlari = set()
    for kodlar in kategori_urunleri.values():
        tum_kategori_kodlari.update(kodlar)
    
    # İptal verilerini çek (hem direkt ürünler hem kategori ürünleri)
    tum_kodlar = list(set(malzeme_kodlari) | tum_kategori_kodlari)
    iptal_data = get_iptal_timestamps_for_magaza(magaza_kodu, tum_kodlar)
    
    # Yeni sütunlar
    kamera_kontrol = []
    
    for _, row in df.iterrows():
        malzeme_kodu = str(row['Malzeme Kodu']).strip()
        kategori = row.get(kategori_col, '') if kategori_col else ''
        
        # Önce direkt ürün için iptal ara
        sonuc = _ara_iptal_kaydi(malzeme_kodu, iptal_data, kamera_limit)
        
        if sonuc['bulundu']:
            # Ürünün kendisi için kayıt var
            kamera_kontrol.append(sonuc['detay'])
        else:
            # Ürün için kayıt yok, kategorideki diğer 100+ TL ürünlere bak
            alternatif_bulundu = False
            alternatif_detay = ""
            
            if kategori and kategori in kategori_urunleri:
                for alt_kod in kategori_urunleri[kategori]:
                    if alt_kod != malzeme_kodu:
                        alt_sonuc = _ara_iptal_kaydi(alt_kod, iptal_data, kamera_limit)
                        if alt_sonuc['bulundu']:
                            alternatif_bulundu = True
                            # Alternatif ürün adını bul
                            alt_ad = ""
                            if full_df is not None:
                                alt_rows = full_df[full_df['Malzeme Kodu'].astype(str) == alt_kod]
                                if len(alt_rows) > 0:
                                    alt_ad = alt_rows['Malzeme Tanımı'].iloc[0] if 'Malzeme Tanımı' in alt_rows.columns else alt_kod
                            
                            alternatif_detay = f"🔄 KATEGORİ: {alt_ad[:30] if alt_ad else alt_kod} → {alt_sonuc['detay']}"
                            break
            
            if alternatif_bulundu:
                kamera_kontrol.append(alternatif_detay)
            else:
                # Ne ürün ne kategori için kayıt yok
                kamera_kontrol.append(f"❌ {kategori} kategorisinde 100+ TL iptal yok" if kategori else "❌ İptal kaydı yok")
    
    df['KAMERA KONTROL DETAY'] = kamera_kontrol
    
    return df


def _ara_iptal_kaydi(malzeme_kodu, iptal_data, kamera_limit):
    """Bir ürün için iptal kaydı ara ve formatla"""
    if malzeme_kodu not in iptal_data:
        return {'bulundu': False, 'detay': ''}
    
    iptaller = iptal_data[malzeme_kodu]
    son_15_gun = []
    
    for iptal in iptaller:
        tarih_str = str(iptal['tarih'])
        
        try:
            for fmt in ['%d.%m.%Y', '%Y-%m-%d', '%d/%m/%Y']:
                try:
                    tarih = datetime.strptime(tarih_str.split()[0], fmt)
                    break
                except:
                    continue
            else:
                continue
            
            if tarih >= kamera_limit:
                son_15_gun.append({**iptal, 'tarih_dt': tarih})
        except:
            pass
    
    if not son_15_gun:
        return {'bulundu': False, 'detay': ''}
    
    # Tarihe göre sırala ve formatla
    son_15_gun_sorted = sorted(son_15_gun, key=lambda x: x['tarih_dt'], reverse=True)
    
    detaylar = []
    for iptal in son_15_gun_sorted[:3]:  # En fazla 3 kayıt göster
        tarih = iptal['tarih_dt'].strftime('%d.%m.%Y')
        saat = str(iptal.get('saat', ''))[:8]
        islem_no = str(iptal.get('islem_no', ''))
        
        # İşlem numarasından kasa numarasını çıkar (örn: 79150012711503250661 -> pozisyon 4-5)
        kasa_no = ""
        if len(islem_no) >= 6:
            try:
                kasa_no = f"Kasa:{int(islem_no[4:6])}"
            except:
                kasa_no = ""
        
        detaylar.append(f"{tarih} {saat} {kasa_no}".strip())
    
    return {
        'bulundu': True,
        'detay': "✅ KAMERA BAK " + " | ".join(detaylar)
    }


# ==================== SUPABASE BAĞLANTISI ====================
# Güvenlik: Credentials st.secrets'tan okunuyor
SUPABASE_URL = st.secrets.get("SUPABASE_URL", "https://tlcgcdiycgfxpxwzkwuf.supabase.co")
SUPABASE_KEY = st.secrets.get("SUPABASE_KEY", "")

@st.cache_resource
def get_supabase_client():
    from supabase import ClientOptions
    # Timeout'u 60 saniyeye çıkar (default 5 saniye)
    options = ClientOptions(
        postgrest_client_timeout=60,
    )
    return create_client(SUPABASE_URL, SUPABASE_KEY, options=options)

supabase: Client = get_supabase_client()

# ==================== GİRİŞ SİSTEMİ ====================
USERS = {
    "ziya": "Gm2025!",
    "sm1": "Sm12025!",
    "sm2": "Sm22025!",
    "sm3": "Sm32025!",
    "sm4": "Sm42025!",
    "sma": "Sma2025!",
}

def login():
    if "user" not in st.session_state:
        st.session_state.user = None
    
    if st.session_state.user is None:
        st.markdown("""
        <div style="max-width: 400px; margin: 100px auto; padding: 40px; 
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                    border-radius: 15px; text-align: center;">
            <h1 style="color: white;">📊 Envanter Risk Analizi</h1>
            <p style="color: #eee;">Mağaza Detay Analizi</p>
        </div>
        """, unsafe_allow_html=True)
        
        col1, col2, col3 = st.columns([1, 1, 1])
        with col2:
            st.markdown("### 🔐 Giriş Yap")
            username = st.text_input("Kullanıcı Adı", key="login_user")
            password = st.text_input("Şifre", type="password", key="login_pass")
            
            if st.button("Giriş", use_container_width=True):
                if username.lower() in USERS and USERS[username.lower()] == password:
                    st.session_state.user = username.lower()
                    st.rerun()
                else:
                    st.error("❌ Hatalı kullanıcı adı veya şifre")
        st.stop()

login()

# ==================== SUPABASE FONKSİYONLARI ====================

def save_to_supabase(df_original):
    """
    Excel verisini Supabase'e kaydet
    
    Duplicate kontrolü: Mağaza Kodu + Envanter Dönemi + Depolama Koşulu Grubu
    - Aynı kombinasyon zaten varsa → O envanter ATLANIR
    - Yoksa → Yüklenir
    
    NOT: Aynı ürün farklı depolama gruplarında OLAMAZ (Soğuk'ta olan Gıda Dışı'nda yok)
    """
    try:
        df = df_original.copy()
        
        # Gerekli sütunlar var mı kontrol
        required_cols = ['Mağaza Kodu', 'Depolama Koşulu Grubu', 'Envanter Dönemi', 'Malzeme Kodu']
        for col in required_cols:
            if col not in df.columns:
                return 0, 0, f"'{col}' sütunu bulunamadı"
        
        # Unique envanter kombinasyonları bul (Mağaza + Dönem + Depolama Grubu)
        df['_env_key'] = (df['Mağaza Kodu'].astype(str) + '|' + 
                         df['Envanter Dönemi'].astype(str) + '|' + 
                         df['Depolama Koşulu Grubu'].astype(str))
        
        unique_envs = df[['Mağaza Kodu', 'Envanter Dönemi', 'Depolama Koşulu Grubu', '_env_key']].drop_duplicates()
        
        # Supabase'de hangileri mevcut kontrol et
        existing_envs = set()
        for _, env_row in unique_envs.iterrows():
            try:
                result = supabase.table('envanter_veri').select('id').eq(
                    'magaza_kodu', str(env_row['Mağaza Kodu'])
                ).eq(
                    'envanter_donemi', str(env_row['Envanter Dönemi'])
                ).eq(
                    'depolama_kosulu_grubu', str(env_row['Depolama Koşulu Grubu'])
                ).limit(1).execute()
                
                if result.data and len(result.data) > 0:
                    existing_envs.add(env_row['_env_key'])
            except:
                pass
        
        # Sadece yeni envanterler
        new_env_keys = set(unique_envs['_env_key']) - existing_envs
        skipped_env_keys = existing_envs
        
        if not new_env_keys:
            skipped_list = [k.replace('|', ' / ') for k in skipped_env_keys]
            return 0, len(skipped_env_keys), f"Tüm envanterler zaten mevcut: {', '.join(skipped_list[:3])}..."
        
        # Sadece yeni envanterlerin verilerini filtrele
        df_new = df[df['_env_key'].isin(new_env_keys)].copy()
        
        # DataFrame içinde duplicate satırları kaldır (aynı malzeme kodu)
        duplicate_key_cols = ['Mağaza Kodu', 'Envanter Dönemi', 'Depolama Koşulu Grubu', 'Malzeme Kodu']
        df_new = df_new.drop_duplicates(subset=duplicate_key_cols, keep='last')
        
        # Sütun mapping
        col_mapping = {
            'Mağaza Kodu': 'magaza_kodu',
            'Mağaza Tanım': 'magaza_tanim',
            'Satış Müdürü': 'satis_muduru',
            'Bölge Sorumlusu': 'bolge_sorumlusu',
            'Depolama Koşulu Grubu': 'depolama_kosulu_grubu',
            'Depolama Koşulu': 'depolama_kosulu',
            'Envanter Dönemi': 'envanter_donemi',
            'Envanter Tarihi': 'envanter_tarihi',
            'Envanter Başlangıç Tarihi': 'envanter_baslangic_tarihi',
            'Ürün Grubu Kodu': 'urun_grubu_kodu',
            'Ürün Grubu Tanımı': 'urun_grubu_tanimi',
            'Mal Grubu Kodu': 'mal_grubu_kodu',
            'Mal Grubu Tanımı': 'mal_grubu_tanimi',
            'Malzeme Kodu': 'malzeme_kodu',
            'Malzeme Tanımı': 'malzeme_tanimi',
            'Satış Fiyatı': 'satis_fiyati',
            'Sayım Miktarı': 'sayim_miktari',
            'Sayım Tutarı': 'sayim_tutari',
            'Kaydi Miktar': 'kaydi_miktar',
            'Kaydi Tutar': 'kaydi_tutar',
            'Fark Miktarı': 'fark_miktari',
            'Fark Tutarı': 'fark_tutari',
            'Kısmi Envanter Miktarı': 'kismi_envanter_miktari',
            'Kısmi Envanter Tutarı': 'kismi_envanter_tutari',
            'Fire Miktarı': 'fire_miktari',
            'Fire Tutarı': 'fire_tutari',
            'Önceki Fark Miktarı': 'onceki_fark_miktari',
            'Önceki Fark Tutarı': 'onceki_fark_tutari',
            'Önceki Fire Miktarı': 'onceki_fire_miktari',
            'Önceki Fire Tutarı': 'onceki_fire_tutari',
            'Satış Miktarı': 'satis_miktari',
            'Satış Hasılatı': 'satis_hasilati',
            'İade Miktarı': 'iade_miktari',
            'İade Tutarı': 'iade_tutari',
            'İptal Fişteki Miktar': 'iptal_fisteki_miktar',
            'İptal Fiş Tutarı': 'iptal_fis_tutari',
            'İptal GP Miktarı': 'iptal_gp_miktari',
            'İptal GP Tutarı': 'iptal_gp_tutari',
            'İptal Satır Miktarı': 'iptal_satir_miktari',
            'İptal Satır Tutarı': 'iptal_satir_tutari',
        }
        
        # Veriyi hazırla
        records = []
        for _, row in df_new.iterrows():
            record = {}
            for excel_col, db_col in col_mapping.items():
                if excel_col in row.index:
                    val = row[excel_col]
                    if pd.isna(val):
                        val = None
                    elif isinstance(val, pd.Timestamp):
                        val = val.strftime('%Y-%m-%d')
                    elif isinstance(val, (np.integer, np.int64)):
                        val = int(val)
                    elif isinstance(val, (np.floating, np.float64)):
                        val = float(val) if not np.isnan(val) else None
                    record[db_col] = val
            records.append(record)
        
        # Batch insert
        batch_size = 500
        inserted = 0
        
        for i in range(0, len(records), batch_size):
            batch = records[i:i+batch_size]
            try:
                supabase.table('envanter_veri').insert(batch).execute()
                inserted += len(batch)
            except Exception as e:
                st.warning(f"Batch {i//batch_size + 1} hatası: {str(e)[:100]}")
        
        new_list = [k.replace('|', ' / ') for k in new_env_keys]
        return inserted, len(skipped_env_keys), f"Yüklenen: {', '.join(new_list[:3])}..."
        
    except Exception as e:
        return 0, 0, f"Hata: {str(e)}"


@st.cache_data(ttl=600)  # 10 dakika cache
# ⚠️ SİLİNDİ: get_available_periods_from_supabase
# Artık VIEW üzerinden alınıyor: get_available_periods_cached()


# ⚠️ SİLİNDİ: get_available_sms_from_supabase
# Artık VIEW üzerinden alınıyor: get_available_sms_cached()


@st.cache_data(ttl=600)  # 10 dakika cache
def get_available_stores_from_supabase():
    """Mevcut mağazaları al - dropdown için"""
    try:
        all_stores = {}
        offset = 0
        batch_size = 1000
        
        while True:
            result = supabase.table('envanter_veri').select('magaza_kodu,magaza_tanim').range(offset, offset + batch_size - 1).execute()
            if not result.data:
                break
            
            for r in result.data:
                if r.get('magaza_kodu'):
                    all_stores[r['magaza_kodu']] = r.get('magaza_tanim', '')
            
            if len(result.data) < batch_size:
                break
            offset += batch_size
            
            if offset > 50000:
                break
        
        return all_stores
    except:
        return {}


@st.cache_data(ttl=300, show_spinner=False)
def get_single_store_data(magaza_kodu, donemler=None):
    """
    Tek mağaza için veri çek - HIZLI
    Sadece belirli mağazanın verisini çeker, tüm bölgeyi değil
    """
    try:
        all_data = []
        batch_size = 1000
        offset = 0
        
        required_columns = ','.join([
            'magaza_kodu', 'magaza_tanim', 'satis_muduru', 'bolge_sorumlusu',
            'depolama_kosulu_grubu', 'depolama_kosulu', 'envanter_donemi', 'envanter_tarihi', 'envanter_baslangic_tarihi',
            'mal_grubu_tanimi', 'malzeme_kodu', 'malzeme_tanimi', 'satis_fiyati',
            'fark_miktari', 'fark_tutari', 'kismi_envanter_miktari', 'kismi_envanter_tutari',
            'fire_miktari', 'fire_tutari', 'onceki_fark_miktari', 'onceki_fire_miktari',
            'satis_miktari', 'satis_hasilati', 'iptal_satir_miktari'
        ])
        
        for _ in range(50):  # Max 50K satır
            query = supabase.table('envanter_veri').select(required_columns)
            query = query.eq('magaza_kodu', str(magaza_kodu))
            
            if donemler and len(donemler) > 0:
                query = query.in_('envanter_donemi', list(donemler))
            
            query = query.range(offset, offset + batch_size - 1)
            result = query.execute()
            
            if not result.data:
                break
            
            all_data.extend(result.data)
            
            if len(result.data) < batch_size:
                break
            
            offset += batch_size
        
        if not all_data:
            return pd.DataFrame()
        
        df = pd.DataFrame(all_data)
        
        reverse_mapping = {
            'magaza_kodu': 'Mağaza Kodu',
            'magaza_tanim': 'Mağaza Adı',
            'satis_muduru': 'Satış Müdürü',
            'bolge_sorumlusu': 'Bölge Sorumlusu',
            'depolama_kosulu_grubu': 'Depolama Koşulu Grubu',
            'depolama_kosulu': 'Depolama Koşulu',
            'envanter_donemi': 'Envanter Dönemi',
            'envanter_tarihi': 'Envanter Tarihi',
            'envanter_baslangic_tarihi': 'Envanter Başlangıç Tarihi',
            'mal_grubu_tanimi': 'Mal Grubu Tanımı',
            'malzeme_kodu': 'Malzeme Kodu',
            'malzeme_tanimi': 'Malzeme Tanımı',
            'satis_fiyati': 'Satış Fiyatı',
            'fark_miktari': 'Fark Miktarı',
            'fark_tutari': 'Fark Tutarı',
            'kismi_envanter_miktari': 'Kısmi Envanter Miktarı',
            'kismi_envanter_tutari': 'Kısmi Envanter Tutarı',
            'fire_miktari': 'Fire Miktarı',
            'fire_tutari': 'Fire Tutarı',
            'onceki_fark_miktari': 'Önceki Fark Miktarı',
            'onceki_fire_miktari': 'Önceki Fire Miktarı',
            'satis_miktari': 'Satış Miktarı',
            'satis_hasilati': 'Satış Tutarı',
            'iptal_satir_miktari': 'İptal Satır Miktarı'
        }
        
        df = df.rename(columns=reverse_mapping)
        return df
        
    except Exception as e:
        st.error(f"Veri çekme hatası: {e}")
        return pd.DataFrame()


def get_data_from_supabase(satis_muduru=None, donemler=None):
    """Supabase'den veri çek ve DataFrame'e çevir - Optimize edilmiş"""
    try:
        all_data = []
        batch_size = 1000  # Supabase max limit
        offset = 0
        max_iterations = 500  # Sonsuz döngü koruması (500K satır max)
        
        # Sadece gerekli sütunları çek
        required_columns = ','.join([
            'magaza_kodu', 'magaza_tanim', 'satis_muduru', 'bolge_sorumlusu',
            'depolama_kosulu_grubu', 'depolama_kosulu', 'envanter_donemi', 'envanter_tarihi', 'envanter_baslangic_tarihi',
            'mal_grubu_tanimi', 'malzeme_kodu', 'malzeme_tanimi', 'satis_fiyati',
            'fark_miktari', 'fark_tutari', 'kismi_envanter_miktari', 'kismi_envanter_tutari',
            'fire_miktari', 'fire_tutari', 'onceki_fark_miktari', 'onceki_fire_miktari',
            'satis_miktari', 'satis_hasilati', 'iptal_satir_miktari'
        ])
        
        iteration = 0
        while iteration < max_iterations:
            iteration += 1
            
            # Sorgu oluştur - sadece gerekli sütunlar
            query = supabase.table('envanter_veri').select(required_columns)
            
            if satis_muduru:
                query = query.eq('satis_muduru', satis_muduru)
            
            # Dönem filtresi
            if donemler and len(donemler) > 0:
                query = query.in_('envanter_donemi', donemler)
            
            # Pagination - limit ve offset
            query = query.range(offset, offset + batch_size - 1)
            
            result = query.execute()
            
            if not result.data or len(result.data) == 0:
                break
            
            all_data.extend(result.data)
            
            # Son batch'te batch_size'dan az veri geldiyse bitir
            if len(result.data) < batch_size:
                break
            
            offset += batch_size
        
        if not all_data:
            return pd.DataFrame()
        
        df = pd.DataFrame(all_data)
        
        # Sütun isimlerini geri çevir
        reverse_mapping = {
            'magaza_kodu': 'Mağaza Kodu',
            'magaza_tanim': 'Mağaza Adı',
            'satis_muduru': 'Satış Müdürü',
            'bolge_sorumlusu': 'Bölge Sorumlusu',
            'depolama_kosulu_grubu': 'Depolama Koşulu Grubu',
            'depolama_kosulu': 'Depolama Koşulu',
            'envanter_donemi': 'Envanter Dönemi',
            'envanter_tarihi': 'Envanter Tarihi',
            'envanter_baslangic_tarihi': 'Envanter Başlangıç Tarihi',
            'mal_grubu_tanimi': 'Mal Grubu Tanımı',
            'malzeme_kodu': 'Malzeme Kodu',
            'malzeme_tanimi': 'Malzeme Adı',
            'satis_fiyati': 'Satış Fiyatı',
            'fark_miktari': 'Fark Miktarı',
            'fark_tutari': 'Fark Tutarı',
            'kismi_envanter_miktari': 'Kısmi Envanter Miktarı',
            'kismi_envanter_tutari': 'Kısmi Envanter Tutarı',
            'fire_miktari': 'Fire Miktarı',
            'fire_tutari': 'Fire Tutarı',
            'onceki_fark_miktari': 'Önceki Fark Miktarı',
            'onceki_fire_miktari': 'Önceki Fire Miktarı',
            'satis_miktari': 'Satış Miktarı',
            'satis_hasilati': 'Satış Tutarı',
            'iptal_satir_miktari': 'İptal Satır Miktarı',
        }
        
        df = df.rename(columns=reverse_mapping)
        
        return df
        
    except Exception as e:
        st.error(f"Supabase hatası: {str(e)}")
        return pd.DataFrame()


@st.cache_data(ttl=900)  # 15 dakika cache
def get_sm_summary_from_view(satis_muduru=None, donemler=None, tarih_baslangic=None, tarih_bitis=None):
    """
    SM Özet ekranı için Supabase VIEW'den veri çek
    PAGINATION YOK - Tek sorguda tüm mağaza özetleri gelir (~200-300 satır)
    
    tarih_baslangic, tarih_bitis: Envanter tarihi aralığı filtresi (opsiyonel)
    """
    try:
        query = supabase.table('v_magaza_ozet').select('*')
        
        if satis_muduru:
            query = query.eq('satis_muduru', satis_muduru)
        
        if donemler and len(donemler) > 0:
            query = query.in_('envanter_donemi', donemler)
        
        # Tarih aralığı filtresi
        if tarih_baslangic:
            query = query.gte('envanter_tarihi', tarih_baslangic.strftime('%Y-%m-%d'))
        if tarih_bitis:
            query = query.lte('envanter_tarihi', tarih_bitis.strftime('%Y-%m-%d'))
        
        result = query.execute()
        
        if not result.data:
            return pd.DataFrame()
        
        df = pd.DataFrame(result.data)
        
        # Kolon isimlerini düzenle
        column_mapping = {
            'magaza_kodu': 'Mağaza Kodu',
            'magaza_tanim': 'Mağaza Adı',
            'satis_muduru': 'Satış Müdürü',
            'bolge_sorumlusu': 'Bölge Sorumlusu',
            'envanter_donemi': 'Envanter Dönemi',
            'envanter_tarihi': 'Envanter Tarihi',
            'envanter_baslangic_tarihi': 'Envanter Başlangıç Tarihi',
            'fark_tutari': 'Fark Tutarı',
            'kismi_tutari': 'Kısmi Tutarı',
            'fire_tutari': 'Fire Tutarı',
            'satis': 'Satış',
            'fark_miktari': 'Fark Miktarı',
            'kismi_miktari': 'Kısmi Miktarı',
            'onceki_fark_miktari': 'Önceki Fark Miktarı',
            'sigara_net': 'Sigara Net',
            'ic_hirsizlik': 'İç Hırs.',
            'kronik_acik': 'Kronik',
            'kronik_fire': 'Kronik Fire',
            'kasa_adet': 'Kasa Adet',
            'kasa_tutar': 'Kasa Tutar',
        }
        df = df.rename(columns=column_mapping)
        
        # Hesaplamalar
        df['Fark'] = df['Fark Tutarı'].fillna(0) + df['Kısmi Tutarı'].fillna(0)
        df['Fire'] = df['Fire Tutarı'].fillna(0)
        df['Toplam Açık'] = df['Fark'] + df['Fire']
        
        # Oranlar
        df['Fark %'] = (abs(df['Fark']) / df['Satış'] * 100).fillna(0)
        df['Fire %'] = (abs(df['Fire']) / df['Satış'] * 100).fillna(0)
        df['Toplam %'] = (abs(df['Toplam Açık']) / df['Satış'] * 100).fillna(0)
        
        # Gün hesabı
        try:
            df['Gün'] = (pd.to_datetime(df['Envanter Tarihi']) - 
                        pd.to_datetime(df['Envanter Başlangıç Tarihi'])).dt.days
            df['Gün'] = df['Gün'].apply(lambda x: max(1, x) if pd.notna(x) else 1)
        except:
            df['Gün'] = 1
        
        df['Günlük Fark'] = df['Fark'] / df['Gün']
        df['Günlük Fire'] = df['Fire'] / df['Gün']
        
        # Sigara açığı (negatifse açık var)
        df['Sigara'] = df['Sigara Net'].apply(lambda x: abs(x) if x < 0 else 0)
        
        # Bölge ortalamalarını hesapla (VIEW'den)
        bolge_ort = {
            'kayip_oran': df['Toplam %'].mean() if len(df) > 0 else 1,
            'ic_hirsizlik': df['İç Hırs.'].mean() if len(df) > 0 else 10,
            'kronik': df['Kronik'].mean() if len(df) > 0 else 50,
            'sigara': df['Sigara'].mean() if len(df) > 0 else 0,
        }
        
        # Risk puanı hesapla (tam formül)
        def calc_risk_score(row):
            """
            Risk puanı hesaplama (0-100)
            Ağırlıklar:
            - Kayıp Oranı: %30 (bölge ortalamasına göre)
            - Sigara Açığı: %30
            - İç Hırsızlık: %30 (bölge ortalamasına göre)
            - Kronik Açık: %5
            - 10TL Ürünleri: %5
            """
            puan = 0
            
            # Kayıp Oranı (30 puan) - Bölge ortalamasına göre
            kayip_oran = row.get('Toplam %', 0)
            if bolge_ort['kayip_oran'] > 0:
                kayip_ratio = kayip_oran / bolge_ort['kayip_oran']
                kayip_puan = min(30, kayip_ratio * 15)
            else:
                kayip_puan = min(30, kayip_oran * 20)
            puan += kayip_puan
            
            # Sigara Açığı (30 puan) - Her sigara kritik
            sigara_count = row.get('Sigara', 0)
            if sigara_count > 10:
                sigara_puan = 30
            elif sigara_count > 5:
                sigara_puan = 25
            elif sigara_count > 0:
                sigara_puan = sigara_count * 4
            else:
                sigara_puan = 0
            puan += sigara_puan
            
            # İç Hırsızlık (30 puan) - Bölge ortalamasına göre
            ic_hirsizlik_count = row.get('İç Hırs.', 0)
            if bolge_ort['ic_hirsizlik'] > 0:
                ic_ratio = ic_hirsizlik_count / bolge_ort['ic_hirsizlik']
                ic_puan = min(30, ic_ratio * 15)
            else:
                ic_puan = min(30, ic_hirsizlik_count * 0.5)
            puan += ic_puan
            
            # Kronik Açık (5 puan)
            kronik_count = row.get('Kronik', 0)
            if bolge_ort['kronik'] > 0:
                kronik_ratio = kronik_count / bolge_ort['kronik']
                kronik_puan = min(5, kronik_ratio * 2.5)
            else:
                kronik_puan = min(5, kronik_count * 0.05)
            puan += kronik_puan
            
            # 10TL Ürünleri (5 puan) - Fazla = şüpheli
            kasa_adet = abs(row.get('Kasa Adet', 0))
            if kasa_adet > 20:
                kasa_puan = 5
            elif kasa_adet > 10:
                kasa_puan = 3
            elif kasa_adet > 0:
                kasa_puan = 1
            else:
                kasa_puan = 0
            puan += kasa_puan
            
            return min(100, max(0, puan))
        
        df['Risk Puan'] = df.apply(calc_risk_score, axis=1)
        
        # Risk seviyesi (puana göre)
        def get_risk_level(puan):
            if puan >= 60:
                return '🔴 KRİTİK'
            elif puan >= 40:
                return '🟠 RİSKLİ'
            elif puan >= 20:
                return '🟡 DİKKAT'
            else:
                return '🟢 TEMİZ'
        
        df['Risk'] = df['Risk Puan'].apply(get_risk_level)
        
        # BS kolonu
        df['BS'] = df['Bölge Sorumlusu']
        
        return df
        
    except Exception as e:
        st.error(f"VIEW hatası: {str(e)}")
        return pd.DataFrame()


# ⚠️ SİLİNDİ: get_store_summary_fast
# Artık VIEW kullanılıyor: get_sm_summary_from_view()
# Bu fonksiyon performans katiliydi - mağaza mağaza loop yapıyordu


# ==================== ANA UYGULAMA ====================

# Çıkış butonu sağ üstte
col_title, col_user = st.columns([4, 1])
with col_title:
    st.title("🔍 Envanter Risk Analizi")
with col_user:
    st.markdown(f"👤 **{st.session_state.user.upper()}**")
    if st.button("🚪 Çıkış", key="logout_btn"):
        # Çıkışta cache'i temizle
        if "df_all" in st.session_state:
            del st.session_state.df_all
        if "df_all_analyzed" in st.session_state:
            del st.session_state.df_all_analyzed
        st.session_state.user = None
        st.rerun()

# ==================== VERİ YÜKLEME (1 KEZ) ====================
# ⚠️ SADECE TEK MAĞAZA MODU İÇİN - SM/GM Özet'te KULLANILMAMALI
# SM/GM Özet → get_sm_summary_from_view() kullanır

def load_all_data_once():
    """
    ⚠️ SADECE TEK MAĞAZA MODU İÇİN
    SM/GM Özet'te bu fonksiyon ÇAĞRILMAMALI - VIEW kullanılmalı
    """
    if "df_all" not in st.session_state or st.session_state.df_all is None:
        progress_text = st.empty()
        progress_bar = st.progress(0)
        
        progress_text.text("📊 Veriler yükleniyor...")
        progress_bar.progress(10)
        
        df_raw = get_data_from_supabase(satis_muduru=None, donemler=None)
        progress_bar.progress(70)
        
        if len(df_raw) > 0:
            progress_text.text("🔄 Analiz yapılıyor...")
            df_analyzed = analyze_inventory(df_raw)
            progress_bar.progress(90)
            
            # Duplicate'ları kaldır (aynı mağaza + dönem + depolama + malzeme)
            duplicate_cols = ['Mağaza Kodu', 'Envanter Dönemi', 'Depolama Koşulu Grubu', 'Malzeme Kodu']
            existing_cols = [c for c in duplicate_cols if c in df_analyzed.columns]
            
            if existing_cols:
                before_count = len(df_analyzed)
                df_analyzed = df_analyzed.drop_duplicates(subset=existing_cols, keep='last')
                after_count = len(df_analyzed)
                if before_count > after_count:
                    st.info(f"🧹 {before_count - after_count:,} duplicate kayıt kaldırıldı")
            
            st.session_state.df_all = df_analyzed
            st.session_state.df_all_loaded_at = datetime.now()
            progress_bar.progress(100)
            progress_text.text(f"✅ {len(df_analyzed):,} kayıt yüklendi")
        else:
            st.session_state.df_all = pd.DataFrame()
            progress_text.text("⚠️ Veri bulunamadı")
        
        # Progress bar'ı temizle
        import time
        time.sleep(0.5)
        progress_bar.empty()
        progress_text.empty()
        
    return st.session_state.df_all

def filter_data(df, satis_muduru=None, donemler=None, magaza_kodu=None):
    """DataFrame'i filtrele - Supabase çağırmadan"""
    if df is None or len(df) == 0:
        return pd.DataFrame()
    
    filtered = df.copy()
    
    if satis_muduru:
        filtered = filtered[filtered['Satış Müdürü'] == satis_muduru]
    
    if donemler and len(donemler) > 0:
        filtered = filtered[filtered['Envanter Dönemi'].isin(donemler)]
    
    if magaza_kodu:
        filtered = filtered[filtered['Mağaza Kodu'] == magaza_kodu]
    
    return filtered

@st.cache_data(ttl=300)
def get_available_periods_cached():
    """Dönemleri distinct VIEW'den al - HIZLI"""
    try:
        # v_distinct_donem VIEW'ı yoksa fallback
        try:
            result = supabase.table('v_distinct_donem').select('envanter_donemi').execute()
        except:
            # Fallback: ana tablodan distinct çek
            result = supabase.rpc('get_distinct_donemler').execute()
            if not result.data:
                # Son fallback
                result = supabase.table('envanter_veri').select('envanter_donemi').limit(1000).execute()
        
        if result.data:
            periods = list(set([r['envanter_donemi'] for r in result.data if r.get('envanter_donemi')]))
            return sorted(periods, reverse=True)
    except Exception as e:
        st.error(f"Dönem verisi alınamadı: {e}")
    return []

@st.cache_data(ttl=300)
def get_available_sms_cached():
    """SM'leri distinct VIEW'den al - HIZLI"""
    try:
        # v_distinct_sm VIEW'ı yoksa fallback
        try:
            result = supabase.table('v_distinct_sm').select('satis_muduru').execute()
        except:
            # Fallback
            result = supabase.table('envanter_veri').select('satis_muduru').limit(1000).execute()
        
        if result.data:
            sms = list(set([r['satis_muduru'] for r in result.data if r.get('satis_muduru')]))
            return sorted(sms)
    except Exception as e:
        st.error(f"SM verisi alınamadı: {e}")
    return []

@st.cache_data(ttl=300)
def get_envanter_tarihleri_by_donem(donemler_tuple):
    """Seçilen dönemlerdeki envanter tarihlerini getir - CACHED"""
    try:
        if not donemler_tuple:
            return []
        donemler = list(donemler_tuple)  # tuple'ı list'e çevir
        query = supabase.table('v_magaza_ozet').select('envanter_tarihi').in_('envanter_donemi', donemler)
        result = query.execute()
        if result.data:
            tarihler = list(set([r['envanter_tarihi'] for r in result.data if r.get('envanter_tarihi')]))
            # Tarihleri datetime'a çevir ve sırala
            tarih_dates = []
            for t in tarihler:
                try:
                    if isinstance(t, str):
                        tarih_dates.append(pd.to_datetime(t).date())
                    else:
                        tarih_dates.append(t)
                except:
                    pass
            return sorted(tarih_dates)
    except:
        pass
    return []

# Mobil uyumlu CSS
st.markdown("""
<style>
    .risk-kritik { background-color: #ff4444; color: white; padding: 10px; border-radius: 5px; text-align: center; font-weight: bold; }
    .risk-riskli { background-color: #ff8800; color: white; padding: 10px; border-radius: 5px; text-align: center; font-weight: bold; }
    .risk-dikkat { background-color: #ffcc00; color: black; padding: 10px; border-radius: 5px; text-align: center; font-weight: bold; }
    .risk-temiz { background-color: #00cc66; color: white; padding: 10px; border-radius: 5px; text-align: center; font-weight: bold; }
    
    /* Mobil uyumluluk */
    @media (max-width: 768px) {
        .stMetric { font-size: 0.8rem; }
        .stDataFrame { font-size: 0.7rem; }
        div[data-testid="column"] { padding: 0.25rem !important; }
    }
    
    /* Tablo kaydırma */
    .stDataFrame { overflow-x: auto; }
</style>
""", unsafe_allow_html=True)

# Mod seçimi - Kullanıcıya göre
current_user = st.session_state.user
is_gm = current_user == "ziya"

# Mod ve yenileme butonları
col_mode, col_refresh = st.columns([6, 1])

with col_mode:
    if is_gm:
        analysis_mode = st.radio("📊 Analiz Modu", ["🏪 Tek Mağaza", "🌍 Bölge Özeti", "👔 SM Özet", "🌍 GM Özet"], horizontal=True)
    else:
        analysis_mode = st.radio("📊 Analiz Modu", ["🏪 Tek Mağaza", "🌍 Bölge Özeti", "👔 SM Özet"], horizontal=True)

with col_refresh:
    if analysis_mode in ["👔 SM Özet", "🌍 GM Özet"]:
        if st.button("🔄", help="Verileri yenile"):
            if "df_all" in st.session_state:
                del st.session_state.df_all
            st.rerun()

# SM Özet ve GM Özet modları için dosya yükleme gerekmez
if analysis_mode not in ["👔 SM Özet", "🌍 GM Özet"]:
    # Dosya yükleme - direkt ekranda
    uploaded_file = st.file_uploader("📁 Excel dosyası yükleyin", type=['xlsx', 'xls'])
else:
    uploaded_file = None


def analyze_inventory(df):
    """Veriyi analiz için hazırla"""
    df = df.copy()
    
    # DUPLICATE TEMİZLEME - Doğru key ile
    # Aynı mağaza + dönem + depolama + malzeme sadece 1 kez olmalı
    dup_key = ['Mağaza Kodu', 'Envanter Dönemi', 'Depolama Koşulu Grubu', 'Malzeme Kodu']
    dup_key = [c for c in dup_key if c in df.columns]
    if dup_key:
        # Envanter tarihi varsa en yeniyi tut
        if 'Envanter Tarihi' in df.columns:
            df['Envanter Tarihi'] = pd.to_datetime(df['Envanter Tarihi'], errors='coerce')
            df = df.sort_values('Envanter Tarihi', ascending=False)
        df = df.drop_duplicates(subset=dup_key, keep='first')
    
    col_mapping = {
        'Mağaza Kodu': 'Mağaza Kodu',
        'Mağaza Tanım': 'Mağaza Adı',
        'Malzeme Kodu': 'Malzeme Kodu',
        'Malzeme Tanımı': 'Malzeme Adı',
        'Mal Grubu Tanımı': 'Ürün Grubu',
        'Ürün Grubu Tanımı': 'Ana Grup',
        'Fark Miktarı': 'Fark Miktarı',
        'Fark Tutarı': 'Fark Tutarı',
        'Kısmi Envanter Miktarı': 'Kısmi Envanter Miktarı',
        'Kısmi Envanter Tutarı': 'Kısmi Envanter Tutarı',
        'Önceki Fark Miktarı': 'Önceki Fark Miktarı',
        'Önceki Fark Tutarı': 'Önceki Fark Tutarı',
        'Önceki Fire Miktarı': 'Önceki Fire Miktarı',
        'Önceki Fire Tutarı': 'Önceki Fire Tutarı',
        'İptal Satır Miktarı': 'İptal Satır Miktarı',
        'İptal Satır Tutarı': 'İptal Satır Tutarı',
        'Fire Miktarı': 'Fire Miktarı',
        'Fire Tutarı': 'Fire Tutarı',
        'Satış Miktarı': 'Satış Miktarı',
        'Satış Hasılatı': 'Satış Tutarı',
        'Satış Fiyatı': 'Birim Fiyat',
        'Fark+Fire+Kısmi Envanter Tutarı': 'NET_ENVANTER_ETKİ_TUTARI',
        'Envanter Dönemi': 'Envanter Dönemi',
        'Envanter Tarihi': 'Envanter Tarihi',
    }
    
    for old_col, new_col in col_mapping.items():
        if old_col in df.columns:
            df[new_col] = df[old_col]
    
    numeric_cols = ['Fark Miktarı', 'Fark Tutarı', 'Kısmi Envanter Miktarı', 'Kısmi Envanter Tutarı',
                    'Önceki Fark Miktarı', 'Önceki Fark Tutarı', 'İptal Satır Miktarı', 'İptal Satır Tutarı',
                    'Fire Miktarı', 'Fire Tutarı', 'Satış Miktarı', 'Satış Tutarı', 'Önceki Fire Miktarı', 
                    'Önceki Fire Tutarı', 'Birim Fiyat']
    
    for col in numeric_cols:
        if col not in df.columns:
            df[col] = 0
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    
    if 'NET_ENVANTER_ETKİ_TUTARI' not in df.columns:
        df['NET_ENVANTER_ETKİ_TUTARI'] = df['Fark Tutarı'] + df['Fire Tutarı'] + df['Kısmi Envanter Tutarı']
    
    df['TOPLAM_MIKTAR'] = df['Fark Miktarı'] + df['Kısmi Envanter Miktarı'] + df['Önceki Fark Miktarı']
    
    return df


def is_balanced(row):
    """Dengelenmiş mi? Fark + Kısmi + Önceki = 0"""
    toplam = row['Fark Miktarı'] + row['Kısmi Envanter Miktarı'] + row['Önceki Fark Miktarı']
    return abs(toplam) <= 0.01


def get_first_two_words(text):
    """İlk 2 kelimeyi al"""
    if pd.isna(text):
        return ""
    words = str(text).strip().split()
    return " ".join(words[:2]).upper() if len(words) >= 2 else str(text).upper()


def get_last_word(text):
    """Son kelimeyi (marka) al"""
    if pd.isna(text):
        return ""
    words = str(text).strip().split()
    return words[-1].upper() if words else ""


def extract_quantity(text):
    """Gramaj/ML çıkar: '750 ML' → 750, 'ML'"""
    import re
    if pd.isna(text):
        return None, None
    
    text = str(text).upper()
    
    # Patterns: 750ML, 750 ML, 1.5L, 1,5 LT, 220G, 220 G, 1KG
    patterns = [
        r'(\d+[.,]?\d*)\s*(ML|LT|L|G|GR|KG|MG)\b',
    ]
    
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            value = float(match.group(1).replace(',', '.'))
            unit = match.group(2)
            
            # Normalize units to base (ML, G)
            if unit in ['LT', 'L']:
                value = value * 1000  # to ML
                unit = 'ML'
            elif unit == 'KG':
                value = value * 1000  # to G
                unit = 'G'
            elif unit == 'GR':
                unit = 'G'
            
            return value, unit
    
    return None, None


def is_quantity_similar(qty1, unit1, qty2, unit2, tolerance=0.30):
    """Gramaj benzer mi? Aynı boyut kategorisinde mi?"""
    if qty1 is None or qty2 is None:
        return True  # Gramaj bulunamadıysa benzer say
    
    if unit1 != unit2:
        return False  # Farklı birim (ML vs G) benzer değil
    
    if qty1 == 0 or qty2 == 0:
        return True
    
    # Oran kontrolü: max 3x fark olabilir
    ratio = max(qty1, qty2) / min(qty1, qty2)
    if ratio > 3:
        return False  # 3 kattan fazla fark varsa benzer değil
    
    # Boyut kategorileri
    def get_size_category(qty, unit):
        if unit == 'ML':
            if qty <= 400: return 'S'      # Küçük: 0-400ml
            elif qty <= 1000: return 'M'   # Orta: 400-1000ml
            else: return 'L'               # Büyük: 1000ml+
        elif unit == 'G':
            if qty <= 100: return 'S'      # Küçük: 0-100g
            elif qty <= 400: return 'M'    # Orta: 100-400g
            else: return 'L'               # Büyük: 400g+
        return 'M'
    
    cat1 = get_size_category(qty1, unit1)
    cat2 = get_size_category(qty2, unit2)
    
    # Sadece aynı kategorideyse benzer
    return cat1 == cat2


def detect_internal_theft(df):
    """
    İÇ HIRSIZLIK TESPİTİ:
    - Satış Fiyatı >= 100 TL
    - Dengelenmemiş (Fark + Kısmi + Önceki ≠ 0)
    - |Toplam| ≈ İptal Satır, fark büyüdükçe risk AZALIR
    """
    results = []
    
    for idx, row in df.iterrows():
        # Dengelenmiş ise atla
        if is_balanced(row):
            continue
        
        satis_fiyati = row.get('Birim Fiyat', 0) or 0
        if satis_fiyati < 100:
            continue
        
        fark = row['Fark Miktarı']
        kismi = row['Kısmi Envanter Miktarı']
        onceki = row['Önceki Fark Miktarı']
        iptal = row['İptal Satır Miktarı']
        
        toplam = fark + kismi + onceki
        
        if toplam >= 0 or iptal <= 0:
            continue
        
        fark_mutlak = abs(abs(toplam) - iptal)
        
        if fark_mutlak == 0:
            risk = "ÇOK YÜKSEK"
            esitlik = "TAM EŞİT"
        elif fark_mutlak <= 2:
            risk = "YÜKSEK"
            esitlik = "YAKIN (±2)"
        elif fark_mutlak <= 5:
            risk = "ORTA"
            esitlik = "YAKIN (±5)"
        elif fark_mutlak <= 10:
            risk = "DÜŞÜK-ORTA"
            esitlik = f"FARK: {fark_mutlak}"
        else:
            continue
        
        results.append({
            'Malzeme Kodu': row.get('Malzeme Kodu', ''),
            'Malzeme Adı': row.get('Malzeme Adı', ''),
            'Ürün Grubu': row.get('Mal Grubu Tanımı', row.get('Ürün Grubu', '')),
            'Satış Fiyatı': satis_fiyati,
            'Fark Miktarı': fark,
            'Kısmi Env.': kismi,
            'Önceki Fark': onceki,
            'TOPLAM': toplam,
            'İptal Satır': iptal,
            'Fark': fark_mutlak,
            'Durum': esitlik,
            'Fark Tutarı (TL)': row['Fark Tutarı'],
            'Risk': risk
        })
    
    result_df = pd.DataFrame(results)
    
    if len(result_df) > 0:
        # DUPLICATE TEMİZLEME - Aynı malzeme kodu sadece 1 kez görünsün
        result_df = result_df.drop_duplicates(subset=['Malzeme Kodu'], keep='first')
        
        # Risk sıralaması
        risk_order = {'ÇOK YÜKSEK': 0, 'YÜKSEK': 1, 'ORTA': 2, 'DÜŞÜK-ORTA': 3}
        result_df['_risk_sort'] = result_df['Risk'].map(risk_order)
        result_df = result_df.sort_values(['_risk_sort', 'Fark Tutarı (TL)'], ascending=[True, True])
        result_df = result_df.drop('_risk_sort', axis=1)
    
    return result_df


def detect_chronic_products(df):
    """Kronik açık - her iki dönemde de Fark < 0"""
    results = []
    
    for idx, row in df.iterrows():
        if is_balanced(row):
            continue
        
        if row['Önceki Fark Miktarı'] < 0 and row['Fark Miktarı'] < 0:
            results.append({
                'Malzeme Kodu': row.get('Malzeme Kodu', ''),
                'Malzeme Adı': row.get('Malzeme Adı', ''),
                'Ürün Grubu': row.get('Mal Grubu Tanımı', row.get('Ürün Grubu', '')),
                'Bu Dönem Fark': row['Fark Miktarı'],
                'Bu Dönem Tutar': row['Fark Tutarı'],
                'Önceki Fark': row['Önceki Fark Miktarı'],
                'Önceki Tutar': row['Önceki Fark Tutarı'],
                'Toplam Tutar': row['Fark Tutarı'] + row['Önceki Fark Tutarı']
            })
    
    result_df = pd.DataFrame(results)
    if len(result_df) > 0:
        # DUPLICATE TEMİZLEME
        result_df = result_df.drop_duplicates(subset=['Malzeme Kodu'], keep='first')
        result_df = result_df.sort_values('Bu Dönem Tutar', ascending=True)
    
    return result_df


def detect_chronic_fire(df):
    """Kronik Fire - her iki dönemde de fire var VE dengelenmemiş"""
    results = []
    
    for idx, row in df.iterrows():
        onceki_fire = row.get('Önceki Fire Miktarı', 0) or 0
        bu_fire = row['Fire Miktarı']
        
        # Her iki dönemde de fire varsa
        if onceki_fire != 0 and bu_fire != 0:
            # Önceki Fark + Fark = 0 ise dengelenmiş, kronik değil
            onceki_fark = row.get('Önceki Fark Miktarı', 0) or 0
            bu_fark = row['Fark Miktarı']
            
            if abs(onceki_fark + bu_fark) <= 0.01:
                continue  # Dengelenmiş, kronik fire değil
            
            results.append({
                'Malzeme Kodu': row.get('Malzeme Kodu', ''),
                'Malzeme Adı': row.get('Malzeme Adı', ''),
                'Ürün Grubu': row.get('Mal Grubu Tanımı', row.get('Ürün Grubu', '')),
                'Bu Dönem Fire': bu_fire,
                'Bu Dönem Fire Tutarı': row['Fire Tutarı'],
                'Önceki Fire': onceki_fire,
                'Önceki Fire Tutarı': row.get('Önceki Fire Tutarı', 0),
                'Toplam Fire Tutarı': row['Fire Tutarı'] + row.get('Önceki Fire Tutarı', 0)
            })
    
    result_df = pd.DataFrame(results)
    if len(result_df) > 0:
        # DUPLICATE TEMİZLEME
        result_df = result_df.drop_duplicates(subset=['Malzeme Kodu'], keep='first')
        result_df = result_df.sort_values('Bu Dönem Fire Tutarı', ascending=True)
    
    return result_df


def detect_fire_manipulation(df):
    """Fire manipülasyonu: Fire var AMA Fark+Kısmi > 0 VE dengelenmemiş"""
    results = []
    
    for idx, row in df.iterrows():
        fark = row['Fark Miktarı']
        kismi = row['Kısmi Envanter Miktarı']
        onceki_fark = row.get('Önceki Fark Miktarı', 0) or 0
        fire = row['Fire Miktarı']
        
        fark_kismi = fark + kismi
        
        # Önceki Fark + Fark = 0 ise dengelenmiş, manipülasyon değil
        if abs(onceki_fark + fark) <= 0.01:
            continue
        
        if fire < 0 and fark_kismi > 0:
            results.append({
                'Malzeme Kodu': row.get('Malzeme Kodu', ''),
                'Malzeme Adı': row.get('Malzeme Adı', ''),
                'Ürün Grubu': row.get('Mal Grubu Tanımı', row.get('Ürün Grubu', '')),
                'Fark Miktarı': fark,
                'Kısmi Env.': kismi,
                'Önceki Fark': onceki_fark,
                'Fark + Kısmi': fark_kismi,
                'Fire Miktarı': fire,
                'Fire Tutarı': row['Fire Tutarı'],
                'Sonuç': 'FAZLA FİRE GİRİLMİŞ'
            })
    
    result_df = pd.DataFrame(results)
    if len(result_df) > 0:
        # DUPLICATE TEMİZLEME
        result_df = result_df.drop_duplicates(subset=['Malzeme Kodu'], keep='first')
        result_df = result_df.sort_values('Fire Tutarı', ascending=True)
    
    return result_df


def detect_cigarette_shortage(df):
    """
    Sigara açığı - Tüm sigaraların TOPLAM (Fark + Kısmi + Önceki) değerine bakılır
    Eğer toplam < 0 ise sigara açığı var demektir
    
    NET = Fark Miktarı + Kısmi Envanter Miktarı + Önceki Fark Miktarı
    
    Sigara tespiti kuralları:
    - Mal Grubu Tanımı veya Ürün Grubu içinde 'SİGARA' veya 'TÜTÜN' geçenler
    - MAKARON tek başına sigara DEĞİLDİR (bilinçli olarak dışarıda tutulur)
    - "MAKARON JEL KALEM" gibi ürünler yanlışlıkla yakalanmasın diye MAKARON dahil edilmez
    """
    
    # Sigara kontrolü yapılacak kolonları belirle (öncelik sırasına göre)
    # NOT: Malzeme Adı dahil değil - sadece kategori bazlı filtre yapılır
    check_cols = []
    for col in ['Mal Grubu Tanımı', 'Ürün Grubu', 'Ana Grup']:
        if col in df.columns:
            check_cols.append(col)
    
    if not check_cols:
        return pd.DataFrame()
    
    # Sigara mask oluştur - CONTAINS kullan (eşitlik değil!)
    sigara_mask = pd.Series([False] * len(df), index=df.index)
    
    for col in check_cols:
        # Türkçe karakterleri normalize et
        col_values = df[col].fillna('').astype(str).str.upper()
        col_values = col_values.str.replace('İ', 'I', regex=False)
        col_values = col_values.str.replace('Ş', 'S', regex=False)
        col_values = col_values.str.replace('Ğ', 'G', regex=False)
        col_values = col_values.str.replace('Ü', 'U', regex=False)
        col_values = col_values.str.replace('Ö', 'O', regex=False)
        col_values = col_values.str.replace('Ç', 'C', regex=False)
        col_values = col_values.str.replace('ı', 'I', regex=False)
        
        # SIGARA veya TUTUN içeren satırları bul
        # NOT: MAKARON tek başına dahil DEĞİL - sadece SIGARA veya TUTUN varsa
        mask = col_values.str.contains('SIGARA|TUTUN', case=False, na=False, regex=True)
        sigara_mask = sigara_mask | mask
    
    # MAKARON'u açıkça dışarıda tut (eğer SIGARA/TUTUN yoksa)
    # Bu satır gereksiz görünebilir ama gelecekte güvenlik sağlar
    # Şu an mask zaten sadece SIGARA|TUTUN içerenleri yakalar
    
    sigara_df = df[sigara_mask].copy()
    
    if len(sigara_df) == 0:
        return pd.DataFrame()
    
    # Net hesapla: Fark + Kısmi + Önceki
    toplam_fark = sigara_df['Fark Miktarı'].fillna(0).sum()
    toplam_kismi = sigara_df['Kısmi Envanter Miktarı'].fillna(0).sum()
    toplam_onceki = sigara_df['Önceki Fark Miktarı'].fillna(0).sum()
    net_toplam = toplam_fark + toplam_kismi + toplam_onceki
    
    # Eğer net toplam < 0 ise açık var
    if net_toplam >= 0:
        return pd.DataFrame()
    
    # Açık varsa, detay göster
    results = []
    for idx, row in sigara_df.iterrows():
        fark = row['Fark Miktarı'] if pd.notna(row['Fark Miktarı']) else 0
        kismi = row['Kısmi Envanter Miktarı'] if pd.notna(row['Kısmi Envanter Miktarı']) else 0
        onceki = row['Önceki Fark Miktarı'] if pd.notna(row['Önceki Fark Miktarı']) else 0
        urun_net = fark + kismi + onceki
        
        # Sadece 0 olmayan kayıtları göster
        if fark != 0 or kismi != 0 or onceki != 0:
            results.append({
                'Malzeme Kodu': row.get('Malzeme Kodu', ''),
                'Malzeme Adı': row.get('Malzeme Adı', ''),
                'Fark': fark,
                'Kısmi': kismi,
                'Önceki': onceki,
                'Ürün Toplam': urun_net,
                'Risk': 'SİGARA'
            })
    
    result_df = pd.DataFrame(results)
    if len(result_df) > 0:
        # DUPLICATE TEMİZLEME
        result_df = result_df.drop_duplicates(subset=['Malzeme Kodu'], keep='first')
        result_df = result_df.sort_values('Ürün Toplam', ascending=True)
        # En sona toplam satırı ekle
        toplam_row = pd.DataFrame([{
            'Malzeme Kodu': '*** TOPLAM ***',
            'Malzeme Adı': f'SİGARA AÇIĞI: {abs(net_toplam):.0f} adet',
            'Fark': toplam_fark,
            'Kısmi': toplam_kismi,
            'Önceki': toplam_onceki,
            'Ürün Toplam': net_toplam,
            'Risk': '⚠️ AÇIK VAR'
        }])
        result_df = pd.concat([result_df, toplam_row], ignore_index=True)
    
    return result_df


def find_product_families(df):
    """
    Benzer ürün ailesi analizi
    Kural: İlk 2 kelime + Son kelime (marka) + Mal Grubu + Gramaj (±%30) aynıysa = AİLE
    """
    df_copy = df.copy()
    df_copy['İlk2Kelime'] = df_copy['Malzeme Adı'].apply(get_first_two_words)
    df_copy['Marka'] = df_copy['Malzeme Adı'].apply(get_last_word)
    df_copy['Gramaj'] = df_copy['Malzeme Adı'].apply(lambda x: extract_quantity(x)[0])
    df_copy['GramajBirim'] = df_copy['Malzeme Adı'].apply(lambda x: extract_quantity(x)[1])
    
    families = []
    processed_indices = set()
    
    # Her ürün için potansiyel aile bul
    for idx, row in df_copy.iterrows():
        if idx in processed_indices:
            continue
        
        ilk2 = row['İlk2Kelime']
        marka = row['Marka']
        urun_grubu = row['Ürün Grubu']
        gramaj = row['Gramaj']
        birim = row['GramajBirim']
        
        if not ilk2 or not marka:
            continue
        
        # Aynı grup içinde benzer ürünleri bul
        family_mask = (
            (df_copy['İlk2Kelime'] == ilk2) & 
            (df_copy['Marka'] == marka) & 
            (df_copy['Ürün Grubu'] == urun_grubu)
        )
        
        potential_family = df_copy[family_mask]
        
        if len(potential_family) <= 1:
            continue
        
        # Gramaj kontrolü - benzer gramajlı olanları filtrele
        family_members = []
        for fam_idx, fam_row in potential_family.iterrows():
            if is_quantity_similar(gramaj, birim, fam_row['Gramaj'], fam_row['GramajBirim']):
                family_members.append(fam_idx)
                processed_indices.add(fam_idx)
        
        if len(family_members) <= 1:
            continue
        
        family_df = df_copy.loc[family_members]
        
        toplam_fark = family_df['Fark Miktarı'].sum()
        toplam_kismi = family_df['Kısmi Envanter Miktarı'].sum()
        toplam_onceki = family_df['Önceki Fark Miktarı'].sum()
        aile_toplami = toplam_fark + toplam_kismi + toplam_onceki
        
        if family_df['Fark Miktarı'].abs().sum() > 0:
            if abs(aile_toplami) <= 2:
                sonuc = "KOD KARIŞIKLIĞI - HIRSIZLIK DEĞİL"
                risk = "DÜŞÜK"
            elif aile_toplami < -2:
                sonuc = "AİLEDE NET AÇIK VAR"
                risk = "ORTA"
            else:
                sonuc = "AİLEDE FAZLA VAR"
                risk = "DÜŞÜK"
            
            urunler = family_df['Malzeme Adı'].tolist()
            farklar = family_df['Fark Miktarı'].tolist()
            
            families.append({
                'Mal Grubu': urun_grubu,
                'İlk 2 Kelime': ilk2,
                'Marka': marka,
                'Ürün Sayısı': len(family_members),
                'Toplam Fark': toplam_fark,
                'Toplam Kısmi': toplam_kismi,
                'Toplam Önceki': toplam_onceki,
                'AİLE TOPLAMI': aile_toplami,
                'Sonuç': sonuc,
                'Risk': risk,
                'Ürünler': ' | '.join([f"{u[:25]}({f})" for u, f in zip(urunler[:5], farklar[:5])])
            })
    
    result_df = pd.DataFrame(families)
    if len(result_df) > 0:
        result_df = result_df.sort_values('AİLE TOPLAMI', ascending=True)
    
    return result_df


def detect_external_theft(df):
    """Dış hırsızlık - açık var ama fire/iptal yok"""
    results = []
    
    for idx, row in df.iterrows():
        if is_balanced(row):
            continue
        
        if row['Fark Miktarı'] < 0 and row['Fire Miktarı'] == 0 and row['İptal Satır Miktarı'] == 0:
            if abs(row['Fark Tutarı']) > 50:
                results.append({
                    'Malzeme Kodu': row.get('Malzeme Kodu', ''),
                    'Malzeme Adı': row.get('Malzeme Adı', ''),
                    'Ürün Grubu': row.get('Ürün Grubu', ''),
                    'Fark Miktarı': row['Fark Miktarı'],
                    'Fark Tutarı': row['Fark Tutarı'],
                    'Önceki Fark': row['Önceki Fark Miktarı'],
                    'Risk': 'DIŞ HIRSIZLIK / SAYIM HATASI'
                })
    
    result_df = pd.DataFrame(results)
    if len(result_df) > 0:
        result_df = result_df.sort_values('Fark Tutarı', ascending=True)
    
    return result_df


def check_kasa_activity_products(df, kasa_kodlari):
    """
    10 TL Ürünleri Kontrolü
    Fiyat değişikliği olan ürünlerde manipülasyon riski
    Toplam adet ve tutar etkisini hesapla
    FORMÜL: Fark + Kısmi (Önceki dahil değil)
    """
    results = []
    
    toplam_adet = 0
    toplam_tutar = 0
    eslesen_urun = 0
    
    for idx, row in df.iterrows():
        # Kod eşleştirme - hem string hem int formatını dene
        kod_raw = row.get('Malzeme Kodu', '')
        kod_str = str(kod_raw).replace('.0', '').strip()  # Float'tan gelen .0'ı kaldır
        
        if kod_str in kasa_kodlari:
            eslesen_urun += 1
            fark = row['Fark Miktarı'] if pd.notna(row['Fark Miktarı']) else 0
            kismi = row['Kısmi Envanter Miktarı'] if pd.notna(row['Kısmi Envanter Miktarı']) else 0
            toplam = fark + kismi  # Önceki dahil değil!
            
            # Tutar hesabı - Fark + Kısmi tutarları
            fark_tutari = row.get('Fark Tutarı', 0) or 0
            kismi_tutari = row.get('Kısmi Envanter Tutarı', 0) or 0
            urun_toplam_tutar = fark_tutari + kismi_tutari  # Önceki dahil değil!
            
            toplam_adet += toplam
            toplam_tutar += urun_toplam_tutar
            
            if toplam != 0:  # Sadece sıfır olmayanları göster
                if toplam > 0:
                    durum = "FAZLA (+)"
                else:
                    durum = "AÇIK (-)"
                
                results.append({
                    'Malzeme Kodu': kod_str,
                    'Malzeme Adı': row.get('Malzeme Adı', ''),
                    'Fark': fark,
                    'Kısmi': kismi,
                    'TOPLAM': toplam,
                    'Tutar': urun_toplam_tutar,
                    'Durum': durum
                })
    
    result_df = pd.DataFrame(results)
    if len(result_df) > 0:
        # Önce fazla (+) olanlar, sonra açık (-) olanlar
        result_df['_sort'] = result_df['TOPLAM'].apply(lambda x: 0 if x > 0 else 1)
        result_df = result_df.sort_values(['_sort', 'TOPLAM'], ascending=[True, False])
        result_df = result_df.drop('_sort', axis=1)
    
    # Özet bilgileri de döndür
    summary = {
        'toplam_urun': eslesen_urun,
        'sorunlu_urun': len(results),
        'toplam_adet': toplam_adet,
        'toplam_tutar': toplam_tutar
    }
    
    return result_df, summary


# 10 TL Ürünleri Ürün Kodları (209 adet)
# Bu ürünlerde fiyat değişikliği olduğu için manipülasyon riski var
KASA_AKTIVITESI_KODLARI = {
    '25006448', '12002256', '12002046', '22001972', '12003295', '22002759', '22002500', '11002886', '22002215', '22002214',
    '22002259', '22002349', '16002163', '22002717', '16001587', '13001073', '30000944', '18002488', '17003609', '22002296',
    '22002652', '24004136', '24004137', '12003073', '22002328', '24005228', '24006215', '24005232', '24005231', '24006214',
    '24006212', '16002332', '16002342', '23001397', '16002310', '24001063', '24004020', '13002613', '13002317', '13002506',
    '16002285', '16002219', '16002286', '16002218', '13000258', '13000257', '13000256', '13000260', '13002533', '22002611',
    '22002579', '13002559', '13000187', '13002904', '13000189', '13000190', '13002908', '13001872', '13001874', '30000838',
    '30000926', '22002605', '22002604', '22002603', '12003241', '16002194', '16001734', '25005580', '25000237', '25000049',
    '16002099', '23001367', '23001510', '23001177', '23001403', '23001278', '22002732', '22002576', '22002577', '25006483',
    '23001240', '16002317', '30000958', '30000956', '24005155', '24005154', '24005156', '24005157', '24005153', '22000280',
    '22002773', '22002774', '22002501', '22002225', '22000397', '22001395', '22000396', '16001859', '18002956', '17003542',
    '16002338', '16002339', '16002341', '16002009', '16000856', '22002715', '16002235', '24006067', '24006069', '24006068',
    '24006066', '22002686', '22002687', '22002688', '16002220', '24005291', '24005290', '24006078', '24006084', '24005288',
    '24006082', '24006079', '24005289', '24006085', '22002763', '22002762', '22001032', '18003049', '24006126', '24004420',
    '24005183', '24005649', '24005650', '14002481', '13002315', '22001229', '13002478', '30000880', '24005798', '24005796',
    '24005799', '24005797', '24005795', '24006159', '24003492', '24006171', '24006170', '24006174', '24006172', '24006173',
    '22002640', '22002553', '22002764', '22002223', '22002679', '22002221', '22002224', '22002572', '27002662', '24005441',
    '24005897', '24005898', '24005900', '24006081', '24006080', '16002087', '22002282', '22002283', '24005893', '24005894',
    '23001198', '23001439', '23001195', '23001199', '23000843', '23000034', '23001445', '23001444', '23001443', '23001522',
    '24004381', '24005184', '23001534', '23001533', '18001591', '27002676', '27002677', '16001956', '24003287', '24000005',
    '24002194', '24002192', '24002764', '24003872', '16001983', '18002969', '27001340', '27001148', '27001563', '24004354',
    '24004196', '24004115', '14002424', '24003641', '24004972', '13001481', '24003327', '24000004', '23000122',
}


def load_kasa_activity_codes():
    """Kasa aktivitesi ürün kodlarını döndür"""
    return KASA_AKTIVITESI_KODLARI


def generate_executive_summary(df, kasa_activity_df=None, kasa_summary=None):
    """Yönetici özeti - mal grubu bazlı yorumlar"""
    comments = []
    
    # Önce toplam tutarı hesapla (Fark + Kısmi + Önceki)
    df_copy = df.copy()
    df_copy['Kısmi Envanter Tutarı'] = df_copy.get('Kısmi Envanter Tutarı', 0).fillna(0)
    df_copy['Önceki Fark Tutarı'] = df_copy.get('Önceki Fark Tutarı', 0).fillna(0)
    df_copy['Toplam Tutar'] = df_copy['Fark Tutarı'] + df_copy['Kısmi Envanter Tutarı'] + df_copy['Önceki Fark Tutarı']
    
    # Mal grubu bazlı analiz
    group_stats = df_copy.groupby('Ürün Grubu').agg({
        'Toplam Tutar': 'sum',
        'Fire Tutarı': 'sum',
        'Satış Tutarı': 'sum',
        'Fark Miktarı': lambda x: (x < 0).sum()
    }).reset_index()
    
    group_stats.columns = ['Ürün Grubu', 'Toplam Fark', 'Toplam Fire', 'Toplam Satış', 'Açık Ürün Sayısı']
    group_stats['Açık Oranı'] = abs(group_stats['Toplam Fark']) / group_stats['Toplam Satış'].replace(0, 1) * 100
    
    # En yüksek açık
    top_acik = group_stats.nsmallest(3, 'Toplam Fark')
    for _, row in top_acik.iterrows():
        if row['Toplam Fark'] < -500:
            comments.append(f"⚠️ {row['Ürün Grubu']}: {row['Toplam Fark']:,.0f} TL açık ({row['Açık Ürün Sayısı']} ürün)")
    
    # En yüksek fire
    top_fire = group_stats.nsmallest(3, 'Toplam Fire')
    for _, row in top_fire.iterrows():
        if row['Toplam Fire'] < -500:
            comments.append(f"🔥 {row['Ürün Grubu']}: {row['Toplam Fire']:,.0f} TL fire")
    
    # 10 TL ürünleri yorumu - TOPLAM ADET VE TUTAR
    if kasa_summary is not None:
        toplam_adet = kasa_summary.get('toplam_adet', 0)
        toplam_tutar = kasa_summary.get('toplam_tutar', 0)
        
        if toplam_adet > 0:
            comments.append(f"💰 10 TL ÜRÜNLERİ: NET +{toplam_adet:.0f} adet / {toplam_tutar:,.0f} TL FAZLA")
            comments.append(f"   ⚠️ Bu fazlalık gerçek envanter açığını gizliyor olabilir!")
        elif toplam_adet < 0:
            comments.append(f"💰 10 TL ÜRÜNLERİ: NET {toplam_adet:.0f} adet / {toplam_tutar:,.0f} TL AÇIK")
    
    return comments, group_stats


def compute_sigara_acik_by_store(df: pd.DataFrame) -> pd.Series:
    """
    Sigara açığını mağaza bazında vektörel hesapla (10x hızlı)
    Loop yerine tek seferde tüm mağazalar için hesaplama yapar
    """
    # Sigara kontrol kolonları
    cols = [c for c in ['Mal Grubu Tanımı', 'Ürün Grubu', 'Ana Grup'] if c in df.columns]
    if not cols:
        return pd.Series(dtype=float)
    
    def norm_turkish(s: pd.Series) -> pd.Series:
        """Türkçe karakterleri normalize et"""
        s = s.fillna('').astype(str).str.upper()
        return (s.str.replace('İ', 'I', regex=False)
                 .str.replace('Ş', 'S', regex=False)
                 .str.replace('Ğ', 'G', regex=False)
                 .str.replace('Ü', 'U', regex=False)
                 .str.replace('Ö', 'O', regex=False)
                 .str.replace('Ç', 'C', regex=False)
                 .str.replace('ı', 'I', regex=False))
    
    # Sigara mask oluştur
    masks = []
    for c in cols:
        v = norm_turkish(df[c])
        masks.append(v.str.contains(r'SIGARA|TUTUN', regex=True, na=False))
    
    sig_mask = masks[0]
    for m in masks[1:]:
        sig_mask = sig_mask | m
    
    # Sigara ürünlerini filtrele
    required_cols = ['Mağaza Kodu', 'Fark Miktarı', 'Kısmi Envanter Miktarı', 'Önceki Fark Miktarı']
    available_cols = [c for c in required_cols if c in df.columns]
    
    if 'Mağaza Kodu' not in available_cols:
        return pd.Series(dtype=float)
    
    sig_df = df.loc[sig_mask, available_cols].copy()
    
    if sig_df.empty:
        return pd.Series(dtype=float)
    
    # Net değeri hesapla
    sig_df['net'] = sig_df.get('Fark Miktarı', pd.Series(0)).fillna(0)
    if 'Kısmi Envanter Miktarı' in sig_df.columns:
        sig_df['net'] += sig_df['Kısmi Envanter Miktarı'].fillna(0)
    if 'Önceki Fark Miktarı' in sig_df.columns:
        sig_df['net'] += sig_df['Önceki Fark Miktarı'].fillna(0)
    
    # Mağaza bazında topla
    net_by_store = sig_df.groupby('Mağaza Kodu')['net'].sum()
    
    # Net negatifse açık var → pozitif "açık adedi" olarak döndür
    sigara_acik = (-net_by_store).clip(lower=0)
    
    return sigara_acik


def analyze_region(df, kasa_kodlari):
    """Bölge geneli analiz - HIZLI VERSİYON (vektörel işlemler)"""
    
    magazalar = df['Mağaza Kodu'].dropna().unique().tolist()
    
    if len(magazalar) == 0:
        return pd.DataFrame()
    
    # Tüm mağazalar için temel metrikleri tek seferde hesapla (vektörel)
    store_metrics = df.groupby('Mağaza Kodu').agg({
        'Mağaza Adı': 'first',
        'Bölge Sorumlusu': 'first',
        'Satış Müdürü': 'first' if 'Satış Müdürü' in df.columns else 'first',
        'Satış Tutarı': 'sum',
        'Fark Tutarı': 'sum',
        'Kısmi Envanter Tutarı': 'sum',
        'Fire Tutarı': 'sum',
        'Envanter Tarihi': 'first',
        'Envanter Başlangıç Tarihi': 'first',
    }).reset_index()
    
    # SM sütunu kontrol
    if 'Satış Müdürü' not in df.columns:
        store_metrics['Satış Müdürü'] = ''
    
    # Hesaplamalar
    store_metrics['Fark'] = store_metrics['Fark Tutarı'].fillna(0) + store_metrics['Kısmi Envanter Tutarı'].fillna(0)
    store_metrics['Fire'] = store_metrics['Fire Tutarı'].fillna(0)
    store_metrics['Toplam Açık'] = store_metrics['Fark'] + store_metrics['Fire']
    store_metrics['Satış'] = store_metrics['Satış Tutarı'].fillna(0)
    
    # Oranlar
    store_metrics['Fark %'] = abs(store_metrics['Fark']) / store_metrics['Satış'] * 100
    store_metrics['Fire %'] = abs(store_metrics['Fire']) / store_metrics['Satış'] * 100
    store_metrics['Toplam %'] = abs(store_metrics['Toplam Açık']) / store_metrics['Satış'] * 100
    store_metrics[['Fark %', 'Fire %', 'Toplam %']] = store_metrics[['Fark %', 'Fire %', 'Toplam %']].fillna(0)
    
    # Gün hesabı
    try:
        store_metrics['Gün'] = (pd.to_datetime(store_metrics['Envanter Tarihi']) - 
                                pd.to_datetime(store_metrics['Envanter Başlangıç Tarihi'])).dt.days
        store_metrics['Gün'] = store_metrics['Gün'].apply(lambda x: max(1, x) if pd.notna(x) else 1)
    except:
        store_metrics['Gün'] = 1
    
    store_metrics['Günlük Fark'] = store_metrics['Fark'] / store_metrics['Gün']
    store_metrics['Günlük Fire'] = store_metrics['Fire'] / store_metrics['Gün']
    
    # ===== HIZLI RİSK ANALİZLERİ (vektörel) =====
    
    # 1. İç Hırsızlık - Satış Fiyatı >= 100 ve Fark < 0 olan ürün sayısı
    if 'Satış Fiyatı' in df.columns:
        ic_hirsizlik = df[(df['Satış Fiyatı'] >= 100) & (df['Fark Miktarı'] < 0)].groupby('Mağaza Kodu').size()
    else:
        ic_hirsizlik = pd.Series(0, index=magazalar)
    
    # 2. Kronik Açık - Önceki Fark < 0 ve Fark < 0 olan ürün sayısı
    kronik = df[(df['Önceki Fark Miktarı'] < 0) & (df['Fark Miktarı'] < 0)].groupby('Mağaza Kodu').size()
    
    # 3. Kronik Fire - Önceki Fire < 0 ve Fire < 0 olan ürün sayısı  
    if 'Önceki Fire Miktarı' in df.columns:
        kronik_fire = df[(df['Önceki Fire Miktarı'] < 0) & (df['Fire Miktarı'] < 0)].groupby('Mağaza Kodu').size()
    else:
        kronik_fire = pd.Series(0, index=magazalar)
    
    # 4. Sigara Açığı - VEKTÖREL HESAPLAMA (10x hızlı)
    sigara_acik_series = compute_sigara_acik_by_store(df)
    
    # 5. Fire Manipülasyonu - Fire > |Fark| olan ürün sayısı
    fire_manip = df[abs(df['Fire Miktarı']) > abs(df['Fark Miktarı'].fillna(0) + df['Kısmi Envanter Miktarı'].fillna(0))].groupby('Mağaza Kodu').size()
    
    # 6. 10TL Ürünleri - Kasa aktivitesi kodları
    kasa_set = set(str(k) for k in kasa_kodlari) if kasa_kodlari else set()
    if len(kasa_set) > 0:
        kasa_mask = df['Malzeme Kodu'].astype(str).isin(kasa_set)
        kasa_agg = df[kasa_mask].groupby('Mağaza Kodu').agg({
            'Fark Miktarı': 'sum',
            'Kısmi Envanter Miktarı': 'sum',
            'Fark Tutarı': 'sum',
            'Kısmi Envanter Tutarı': 'sum'
        })
        if len(kasa_agg) > 0:
            kasa_agg['10TL Adet'] = kasa_agg['Fark Miktarı'].fillna(0) + kasa_agg['Kısmi Envanter Miktarı'].fillna(0)
            kasa_agg['10TL Tutar'] = kasa_agg['Fark Tutarı'].fillna(0) + kasa_agg['Kısmi Envanter Tutarı'].fillna(0)
        else:
            kasa_agg = pd.DataFrame({'10TL Adet': [], '10TL Tutar': []})
    else:
        kasa_agg = pd.DataFrame({'10TL Adet': [], '10TL Tutar': []})
    
    # Sonuçları birleştir
    results = []
    
    # Risk config'i al
    rw = RISK_CONFIG.get('risk_weights', {})
    rl = RISK_CONFIG.get('risk_levels', {})
    max_score = RISK_CONFIG.get('max_risk_score', 100)
    
    for _, row in store_metrics.iterrows():
        mag = row['Mağaza Kodu']
        
        # Risk değerlerini al
        ic_hrs = ic_hirsizlik.get(mag, 0)
        kr_acik = kronik.get(mag, 0)
        kr_fire = kronik_fire.get(mag, 0)
        sig_acik = sigara_acik_series.get(mag, 0)
        fire_man = fire_manip.get(mag, 0)
        kasa_adet = kasa_agg.loc[mag, '10TL Adet'] if mag in kasa_agg.index else 0
        kasa_tutar = kasa_agg.loc[mag, '10TL Tutar'] if mag in kasa_agg.index else 0
        
        # Risk puanı hesapla (config'den ağırlıklar)
        risk_puan = 0
        risk_nedenler = []
        toplam_oran = row['Toplam %']
        
        # Toplam oran bazlı risk
        to = rw.get('toplam_oran', {})
        if toplam_oran > to.get('high', {}).get('threshold', 2):
            risk_puan += to.get('high', {}).get('points', 40)
            risk_nedenler.append(f"Toplam %{toplam_oran:.1f}")
        elif toplam_oran > to.get('medium', {}).get('threshold', 1.5):
            risk_puan += to.get('medium', {}).get('points', 25)
            risk_nedenler.append(f"Toplam %{toplam_oran:.1f}")
        elif toplam_oran > to.get('low', {}).get('threshold', 1):
            risk_puan += to.get('low', {}).get('points', 15)
        
        # İç hırsızlık
        ih = rw.get('ic_hirsizlik', {})
        if ic_hrs > ih.get('high', {}).get('threshold', 50):
            risk_puan += ih.get('high', {}).get('points', 30)
            risk_nedenler.append(f"İç hırs. {ic_hrs}")
        elif ic_hrs > ih.get('medium', {}).get('threshold', 30):
            risk_puan += ih.get('medium', {}).get('points', 20)
            risk_nedenler.append(f"İç hırs. {ic_hrs}")
        elif ic_hrs > ih.get('low', {}).get('threshold', 15):
            risk_puan += ih.get('low', {}).get('points', 10)
        
        # Sigara açığı
        sg = rw.get('sigara', {})
        if sig_acik > sg.get('high', {}).get('threshold', 5):
            risk_puan += sg.get('high', {}).get('points', 35)
            risk_nedenler.append(f"🚬 SİGARA {sig_acik:.0f}")
        elif sig_acik > sg.get('low', {}).get('threshold', 0):
            risk_puan += sg.get('low', {}).get('points', 20)
            risk_nedenler.append(f"🚬 Sigara {sig_acik:.0f}")
        
        # Kronik açık
        kr = rw.get('kronik', {})
        if kr_acik > kr.get('high', {}).get('threshold', 100):
            risk_puan += kr.get('high', {}).get('points', 15)
            risk_nedenler.append(f"Kronik {kr_acik}")
        elif kr_acik > kr.get('low', {}).get('threshold', 50):
            risk_puan += kr.get('low', {}).get('points', 10)
        
        # Fire manipülasyonu
        fm = rw.get('fire_manipulasyon', {})
        if fire_man > fm.get('high', {}).get('threshold', 10):
            risk_puan += fm.get('high', {}).get('points', 20)
            risk_nedenler.append(f"Fire man. {fire_man}")
        elif fire_man > fm.get('low', {}).get('threshold', 5):
            risk_puan += fm.get('low', {}).get('points', 10)
        
        # 10 TL ürünleri
        kt = rw.get('kasa_10tl', {})
        if kasa_adet > kt.get('high', {}).get('threshold', 20):
            risk_puan += kt.get('high', {}).get('points', 15)
            risk_nedenler.append(f"10TL +{kasa_adet:.0f}")
        elif kasa_adet > kt.get('low', {}).get('threshold', 10):
            risk_puan += kt.get('low', {}).get('points', 10)
        
        # Risk puanını sınırla
        risk_puan = min(risk_puan, max_score)
        
        # Risk seviyesi (config'den eşikler)
        if risk_puan >= rl.get('kritik', 60):
            risk_seviye = "🔴 KRİTİK"
        elif risk_puan >= rl.get('riskli', 40):
            risk_seviye = "🟠 RİSKLİ"
        elif risk_puan >= rl.get('dikkat', 20):
            risk_seviye = "🟡 DİKKAT"
        else:
            risk_seviye = "🟢 TEMİZ"
        
        results.append({
            'Mağaza Kodu': mag,
            'Mağaza Adı': row['Mağaza Adı'],
            'SM': row.get('Satış Müdürü', ''),
            'BS': row['Bölge Sorumlusu'],
            'Satış': row['Satış'],
            'Fark': row['Fark'],
            'Fire': row['Fire'],
            'Toplam Açık': row['Toplam Açık'],
            'Fark %': row['Fark %'],
            'Fire %': row['Fire %'],
            'Toplam %': row['Toplam %'],
            'Gün': row['Gün'],
            'Günlük Fark': row['Günlük Fark'],
            'Günlük Fire': row['Günlük Fire'],
            'İç Hırs.': ic_hrs,
            'Kr.Açık': kr_acik,
            'Kr.Fire': kr_fire,
            'Sigara': sig_acik,
            'Fire Man.': fire_man,
            '10TL Adet': kasa_adet,
            '10TL Tutar': kasa_tutar,
            'Risk Puan': risk_puan,
            'Risk': risk_seviye,
            'Risk Nedenleri': " | ".join(risk_nedenler) if risk_nedenler else "-"
        })
    
    result_df = pd.DataFrame(results)
    if len(result_df) > 0:
        result_df = result_df.sort_values('Risk Puan', ascending=False)
    
    return result_df


def aggregate_by_group(store_df, group_col):
    """SM veya BS bazında gruplama - Satış Ağırlıklı Ortalama Risk"""
    if group_col not in store_df.columns:
        return pd.DataFrame()
    
    # Kolon isimlerini kontrol et (VIEW vs analyze_region uyumu)
    kronik_col = 'Kronik' if 'Kronik' in store_df.columns else 'Kr.Açık'
    kasa_adet_col = 'Kasa Adet' if 'Kasa Adet' in store_df.columns else '10TL Adet'
    kasa_tutar_col = 'Kasa Tutar' if 'Kasa Tutar' in store_df.columns else '10TL Tutar'
    
    # Eksik kolonları 0 ile doldur
    if kronik_col not in store_df.columns:
        store_df[kronik_col] = 0
    if kasa_adet_col not in store_df.columns:
        store_df[kasa_adet_col] = 0
    if kasa_tutar_col not in store_df.columns:
        store_df[kasa_tutar_col] = 0
    if 'Gün' not in store_df.columns:
        store_df['Gün'] = 1
    
    # Temel metrikler
    agg_dict = {
        'Mağaza Kodu': 'count',
        'Satış': 'sum',
        'Fark': 'sum',
        'Fire': 'sum',
        'Toplam Açık': 'sum',
        'İç Hırs.': 'sum',
        kronik_col: 'sum',
        'Sigara': 'sum',
        kasa_adet_col: 'sum',
        kasa_tutar_col: 'sum',
        'Gün': 'sum',
    }
    
    grouped = store_df.groupby(group_col).agg(agg_dict).reset_index()
    
    grouped.columns = [group_col, 'Mağaza Sayısı', 'Satış', 'Fark', 'Fire', 'Toplam Açık',
                       'İç Hırs.', 'Kronik', 'Sigara', '10TL Adet', '10TL Tutar', 'Toplam Gün']
    
    # Satış Ağırlıklı Ortalama Risk Puanı hesapla
    for idx, row in grouped.iterrows():
        grup_magazalar = store_df[store_df[group_col] == row[group_col]]
        
        # Ağırlıklı ortalama
        toplam_agirlik = grup_magazalar['Satış'].sum()
        if toplam_agirlik > 0:
            agirlikli_risk = (grup_magazalar['Risk Puan'] * grup_magazalar['Satış']).sum() / toplam_agirlik
        else:
            agirlikli_risk = grup_magazalar['Risk Puan'].mean()
        
        grouped.at[idx, 'Risk Puan'] = agirlikli_risk
        
        # Kritik ve Riskli mağaza sayıları
        kritik_count = len(grup_magazalar[grup_magazalar['Risk'].str.contains('KRİTİK')])
        riskli_count = len(grup_magazalar[grup_magazalar['Risk'].str.contains('RİSKLİ')])
        grouped.at[idx, 'Kritik Mağaza'] = kritik_count
        grouped.at[idx, 'Riskli Mağaza'] = riskli_count
    
    # Oranlar
    grouped['Fark %'] = abs(grouped['Fark']) / grouped['Satış'] * 100
    grouped['Fark %'] = grouped['Fark %'].fillna(0)
    
    grouped['Fire %'] = abs(grouped['Fire']) / grouped['Satış'] * 100
    grouped['Fire %'] = grouped['Fire %'].fillna(0)
    
    grouped['Toplam %'] = abs(grouped['Toplam Açık']) / grouped['Satış'] * 100
    grouped['Toplam %'] = grouped['Toplam %'].fillna(0)
    
    # Günlük fark ve fire
    grouped['Günlük Fark'] = grouped['Fark'] / grouped['Toplam Gün']
    grouped['Günlük Fark'] = grouped['Günlük Fark'].fillna(0)
    grouped['Günlük Fire'] = grouped['Fire'] / grouped['Toplam Gün']
    grouped['Günlük Fire'] = grouped['Günlük Fire'].fillna(0)
    
    # Risk seviyesi (ağırlıklı ortalama risk puanına göre)
    def get_risk_level(puan):
        if puan >= 60:
            return "🔴 KRİTİK"
        elif puan >= 40:
            return "🟠 RİSKLİ"
        elif puan >= 20:
            return "🟡 DİKKAT"
        else:
            return "🟢 TEMİZ"
    
    grouped['Risk'] = grouped['Risk Puan'].apply(get_risk_level)
    
    # Risk puanına göre sırala (yüksekten düşüğe)
    grouped = grouped.sort_values('Risk Puan', ascending=False)
    
    return grouped


def create_gm_excel_report(store_df, sm_df, bs_df, params):
    """GM Dashboard Excel raporu"""
    
    wb = Workbook()
    
    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill('solid', fgColor='1F4E79')
    kritik_fill = PatternFill('solid', fgColor='FF4444')
    riskli_fill = PatternFill('solid', fgColor='FF8800')
    dikkat_fill = PatternFill('solid', fgColor='FFCC00')
    temiz_fill = PatternFill('solid', fgColor='00CC66')
    title_font = Font(bold=True, size=14)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    
    # ===== BÖLGE ÖZETİ =====
    ws = wb.active
    ws.title = "BÖLGE ÖZETİ"
    
    ws['A1'] = "GM BÖLGE DASHBOARD"
    ws['A1'].font = title_font
    ws['A2'] = f"Dönem: {params.get('donem', '')} | Mağaza: {len(store_df)}"
    
    # Toplamlar
    toplam_satis = store_df['Satış'].sum()
    toplam_fark = store_df['Fark'].sum()
    toplam_fire = store_df['Fire'].sum()
    toplam_acik = store_df['Toplam Açık'].sum()
    
    ws['A4'] = "GENEL METRİKLER"
    ws['A4'].font = Font(bold=True, size=11)
    
    ws['A5'] = "Toplam Satış"
    ws['B5'] = f"{toplam_satis:,.0f} TL"
    ws['A6'] = "Toplam Fark"
    ws['B6'] = f"{toplam_fark:,.0f} TL"
    ws['A7'] = "Toplam Fire"
    ws['B7'] = f"{toplam_fire:,.0f} TL"
    ws['A8'] = "Toplam Açık"
    ws['B8'] = f"{toplam_acik:,.0f} TL"
    ws['A9'] = "Kayıp Oranı"
    ws['B9'] = f"%{abs(toplam_acik)/toplam_satis*100:.2f}" if toplam_satis > 0 else "0%"
    
    # Risk dağılımı
    ws['A11'] = "RİSK DAĞILIMI"
    ws['A11'].font = Font(bold=True, size=11)
    
    kritik = len(store_df[store_df['Risk'].str.contains('KRİTİK')])
    riskli = len(store_df[store_df['Risk'].str.contains('RİSKLİ')])
    dikkat = len(store_df[store_df['Risk'].str.contains('DİKKAT')])
    temiz = len(store_df[store_df['Risk'].str.contains('TEMİZ')])
    
    ws['A12'] = "🔴 KRİTİK"
    ws['B12'] = kritik
    ws['A13'] = "🟠 RİSKLİ"
    ws['B13'] = riskli
    ws['A14'] = "🟡 DİKKAT"
    ws['B14'] = dikkat
    ws['A15'] = "🟢 TEMİZ"
    ws['B15'] = temiz
    
    # ===== SM BAZLI =====
    if len(sm_df) > 0:
        ws2 = wb.create_sheet("SM BAZLI")
        headers = ['Satış Müdürü', 'Mağaza', 'Satış', 'Fark', 'Fire', 'Toplam %', 'Sigara', 'İç Hırs.', 'Risk Puan', 'Risk']
        
        for col, header in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for row_idx, (_, row) in enumerate(sm_df.iterrows(), start=2):
            ws2.cell(row=row_idx, column=1, value=row['SM']).border = border
            ws2.cell(row=row_idx, column=2, value=row['Mağaza Sayısı']).border = border
            ws2.cell(row=row_idx, column=3, value=f"{row['Satış']:,.0f}").border = border
            ws2.cell(row=row_idx, column=4, value=f"{row['Fark']:,.0f}").border = border
            ws2.cell(row=row_idx, column=5, value=f"{row['Fire']:,.0f}").border = border
            ws2.cell(row=row_idx, column=6, value=f"%{row['Toplam %']:.1f}").border = border
            ws2.cell(row=row_idx, column=7, value=row['Sigara']).border = border
            ws2.cell(row=row_idx, column=8, value=row['İç Hırs.']).border = border
            ws2.cell(row=row_idx, column=9, value=f"{row['Risk Puan']:.0f}").border = border
            risk_cell = ws2.cell(row=row_idx, column=10, value=row['Risk'])
            risk_cell.border = border
    
    # ===== BS BAZLI =====
    if len(bs_df) > 0:
        ws3 = wb.create_sheet("BS BAZLI")
        headers = ['Bölge Sorumlusu', 'Mağaza', 'Satış', 'Fark', 'Fire', 'Toplam %', 'Sigara', 'İç Hırs.', 'Risk Puan', 'Risk']
        
        for col, header in enumerate(headers, 1):
            cell = ws3.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for row_idx, (_, row) in enumerate(bs_df.iterrows(), start=2):
            ws3.cell(row=row_idx, column=1, value=row['BS']).border = border
            ws3.cell(row=row_idx, column=2, value=row['Mağaza Sayısı']).border = border
            ws3.cell(row=row_idx, column=3, value=f"{row['Satış']:,.0f}").border = border
            ws3.cell(row=row_idx, column=4, value=f"{row['Fark']:,.0f}").border = border
            ws3.cell(row=row_idx, column=5, value=f"{row['Fire']:,.0f}").border = border
            ws3.cell(row=row_idx, column=6, value=f"%{row['Toplam %']:.1f}").border = border
            ws3.cell(row=row_idx, column=7, value=row['Sigara']).border = border
            ws3.cell(row=row_idx, column=8, value=row['İç Hırs.']).border = border
            ws3.cell(row=row_idx, column=9, value=f"{row['Risk Puan']:.0f}").border = border
            risk_cell = ws3.cell(row=row_idx, column=10, value=row['Risk'])
            risk_cell.border = border
    
    # ===== TÜM MAĞAZALAR =====
    ws4 = wb.create_sheet("TÜM MAĞAZALAR")
    headers = ['Mağaza Kodu', 'Mağaza Adı', 'SM', 'BS', 'Satış', 'Fark', 'Fire', 'Toplam %', 
               'Sigara', 'İç Hırs.', '10TL Adet', 'Risk Puan', 'Risk', 'Nedenler']
    
    for col, header in enumerate(headers, 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    for row_idx, (_, row) in enumerate(store_df.iterrows(), start=2):
        ws4.cell(row=row_idx, column=1, value=row['Mağaza Kodu']).border = border
        ws4.cell(row=row_idx, column=2, value=row['Mağaza Adı']).border = border
        ws4.cell(row=row_idx, column=3, value=row.get('SM', '')).border = border
        ws4.cell(row=row_idx, column=4, value=row['BS']).border = border
        ws4.cell(row=row_idx, column=5, value=f"{row['Satış']:,.0f}").border = border
        ws4.cell(row=row_idx, column=6, value=f"{row['Fark']:,.0f}").border = border
        ws4.cell(row=row_idx, column=7, value=f"{row['Fire']:,.0f}").border = border
        ws4.cell(row=row_idx, column=8, value=f"%{row['Toplam %']:.1f}").border = border
        ws4.cell(row=row_idx, column=9, value=row['Sigara']).border = border
        ws4.cell(row=row_idx, column=10, value=row['İç Hırs.']).border = border
        # 10TL Adet - VIEW ve analyze_region uyumu
        kasa_adet = row.get('Kasa Adet', row.get('10TL Adet', 0))
        ws4.cell(row=row_idx, column=11, value=kasa_adet).border = border
        ws4.cell(row=row_idx, column=12, value=f"{row['Risk Puan']:.0f}").border = border
        
        risk_cell = ws4.cell(row=row_idx, column=13, value=row['Risk'])
        risk_cell.border = border
        if 'KRİTİK' in str(row['Risk']):
            risk_cell.fill = kritik_fill
        elif 'RİSKLİ' in str(row['Risk']):
            risk_cell.fill = riskli_fill
        elif 'DİKKAT' in str(row['Risk']):
            risk_cell.fill = dikkat_fill
        else:
            risk_cell.fill = temiz_fill
        
        ws4.cell(row=row_idx, column=14, value=row.get('Risk Nedenleri', '')).border = border
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def create_region_excel_report(region_df, df_all, kasa_kodlari, params):
    """Bölge özet Excel raporu"""
    
    wb = Workbook()
    
    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill('solid', fgColor='1F4E79')
    kritik_fill = PatternFill('solid', fgColor='FF4444')
    riskli_fill = PatternFill('solid', fgColor='FF8800')
    dikkat_fill = PatternFill('solid', fgColor='FFCC00')
    temiz_fill = PatternFill('solid', fgColor='00CC66')
    title_font = Font(bold=True, size=14)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    
    # ===== BÖLGE ÖZETİ =====
    ws = wb.active
    ws.title = "BÖLGE ÖZETİ"
    
    ws['A1'] = f"BÖLGE ENVANTER ANALİZİ"
    ws['A1'].font = title_font
    ws['A2'] = f"Dönem: {params.get('donem', '')} | Tarih: {params.get('tarih', '')} | Mağaza Sayısı: {len(region_df)}"
    
    # Bölge toplamları
    ws['A4'] = "BÖLGE TOPLAMI"
    ws['A4'].font = Font(bold=True, size=11)
    
    toplam_satis = region_df['Satış'].sum()
    toplam_fark = region_df['Fark'].sum()
    toplam_fire = region_df['Fire'].sum()
    # Kayıp Oranı = |Fark + Fire| / Satış × 100
    genel_oran = abs(toplam_fark + toplam_fire) / toplam_satis * 100 if toplam_satis > 0 else 0
    
    ws['A5'] = "Toplam Satış"
    ws['B5'] = f"{toplam_satis:,.0f} TL"
    ws['A6'] = "Toplam Fark"
    ws['B6'] = f"{toplam_fark:,.0f} TL"
    ws['A7'] = "Toplam Fire"
    ws['B7'] = f"{toplam_fire:,.0f} TL"
    ws['A8'] = "Genel Kayıp Oranı"
    ws['B8'] = f"%{genel_oran:.2f}"
    
    # Risk dağılımı
    ws['A10'] = "RİSK DAĞILIMI"
    ws['A10'].font = Font(bold=True, size=11)
    
    kritik_sayisi = len(region_df[region_df['Risk'].str.contains('KRİTİK')])
    riskli_sayisi = len(region_df[region_df['Risk'].str.contains('RİSKLİ')])
    dikkat_sayisi = len(region_df[region_df['Risk'].str.contains('DİKKAT')])
    temiz_sayisi = len(region_df[region_df['Risk'].str.contains('TEMİZ')])
    
    ws['A11'] = "🔴 KRİTİK"
    ws['B11'] = kritik_sayisi
    ws['A12'] = "🟠 RİSKLİ"
    ws['B12'] = riskli_sayisi
    ws['A13'] = "🟡 DİKKAT"
    ws['B13'] = dikkat_sayisi
    ws['A14'] = "🟢 TEMİZ"
    ws['B14'] = temiz_sayisi
    
    # Mağaza sıralaması
    ws['A16'] = "MAĞAZA SIRALAMASI (Risk Puanına Göre)"
    ws['A16'].font = Font(bold=True, size=11)
    
    headers = ['Mağaza', 'Adı', 'Satış', 'Fark', 'Toplam %', 'İç Hırs.', 'Sigara', 'Kr.Açık', 'Risk', 'Neden']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=17, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    for row_idx, (_, row) in enumerate(region_df.iterrows(), start=18):
        ws.cell(row=row_idx, column=1, value=row['Mağaza Kodu']).border = border
        ws.cell(row=row_idx, column=2, value=row['Mağaza Adı'][:25]).border = border
        ws.cell(row=row_idx, column=3, value=f"{row['Satış']:,.0f}").border = border
        ws.cell(row=row_idx, column=4, value=f"{row['Fark']:,.0f}").border = border
        ws.cell(row=row_idx, column=5, value=f"%{row['Toplam %']:.1f}").border = border
        ws.cell(row=row_idx, column=6, value=row['İç Hırs.']).border = border
        ws.cell(row=row_idx, column=7, value=row['Sigara']).border = border
        ws.cell(row=row_idx, column=8, value=row['Kr.Açık']).border = border
        
        risk_cell = ws.cell(row=row_idx, column=9, value=row['Risk'])
        risk_cell.border = border
        if 'KRİTİK' in row['Risk']:
            risk_cell.fill = kritik_fill
            risk_cell.font = Font(bold=True, color='FFFFFF')
        elif 'RİSKLİ' in row['Risk']:
            risk_cell.fill = riskli_fill
            risk_cell.font = Font(bold=True, color='FFFFFF')
        elif 'DİKKAT' in row['Risk']:
            risk_cell.fill = dikkat_fill
            risk_cell.font = Font(bold=True)
        else:
            risk_cell.fill = temiz_fill
            risk_cell.font = Font(bold=True, color='FFFFFF')
        
        ws.cell(row=row_idx, column=10, value=row['Risk Nedenleri']).border = border
    
    # Sütun genişlikleri
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 35
    
    # ===== DETAY SHEET =====
    ws2 = wb.create_sheet("DETAY")
    
    detail_headers = ['Mağaza Kodu', 'Mağaza Adı', 'Satış', 'Fark', 'Fire', 'Toplam %', 
                      'İç Hırs.', 'Kr.Açık', 'Kr.Fire', 'Sigara', 'Fire Man.', 
                      '10TL Adet', '10TL Tutar', 'Risk Puan', 'Risk', 'Risk Nedenleri']
    
    for col, header in enumerate(detail_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    for row_idx, (_, row) in enumerate(region_df.iterrows(), start=2):
        ws2.cell(row=row_idx, column=1, value=row['Mağaza Kodu']).border = border
        ws2.cell(row=row_idx, column=2, value=row['Mağaza Adı']).border = border
        ws2.cell(row=row_idx, column=3, value=row['Satış']).border = border
        ws2.cell(row=row_idx, column=4, value=row['Fark']).border = border
        ws2.cell(row=row_idx, column=5, value=row['Fire']).border = border
        ws2.cell(row=row_idx, column=6, value=row['Toplam %']).border = border
        ws2.cell(row=row_idx, column=7, value=row['İç Hırs.']).border = border
        ws2.cell(row=row_idx, column=8, value=row['Kr.Açık']).border = border
        ws2.cell(row=row_idx, column=9, value=row['Kr.Fire']).border = border
        ws2.cell(row=row_idx, column=10, value=row['Sigara']).border = border
        ws2.cell(row=row_idx, column=11, value=row['Fire Man.']).border = border
        ws2.cell(row=row_idx, column=12, value=row['10TL Adet']).border = border
        ws2.cell(row=row_idx, column=13, value=row['10TL Tutar']).border = border
        ws2.cell(row=row_idx, column=14, value=row['Risk Puan']).border = border
        ws2.cell(row=row_idx, column=15, value=row['Risk']).border = border
        ws2.cell(row=row_idx, column=16, value=row['Risk Nedenleri']).border = border
    
    auto_adjust_column_width(ws2)
    
    # Excel çıktısı
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()


def calculate_store_risk(df, internal_df, chronic_df, cigarette_df):
    """Mağaza risk seviyesi"""
    toplam_satis = df['Satış Tutarı'].sum()
    fark_tutari = df['Fark Tutarı'].fillna(0).sum()
    fire_tutari = df['Fire Tutarı'].fillna(0).sum()
    kismi_tutari = df['Kısmi Envanter Tutarı'].fillna(0).sum()
    
    # Kayıp Oranı = |Fark + Fire + Kısmi| / Satış × 100
    kayip = fark_tutari + fire_tutari + kismi_tutari
    kayip_orani = abs(kayip) / toplam_satis * 100 if toplam_satis > 0 else 0
    ic_hirsizlik = len(internal_df)
    
    # Sigara açığı - toplam bazlı
    sigara_acik = 0
    if len(cigarette_df) > 0 and 'Ürün Toplam' in cigarette_df.columns:
        son_satir = cigarette_df.iloc[-1]
        if son_satir['Malzeme Kodu'] == '*** TOPLAM ***':
            sigara_acik = abs(son_satir['Ürün Toplam'])
    
    if kayip_orani > 2 or ic_hirsizlik > 50 or sigara_acik > 5:
        return "KRİTİK", "risk-kritik"
    elif kayip_orani > 1.5 or ic_hirsizlik > 30 or sigara_acik > 3:
        return "RİSKLİ", "risk-riskli"
    elif kayip_orani > 1 or ic_hirsizlik > 15 or sigara_acik > 0:
        return "DİKKAT", "risk-dikkat"
    else:
        return "TEMİZ", "risk-temiz"


def create_top_20_risky(df, internal_codes, chronic_codes, family_balanced_codes):
    """En riskli 20 ürün"""
    
    # Dengelenmişleri ve aile dengelenmişlerini çıkar
    risky_df = df[
        (df['NET_ENVANTER_ETKİ_TUTARI'] < 0) & 
        (~df.apply(is_balanced, axis=1)) &
        (~df['Malzeme Kodu'].astype(str).isin(family_balanced_codes))
    ].copy()
    
    if len(risky_df) == 0:
        return pd.DataFrame()
    
    # DUPLICATE TEMİZLEME - önce yap
    risky_df = risky_df.drop_duplicates(subset=['Malzeme Kodu'], keep='first')
    
    def classify(row):
        kod = str(row.get('Malzeme Kodu', ''))
        
        if kod in internal_codes:
            return "İÇ HIRSIZLIK", "Kasa kamera incelemesi"
        elif kod in chronic_codes:
            return "KRONİK AÇIK", "Raf kontrolü, Sayım eğitimi"
        elif row['Fire Miktarı'] < 0:
            return "OPERASYONEL", "Fire kayıt kontrolü"
        else:
            return "DIŞ HIRSIZLIK/SAYIM", "Sayım ve kod kontrolü"
    
    risky_df['Risk Türü'] = risky_df.apply(lambda x: classify(x)[0], axis=1)
    risky_df['Aksiyon'] = risky_df.apply(lambda x: classify(x)[1], axis=1)
    
    risky_df = risky_df.sort_values('NET_ENVANTER_ETKİ_TUTARI', ascending=True).head(20)
    
    result = pd.DataFrame({
        'Sıra': range(1, len(risky_df) + 1),
        'Malzeme Kodu': risky_df['Malzeme Kodu'].values,
        'Malzeme Adı': risky_df['Malzeme Adı'].values,
        'Fark Mik.': risky_df['Fark Miktarı'].values,
        'Kısmi': risky_df['Kısmi Envanter Miktarı'].values,
        'Önceki': risky_df['Önceki Fark Miktarı'].values,
        'TOPLAM': risky_df['TOPLAM_MIKTAR'].values,
        'İptal': risky_df['İptal Satır Miktarı'].values,
        'Fire': risky_df['Fire Miktarı'].values,
        'Fire Tutarı': risky_df['Fire Tutarı'].values,
        'Fark Tutarı': risky_df['Fark Tutarı'].values,
        'Risk Türü': risky_df['Risk Türü'].values,
        'Aksiyon': risky_df['Aksiyon'].values
    })
    
    return result


def auto_adjust_column_width(ws):
    """Excel sütun genişliklerini otomatik ayarla"""
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        
        for cell in column_cells:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width


def create_excel_report(df, internal_df, chronic_df, chronic_fire_df, cigarette_df, 
                       external_df, family_df, fire_manip_df, kasa_activity_df, top20_df, 
                       exec_comments, group_stats, magaza_kodu, magaza_adi, params):
    """Excel raporu - tüm sheet'ler dahil"""
    
    wb = Workbook()
    
    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill('solid', fgColor='1F4E79')
    title_font = Font(bold=True, size=14)
    subtitle_font = Font(bold=True, size=11)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    
    # ===== ÖZET =====
    ws = wb.active
    ws.title = "ÖZET"
    
    ws['A1'] = f"MAĞAZA: {magaza_kodu} - {magaza_adi}"
    ws['A1'].font = title_font
    ws['A2'] = f"Dönem: {params.get('donem', '')} | Tarih: {params.get('tarih', '')}"
    
    ws['A4'] = "GENEL METRIKLER"
    ws['A4'].font = subtitle_font
    
    toplam_satis = df['Satış Tutarı'].sum()
    fark_tutari = df['Fark Tutarı'].fillna(0).sum()
    kismi_tutari = df['Kısmi Envanter Tutarı'].fillna(0).sum()
    fire_tutari = df['Fire Tutarı'].fillna(0).sum()
    
    # Fark = Fark Tutarı + Kısmi
    fark = fark_tutari + kismi_tutari
    # Toplam Açık = Fark + Fire
    toplam_acik = fark + fire_tutari
    
    # Oranlar
    fark_oran = abs(fark) / toplam_satis * 100 if toplam_satis > 0 else 0
    fire_oran = abs(fire_tutari) / toplam_satis * 100 if toplam_satis > 0 else 0
    toplam_oran = abs(toplam_acik) / toplam_satis * 100 if toplam_satis > 0 else 0
    
    metrics = [
        ('Toplam Ürün', len(df)),
        ('Açık Veren Ürün', len(df[df['Fark Miktarı'] < 0])),
        ('Toplam Satış', f"{toplam_satis:,.0f} TL"),
        ('Fark (Fark+Kısmi)', f"{fark:,.0f} TL"),
        ('Fire', f"{fire_tutari:,.0f} TL"),
        ('Toplam Açık', f"{toplam_acik:,.0f} TL"),
        ('Fark Oranı', f"%{fark_oran:.2f}"),
        ('Fire Oranı', f"%{fire_oran:.2f}"),
        ('Toplam Oran', f"%{toplam_oran:.2f}"),
    ]
    
    for i, (label, value) in enumerate(metrics, start=5):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
    
    ws['A15'] = "RİSK DAĞILIMI"
    ws['A15'].font = subtitle_font
    
    # Sigara açığı NET toplamı hesapla (satır sayısı değil!)
    sigara_net_toplam = 0
    if len(cigarette_df) > 0:
        toplam_row = cigarette_df[cigarette_df['Malzeme Kodu'] == '*** TOPLAM ***']
        if len(toplam_row) > 0:
            sigara_net_toplam = abs(toplam_row['Ürün Toplam'].values[0])
        else:
            # Toplam satırı yoksa manuel hesapla
            sigara_net_toplam = abs(cigarette_df['Ürün Toplam'].sum())
    
    risks = [
        ('İç Hırsızlık (≥100TL)', len(internal_df)),
        ('Kronik Açık', len(chronic_df)),
        ('Kronik Fire', len(chronic_fire_df)),
        ('Sigara Açığı', int(sigara_net_toplam)),  # NET TOPLAM, satır sayısı değil!
        ('Fire Manipülasyonu', len(fire_manip_df)),
    ]
    
    for i, (label, value) in enumerate(risks, start=16):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        if 'Sigara' in label and value > 0:
            ws[f'B{i}'].fill = PatternFill('solid', fgColor='FF4444')
            ws[f'B{i}'].font = Font(bold=True, color='FFFFFF')
    
    ws['A22'] = "YÖNETİCİ ÖZETİ"
    ws['A22'].font = subtitle_font
    
    for i, comment in enumerate(exec_comments[:10], start=23):
        ws[f'A{i}'] = comment
    
    auto_adjust_column_width(ws)
    
    # ===== EN RİSKLİ 20 =====
    if len(top20_df) > 0:
        ws2 = wb.create_sheet("EN RİSKLİ 20")
        for col, h in enumerate(top20_df.columns, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for r_idx, row in enumerate(top20_df.values, 2):
            for c_idx, val in enumerate(row, 1):
                cell = ws2.cell(row=r_idx, column=c_idx, value=val)
                cell.border = border
                cell.alignment = wrap_alignment
        
        auto_adjust_column_width(ws2)
    
    # ===== KRONİK AÇIK =====
    if len(chronic_df) > 0:
        ws3 = wb.create_sheet("KRONİK AÇIK")
        for col, h in enumerate(chronic_df.columns, 1):
            cell = ws3.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        
        for r_idx, row in enumerate(chronic_df.head(100).values, 2):
            for c_idx, val in enumerate(row, 1):
                ws3.cell(row=r_idx, column=c_idx, value=val)
        
        auto_adjust_column_width(ws3)
    
    # ===== KRONİK FİRE =====
    if len(chronic_fire_df) > 0:
        ws4 = wb.create_sheet("KRONİK FİRE")
        for col, h in enumerate(chronic_fire_df.columns, 1):
            cell = ws4.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        
        for r_idx, row in enumerate(chronic_fire_df.head(100).values, 2):
            for c_idx, val in enumerate(row, 1):
                ws4.cell(row=r_idx, column=c_idx, value=val)
        
        auto_adjust_column_width(ws4)
    
    # ===== SİGARA AÇIĞI =====
    ws5 = wb.create_sheet("SİGARA AÇIĞI")
    ws5['A1'] = "⚠️ SİGARA AÇIĞI - YÜKSEK RİSK"
    ws5['A1'].font = Font(bold=True, size=14, color='FF0000')
    
    if len(cigarette_df) > 0:
        for col, h in enumerate(cigarette_df.columns, 1):
            cell = ws5.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = PatternFill('solid', fgColor='FF4444')
        
        for r_idx, row in enumerate(cigarette_df.values, 4):
            for c_idx, val in enumerate(row, 1):
                ws5.cell(row=r_idx, column=c_idx, value=val)
        
        auto_adjust_column_width(ws5)
    
    # ===== İÇ HIRSIZLIK =====
    if len(internal_df) > 0:
        ws6 = wb.create_sheet("İÇ HIRSIZLIK")
        ws6['A1'] = "Satış Fiyatı ≥ 100 TL | Fark büyüdükçe risk AZALIR"
        ws6['A1'].font = subtitle_font
        
        for col, h in enumerate(internal_df.columns, 1):
            cell = ws6.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        
        for r_idx, row in enumerate(internal_df.head(100).values, 4):
            for c_idx, val in enumerate(row, 1):
                ws6.cell(row=r_idx, column=c_idx, value=val)
        
        auto_adjust_column_width(ws6)
    
    # ===== AİLE ANALİZİ =====
    if len(family_df) > 0:
        ws7 = wb.create_sheet("AİLE ANALİZİ")
        ws7['A1'] = "Benzer Ürün Ailesi - Kod Karışıklığı Tespiti"
        ws7['A1'].font = subtitle_font
        
        for col, h in enumerate(family_df.columns, 1):
            cell = ws7.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        
        for r_idx, row in enumerate(family_df.head(100).values, 4):
            for c_idx, val in enumerate(row, 1):
                cell = ws7.cell(row=r_idx, column=c_idx, value=val)
                cell.alignment = wrap_alignment
        
        auto_adjust_column_width(ws7)
    
    # ===== FİRE MANİPÜLASYONU =====
    if len(fire_manip_df) > 0:
        ws8 = wb.create_sheet("FİRE MANİPÜLASYONU")
        for col, h in enumerate(fire_manip_df.columns, 1):
            cell = ws8.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        
        for r_idx, row in enumerate(fire_manip_df.head(100).values, 2):
            for c_idx, val in enumerate(row, 1):
                ws8.cell(row=r_idx, column=c_idx, value=val)
        
        auto_adjust_column_width(ws8)
    
    # ===== KASA AKTİVİTESİ =====
    if len(kasa_activity_df) > 0:
        ws9 = wb.create_sheet("KASA AKTİVİTESİ")
        ws9['A1'] = "⚠️ KASA AKTİVİTESİ ÜRÜNLERİ - FAZLA (+) OLANLAR MANİPÜLASYON RİSKİ!"
        ws9['A1'].font = Font(bold=True, size=12, color='FF0000')
        
        for col, h in enumerate(kasa_activity_df.columns, 1):
            cell = ws9.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        
        for r_idx, row in enumerate(kasa_activity_df.values, 4):
            for c_idx, val in enumerate(row, 1):
                cell = ws9.cell(row=r_idx, column=c_idx, value=val)
                # Fazla olanları kırmızı yap
                if c_idx == 6 and isinstance(val, (int, float)) and val > 0:  # TOPLAM sütunu
                    cell.fill = PatternFill('solid', fgColor='FFCCCC')
        
        auto_adjust_column_width(ws9)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ===== ANA UYGULAMA =====

# SM Özet modu - session_state'den filtrele
if analysis_mode == "👔 SM Özet":
    st.subheader("👔 SM Özet")
    
    # ⚡ EKRAN İÇİN HAFİF VERİ - load_all_data_once ÇAĞRILMIYOR!
    # Excel butonu basılınca tam veri çekilecek
    
    # Kullanıcı -> SM eşleştirmesi
    USER_SM_MAPPING = {
        "sm1": "ALİ AKÇAY",
        "sm2": "ŞADAN YURDAKUL",
        "sm3": "VELİ GÖK",
        "sm4": "GİZEM TOSUN",
        "sma": None,  # Asistan - SM seçecek
        "ziya": None,  # GM - tüm SM'leri görebilir
    }
    
    current_user = st.session_state.user
    user_sm = USER_SM_MAPPING.get(current_user)
    is_gm = current_user == "ziya"
    
    # SM ve Dönem seçimi - aynı satırda
    col_sm, col_donem = st.columns([1, 1])
    
    # Cache'den al (hızlı)
    available_sms = get_available_sms_cached()
    available_periods = get_available_periods_cached()
    
    with col_sm:
        if is_gm:
            # GM tüm SM'leri görebilir + TÜMÜ seçeneği
            if available_sms:
                sm_options = ["📊 TÜMÜ (Bölge)"] + available_sms
                selected_sm_option = st.selectbox("👔 Satış Müdürü", sm_options)
                
                if selected_sm_option == "📊 TÜMÜ (Bölge)":
                    selected_sm = None
                    display_sm = "Bölge"
                else:
                    selected_sm = selected_sm_option
                    display_sm = selected_sm
            else:
                st.warning("Henüz veri yüklenmemiş")
                selected_sm = None
                selected_sm_option = None
                display_sm = None
        elif user_sm:
            # SM kendi verilerini görür (sadece kendi ismi gösterilir)
            selected_sm = user_sm
            selected_sm_option = user_sm
            display_sm = user_sm
            st.selectbox("👔 Satış Müdürü", [user_sm], disabled=True)
        else:
            # Asistan veya tanımsız - SM seçebilir
            if available_sms:
                selected_sm = st.selectbox("👔 Satış Müdürü", available_sms)
                selected_sm_option = selected_sm
                display_sm = selected_sm
            else:
                st.warning("Henüz veri yüklenmemiş")
                selected_sm = None
                selected_sm_option = None
                display_sm = None
    
    with col_donem:
        if available_periods:
            selected_periods = st.multiselect("📅 Dönem", available_periods, default=available_periods[:1])
        else:
            selected_periods = []
    
    # Tarih aralığı filtresi (opsiyonel)
    tarih_baslangic = None
    tarih_bitis = None
    
    if selected_periods:
        # Seçilen dönemlerdeki envanter tarihlerini getir
        donem_tarihleri = get_envanter_tarihleri_by_donem(tuple(selected_periods))
        
        if donem_tarihleri and len(donem_tarihleri) > 1:
            with st.expander("📆 Tarih Aralığı Filtresi (Opsiyonel)", expanded=False):
                col_t1, col_t2 = st.columns(2)
                with col_t1:
                    min_tarih = min(donem_tarihleri)
                    max_tarih = max(donem_tarihleri)
                    tarih_baslangic = st.date_input(
                        "Başlangıç Tarihi", 
                        value=min_tarih,
                        min_value=min_tarih,
                        max_value=max_tarih,
                        key="sm_tarih_bas"
                    )
                with col_t2:
                    tarih_bitis = st.date_input(
                        "Bitiş Tarihi", 
                        value=max_tarih,
                        min_value=min_tarih,
                        max_value=max_tarih,
                        key="sm_tarih_bit"
                    )
                
                # Eğer varsayılan değerler seçiliyse filtre uygulanmasın
                if tarih_baslangic == min_tarih and tarih_bitis == max_tarih:
                    tarih_baslangic = None
                    tarih_bitis = None
                else:
                    st.info(f"📆 Filtre: {tarih_baslangic.strftime('%d.%m.%Y')} - {tarih_bitis.strftime('%d.%m.%Y')}")
    
    if selected_sm_option and selected_periods:
        # ⚡ SÜPER HIZLI - Supabase VIEW'den direkt özet veri
        region_df = get_sm_summary_from_view(
            satis_muduru=selected_sm, 
            donemler=selected_periods,
            tarih_baslangic=tarih_baslangic,
            tarih_bitis=tarih_bitis
        )
        
        if len(region_df) == 0:
            st.warning("Seçilen kriterlere uygun veri bulunamadı")
        else:
            # Mağaza bilgisi
            magazalar = region_df['Mağaza Kodu'].dropna().unique().tolist()
            magaza_isimleri = {}
            for mag in magazalar:
                isim = region_df[region_df['Mağaza Kodu'] == mag]['Mağaza Adı'].iloc[0] if 'Mağaza Adı' in region_df.columns else ''
                magaza_isimleri[mag] = f"{mag} - {isim}" if isim else str(mag)
            
            params = {
                'donem': ', '.join(selected_periods),
                'tarih': datetime.now().strftime('%Y-%m-%d'),
            }
            
            # Kasa aktivitesi kodlarını yükle
            kasa_kodlari = load_kasa_activity_codes()
            
            st.subheader(f"📊 {display_sm} - {len(magazalar)} Mağaza")
            
            # ⚡ Risk puanına göre sırala (yüksekten düşüğe)
            region_df = region_df.sort_values('Risk Puan', ascending=False)
            
            if len(region_df) == 0:
                st.warning("Analiz edilecek mağaza bulunamadı!")
            else:
                # Bölge toplamları
                toplam_satis = region_df['Satış'].sum()
                toplam_fark = region_df['Fark'].sum()
                toplam_fire = region_df['Fire'].sum()
                toplam_acik = region_df['Toplam Açık'].sum()
                toplam_gun = region_df['Gün'].sum()
                
                # Oranlar
                fark_oran = abs(toplam_fark) / toplam_satis * 100 if toplam_satis > 0 else 0
                fire_oran = abs(toplam_fire) / toplam_satis * 100 if toplam_satis > 0 else 0
                toplam_oran = abs(toplam_acik) / toplam_satis * 100 if toplam_satis > 0 else 0
                gunluk_fark = toplam_fark / toplam_gun if toplam_gun > 0 else 0
                gunluk_fire = toplam_fire / toplam_gun if toplam_gun > 0 else 0
                
                # Risk dağılımı
                kritik_sayisi = len(region_df[region_df['Risk'] == '🔴 KRİTİK'])
                riskli_sayisi = len(region_df[region_df['Risk'] == '🟠 RİSKLİ'])
                dikkat_sayisi = len(region_df[region_df['Risk'] == '🟡 DİKKAT'])
                temiz_sayisi = len(region_df[region_df['Risk'] == '🟢 TEMİZ'])
                
                # Üst metrikler
                st.markdown("### 💰 Özet Metrikler")
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("💰 Toplam Satış", f"{toplam_satis/1_000_000:.1f}M TL")
                with col2:
                    st.metric("📉 Fark", f"{toplam_fark/1000:.0f}K TL", f"%{fark_oran:.2f} | Günlük: {gunluk_fark/1000:.1f}K")
                with col3:
                    st.metric("🔥 Fire", f"{toplam_fire/1000:.0f}K TL", f"%{fire_oran:.2f} | Günlük: {gunluk_fire/1000:.1f}K")
                with col4:
                    st.metric("📊 Toplam", f"{toplam_acik/1000:.0f}K TL", f"%{toplam_oran:.2f}")
                
                # Risk dağılımı
                st.markdown("### 📊 Risk Dağılımı")
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    if kritik_sayisi > 0:
                        st.markdown(f'<div class="risk-kritik">🔴 KRİTİK: {kritik_sayisi}</div>', unsafe_allow_html=True)
                    else:
                        st.metric("🔴 KRİTİK", kritik_sayisi)
                with col2:
                    if riskli_sayisi > 0:
                        st.markdown(f'<div class="risk-riskli">🟠 RİSKLİ: {riskli_sayisi}</div>', unsafe_allow_html=True)
                    else:
                        st.metric("🟠 RİSKLİ", riskli_sayisi)
                with col3:
                    if dikkat_sayisi > 0:
                        st.markdown(f'<div class="risk-dikkat">🟡 DİKKAT: {dikkat_sayisi}</div>', unsafe_allow_html=True)
                    else:
                        st.metric("🟡 DİKKAT", dikkat_sayisi)
                with col4:
                    st.markdown(f'<div class="risk-temiz">🟢 TEMİZ: {temiz_sayisi}</div>', unsafe_allow_html=True)
                
                # BS Özeti
                st.markdown("### 👔 BS Özeti")
                
                bs_ozet = region_df.groupby('BS').agg({
                    'Mağaza Kodu': 'count',
                    'Satış': 'sum',
                    'Fark': 'sum',
                    'Fire': 'sum',
                    'Toplam Açık': 'sum',
                    'Risk Puan': 'sum',
                    'Sigara': 'sum',
                    'İç Hırs.': 'sum',
                    'Kasa Tutar': 'sum'  # 10TL ürünleri
                }).reset_index()
                
                bs_ozet = bs_ozet.rename(columns={
                    'Mağaza Kodu': 'Mağaza',
                    'Toplam Açık': 'Toplam'
                })
                
                bs_ozet['Kayıp %'] = abs(bs_ozet['Toplam']) / bs_ozet['Satış'] * 100
                bs_ozet = bs_ozet.sort_values('Risk Puan', ascending=False)
                
                # BS tablosu - tam rakamlar ve risk puanı ile
                for _, bs_row in bs_ozet.iterrows():
                    col1, col2, col3, col4, col5, col6 = st.columns([2.5, 1.5, 1.5, 1, 1.2, 1])
                    col1.write(f"**{bs_row['BS']}** ({bs_row['Mağaza']:.0f} mağ.)")
                    col2.write(f"Satış: {bs_row['Satış']/1e6:.1f}M | Fark: {bs_row['Fark']:,.0f}")
                    col3.write(f"Fire: {bs_row['Fire']:,.0f}")
                    col4.write(f"Kayıp: %{bs_row['Kayıp %']:.1f}")
                    col5.write(f"🚬{bs_row['Sigara']:.0f} 🔒{bs_row['İç Hırs.']:.0f} 💰{bs_row['Kasa Tutar']:,.0f}")
                    col6.write(f"**Risk: {bs_row['Risk Puan']:.0f}**")
                
                # Sekmeler - Bölge Özeti ile aynı
                st.markdown("---")
                tabs = st.tabs(["📋 Sıralama", "🔴 Kritik", "🟠 Riskli", "🚬 Sigara", "🔍 Mağaza Detay", "📥 İndir"])
                
                with tabs[0]:
                    st.subheader("📋 Mağaza Sıralaması (Risk Puanına Göre)")
                    
                    # Basit tablo göster - 10TL açığı dahil
                    display_cols = ['Mağaza Kodu', 'Mağaza Adı', 'BS', 'Satış', 'Fark', 'Fire', 
                                   'Toplam %', 'Sigara', 'İç Hırs.', 'Kasa Tutar', 'Risk Puan', 'Risk']
                    
                    # Formatla
                    display_df = region_df[display_cols].copy()
                    display_df['Mağaza Kodu'] = display_df['Mağaza Kodu'].astype(str)  # Dar kolon için
                    display_df['Satış'] = display_df['Satış'].apply(lambda x: f"{x/1000:.0f}K")
                    display_df['Fark'] = display_df['Fark'].apply(lambda x: f"{x/1000:.0f}K")
                    display_df['Fire'] = display_df['Fire'].apply(lambda x: f"{x/1000:.0f}K")
                    display_df['Toplam %'] = display_df['Toplam %'].apply(lambda x: f"%{x:.1f}")
                    display_df['Kasa Tutar'] = display_df['Kasa Tutar'].apply(lambda x: f"{x:,.0f}")
                    display_df['Risk Puan'] = display_df['Risk Puan'].apply(lambda x: f"{x:.0f}")
                    
                    # Kolon isimlerini kısalt
                    display_df = display_df.rename(columns={
                        'Mağaza Kodu': 'Kod',
                        'Mağaza Adı': 'Mağaza',
                        'Kasa Tutar': '10TL'
                    })
                    
                    st.dataframe(display_df, use_container_width=True, hide_index=True, height=500)
                    
                    # Tek mağaza raporu için seçim
                    st.markdown("---")
                    st.markdown("**📥 Mağaza Raporu İndir**")
                    
                    mag_options = [f"{row['Mağaza Kodu']} - {row['Mağaza Adı']}" for _, row in region_df.iterrows()]
                    selected_mag_option = st.selectbox("Mağaza seçin", mag_options, key="sm_mag_select")
                    
                    if st.button("📥 Rapor Oluştur", key="sm_create_report"):
                        selected_mag_kod = selected_mag_option.split(" - ")[0]
                        selected_row = region_df[region_df['Mağaza Kodu'] == selected_mag_kod].iloc[0]
                        
                        with st.spinner("📊 Mağaza verisi yükleniyor (bu işlem 5-10 saniye sürebilir)..."):
                            # ⚡ HIZLI - Sadece bu mağaza için veri çek
                            df_mag = get_single_store_data(selected_mag_kod, tuple(selected_periods) if selected_periods else None)
                            
                            if len(df_mag) > 0:
                                df_mag = analyze_inventory(df_mag)
                                mag_adi = selected_row['Mağaza Adı']
                                
                                # Analizleri yap
                                int_df = detect_internal_theft(df_mag)
                                
                                # Kamera timestamp entegrasyonu (kategori araması için full_df geçir)
                                if len(int_df) > 0:
                                    try:
                                        env_tarihi = df_mag['Envanter Tarihi'].iloc[0]
                                        int_df = enrich_internal_theft_with_camera(int_df, selected_mag_kod, env_tarihi, full_df=df_mag)
                                    except:
                                        pass
                                
                                chr_df = detect_chronic_products(df_mag)
                                chr_fire_df = detect_chronic_fire(df_mag)
                                cig_df = detect_cigarette_shortage(df_mag)
                                ext_df = detect_external_theft(df_mag)
                                fam_df = find_product_families(df_mag)
                                fire_df = detect_fire_manipulation(df_mag)
                                kasa_df, kasa_sum = check_kasa_activity_products(df_mag, kasa_kodlari)
                                
                                int_codes = set(int_df['Malzeme Kodu'].astype(str).tolist()) if len(int_df) > 0 else set()
                                chr_codes = set(chr_df['Malzeme Kodu'].astype(str).tolist()) if len(chr_df) > 0 else set()
                                
                                t20_df = create_top_20_risky(df_mag, int_codes, chr_codes, set())
                                exec_c, grp_s = generate_executive_summary(df_mag, kasa_df, kasa_sum)
                                
                                report_data = create_excel_report(
                                    df_mag, int_df, chr_df, chr_fire_df, cig_df,
                                    ext_df, fam_df, fire_df, kasa_df, t20_df,
                                    exec_c, grp_s, selected_mag_kod, mag_adi, params
                                )
                                
                                mag_adi_clean = mag_adi.replace(' ', '_').replace('/', '_')[:30] if mag_adi else ''
                                
                                st.download_button(
                                    "📥 İndir", 
                                    data=report_data,
                                    file_name=f"{selected_mag_kod}_{mag_adi_clean}_Risk_Raporu.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    key="sm_download_report"
                                )
                                st.success("✅ Rapor hazır!")
                            else:
                                st.error("Veri çekilemedi!")
                
                with tabs[1]:
                    st.subheader("🔴 Kritik Mağazalar")
                    kritik_df = region_df[region_df['Risk'].str.contains('KRİTİK')]
                    if len(kritik_df) > 0:
                        for _, row in kritik_df.iterrows():
                            # Risk nedenlerini hesapla (VIEW'de yok)
                            nedenler = []
                            if row.get('Sigara', 0) > 0:
                                nedenler.append(f"🚬 Sigara: {row['Sigara']:.0f}")
                            if row.get('İç Hırs.', 0) > 5:
                                nedenler.append(f"🔒 İç Hırs: {row['İç Hırs.']:.0f}")
                            if row.get('Toplam %', 0) >= 2:
                                nedenler.append(f"📊 Yüksek Kayıp")
                            neden_str = " | ".join(nedenler) if nedenler else "Yüksek kayıp oranı"
                            
                            st.error(f"**{row['Mağaza Kodu']} - {row['Mağaza Adı']}**\n\n"
                                    f"Kayıp: %{row['Toplam %']:.1f} | Fark: {row['Fark']:,.0f} TL\n\n"
                                    f"**Neden:** {neden_str}")
                    else:
                        st.success("Kritik mağaza yok! 🎉")
                
                with tabs[2]:
                    st.subheader("🟠 Riskli Mağazalar")
                    riskli_df = region_df[region_df['Risk'].str.contains('RİSKLİ')]
                    if len(riskli_df) > 0:
                        for _, row in riskli_df.iterrows():
                            # Risk nedenlerini hesapla
                            nedenler = []
                            if row.get('Sigara', 0) > 0:
                                nedenler.append(f"🚬 Sigara: {row['Sigara']:.0f}")
                            if row.get('İç Hırs.', 0) > 5:
                                nedenler.append(f"🔒 İç Hırs: {row['İç Hırs.']:.0f}")
                            neden_str = " | ".join(nedenler) if nedenler else "Kayıp oranı yüksek"
                            
                            st.warning(f"**{row['Mağaza Kodu']} - {row['Mağaza Adı']}**\n\n"
                                      f"Kayıp: %{row['Toplam %']:.1f} | Fark: {row['Fark']:,.0f} TL\n\n"
                                      f"**Neden:** {neden_str}")
                    else:
                        st.success("Riskli mağaza yok! 🎉")
                
                with tabs[3]:
                    st.subheader("🚬 Sigara Açığı Olan Mağazalar")
                    sigara_df = region_df[region_df['Sigara'] > 0].sort_values('Sigara', ascending=False)
                    if len(sigara_df) > 0:
                        st.error(f"⚠️ {len(sigara_df)} mağazada sigara açığı var!")
                        for _, row in sigara_df.iterrows():
                            st.error(f"**{row['Mağaza Kodu']} - {row['Mağaza Adı']}**: {row['Sigara']:.0f} ürün sigara açığı")
                    else:
                        st.success("Sigara açığı olan mağaza yok! 🎉")
                
                with tabs[4]:
                    st.subheader("🔍 Mağaza Detay Görünümü")
                    st.info("Bir mağaza seçerek İç Hırsızlık, Kronik Ürünler ve Sigara detaylarını görüntüleyebilirsiniz.")
                    
                    # Mağaza seçimi
                    mag_options_detay = [f"{row['Mağaza Kodu']} - {row['Mağaza Adı']}" for _, row in region_df.iterrows()]
                    selected_mag_detay = st.selectbox("📍 Mağaza Seçin", mag_options_detay, key="sm_mag_detay_select")
                    
                    if st.button("🔍 Detayları Getir", key="sm_get_details"):
                        selected_mag_kod_detay = selected_mag_detay.split(" - ")[0]
                        
                        with st.spinner("📊 Mağaza detayları yükleniyor..."):
                            # Sadece bu mağazanın verisini çek
                            df_mag_detay = get_single_store_data(selected_mag_kod_detay, tuple(selected_periods) if selected_periods else None)
                            
                            if len(df_mag_detay) > 0:
                                df_mag_detay = analyze_inventory(df_mag_detay)
                                
                                # İç Hırsızlık analizi
                                int_df_detay = detect_internal_theft(df_mag_detay)
                                
                                # Kamera entegrasyonu
                                if len(int_df_detay) > 0:
                                    try:
                                        env_tarihi_detay = df_mag_detay['Envanter Tarihi'].iloc[0]
                                        int_df_detay = enrich_internal_theft_with_camera(int_df_detay, selected_mag_kod_detay, env_tarihi_detay, full_df=df_mag_detay)
                                    except Exception as e:
                                        st.warning(f"Kamera entegrasyonu hatası: {e}")
                                
                                # Kronik ve Sigara
                                chr_df_detay = detect_chronic_products(df_mag_detay)
                                cig_df_detay = detect_cigarette_shortage(df_mag_detay)
                                
                                # Sonuçları göster
                                detay_tabs = st.tabs(["🔒 İç Hırsızlık", "🔄 Kronik Ürünler", "🚬 Sigara"])
                                
                                with detay_tabs[0]:
                                    st.markdown(f"### 🔒 İç Hırsızlık Şüphelileri ({len(int_df_detay)} ürün)")
                                    if len(int_df_detay) > 0:
                                        # Gösterilecek sütunlar
                                        display_cols_int = ['Malzeme Kodu', 'Malzeme Adı', 'Satış Fiyatı', 'TOPLAM', 
                                                           'İptal Satır', 'Durum', 'Risk', 'Fark Tutarı (TL)']
                                        if 'KAMERA KONTROL DETAY' in int_df_detay.columns:
                                            display_cols_int.append('KAMERA KONTROL DETAY')
                                        
                                        available_cols = [c for c in display_cols_int if c in int_df_detay.columns]
                                        st.dataframe(int_df_detay[available_cols], use_container_width=True, hide_index=True)
                                    else:
                                        st.success("İç hırsızlık şüphelisi ürün bulunamadı! ✅")
                                
                                with detay_tabs[1]:
                                    st.markdown(f"### 🔄 Kronik Açık Ürünler ({len(chr_df_detay)} ürün)")
                                    if len(chr_df_detay) > 0:
                                        st.dataframe(chr_df_detay, use_container_width=True, hide_index=True)
                                    else:
                                        st.success("Kronik açık ürün bulunamadı! ✅")
                                
                                with detay_tabs[2]:
                                    st.markdown(f"### 🚬 Sigara Analizi")
                                    if len(cig_df_detay) > 0:
                                        st.dataframe(cig_df_detay, use_container_width=True, hide_index=True)
                                    else:
                                        st.success("Sigara açığı bulunamadı! ✅")
                            else:
                                st.error("Mağaza verisi bulunamadı")
                
                with tabs[5]:
                    st.subheader("📥 SM Raporu İndir")
                    
                    # ⚡ LAZY LOAD - Excel butonu basılınca tam veri çekilir
                    if st.button("📊 Excel Raporu Hazırla", key="prepare_sm_excel"):
                        with st.spinner("📊 Detaylı veri yükleniyor..."):
                            # Tam veri çek (sadece bu SM için)
                            df_full = get_data_from_supabase(satis_muduru=selected_sm, donemler=selected_periods)
                            
                            if len(df_full) > 0:
                                df_analyzed = analyze_inventory(df_full)
                                
                                # Excel oluştur
                                excel_data = create_region_excel_report(region_df, df_analyzed, kasa_kodlari, params)
                                
                                st.download_button(
                                    label=f"📥 {display_sm} Özet Raporu (Excel)",
                                    data=excel_data,
                                    file_name=f"SM_OZET_{display_sm}_{params.get('donem', '')}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )
                                st.success("✅ Excel hazır!")
                            else:
                                st.error("Veri çekilemedi!")

# GM Özet modu - Sadece GM için
elif analysis_mode == "🌍 GM Özet":
    st.subheader("🌍 GM Özet - Bölge Dashboard")
    
    # ⚡ VIEW KULLAN - load_all_data_once YOK
    
    # Dönem seçimi - cache'den al
    available_periods = get_available_periods_cached()
    
    if available_periods:
        selected_periods = st.multiselect("📅 Dönem Seçin", available_periods, default=available_periods[:1])
    else:
        selected_periods = []
        st.warning("Henüz veri yüklenmemiş. SM'ler Excel yükledikçe veriler burada görünecek.")
    
    # Tarih aralığı filtresi (opsiyonel)
    gm_tarih_baslangic = None
    gm_tarih_bitis = None
    
    if selected_periods:
        # Seçilen dönemlerdeki envanter tarihlerini getir
        donem_tarihleri = get_envanter_tarihleri_by_donem(tuple(selected_periods))
        
        if donem_tarihleri and len(donem_tarihleri) > 1:
            with st.expander("📆 Tarih Aralığı Filtresi (Opsiyonel)", expanded=False):
                col_t1, col_t2 = st.columns(2)
                with col_t1:
                    min_tarih = min(donem_tarihleri)
                    max_tarih = max(donem_tarihleri)
                    gm_tarih_baslangic = st.date_input(
                        "Başlangıç Tarihi", 
                        value=min_tarih,
                        min_value=min_tarih,
                        max_value=max_tarih,
                        key="gm_tarih_bas"
                    )
                with col_t2:
                    gm_tarih_bitis = st.date_input(
                        "Bitiş Tarihi", 
                        value=max_tarih,
                        min_value=min_tarih,
                        max_value=max_tarih,
                        key="gm_tarih_bit"
                    )
                
                # Eğer varsayılan değerler seçiliyse filtre uygulanmasın
                if gm_tarih_baslangic == min_tarih and gm_tarih_bitis == max_tarih:
                    gm_tarih_baslangic = None
                    gm_tarih_bitis = None
                else:
                    st.info(f"📆 Filtre: {gm_tarih_baslangic.strftime('%d.%m.%Y')} - {gm_tarih_bitis.strftime('%d.%m.%Y')}")
    
    if selected_periods:
        # ⚡ SÜPER HIZLI - Supabase VIEW'den direkt özet veri (TÜM SM'ler)
        region_df = get_sm_summary_from_view(
            satis_muduru=None, 
            donemler=selected_periods,
            tarih_baslangic=gm_tarih_baslangic,
            tarih_bitis=gm_tarih_bitis
        )
        
        if len(region_df) == 0:
            st.warning("Seçilen döneme ait veri bulunamadı")
        else:
            magazalar = region_df['Mağaza Kodu'].dropna().unique().tolist()
            
            params = {
                'donem': ', '.join(selected_periods),
                'tarih': datetime.now().strftime('%Y-%m-%d'),
            }
            
            # Kasa aktivitesi kodlarını yükle
            kasa_kodlari = load_kasa_activity_codes()
            
            # SM sütunu ekle (VIEW'de zaten var)
            if 'SM' not in region_df.columns:
                region_df['SM'] = region_df['Satış Müdürü']
            
            # SM ve BS agregasyonları
            sm_df = aggregate_by_group(region_df, 'SM') if 'SM' in region_df.columns else pd.DataFrame()
            bs_df = aggregate_by_group(region_df, 'BS') if 'BS' in region_df.columns else pd.DataFrame()
            
            # ⚡ Risk puanına göre sırala (yüksekten düşüğe)
            region_df = region_df.sort_values('Risk Puan', ascending=False)
            
            if len(region_df) == 0:
                st.error("Analiz edilecek mağaza bulunamadı!")
            else:
                # Bölge toplamları
                toplam_satis = region_df['Satış'].sum()
                toplam_fark = region_df['Fark'].sum()
                toplam_fire = region_df['Fire'].sum()
                toplam_acik = region_df['Toplam Açık'].sum()
                toplam_gun = region_df['Gün'].sum()
                
                # Oranlar
                fark_oran = abs(toplam_fark) / toplam_satis * 100 if toplam_satis > 0 else 0
                fire_oran = abs(toplam_fire) / toplam_satis * 100 if toplam_satis > 0 else 0
                toplam_oran = abs(toplam_acik) / toplam_satis * 100 if toplam_satis > 0 else 0
                gunluk_fark = toplam_fark / toplam_gun if toplam_gun > 0 else 0
                gunluk_fire = toplam_fire / toplam_gun if toplam_gun > 0 else 0
                
                # Risk sayıları
                kritik_sayisi = len(region_df[region_df['Risk'].str.contains('KRİTİK')])
                riskli_sayisi = len(region_df[region_df['Risk'].str.contains('RİSKLİ')])
                dikkat_sayisi = len(region_df[region_df['Risk'].str.contains('DİKKAT')])
                temiz_sayisi = len(region_df[region_df['Risk'].str.contains('TEMİZ')])
                
                # 10TL Özet (VIEW: Kasa Adet/Tutar, analyze_region: 10TL Adet/Tutar)
                kasa_adet_col = 'Kasa Adet' if 'Kasa Adet' in region_df.columns else '10TL Adet'
                kasa_tutar_col = 'Kasa Tutar' if 'Kasa Tutar' in region_df.columns else '10TL Tutar'
                toplam_10tl_adet = region_df[kasa_adet_col].sum() if kasa_adet_col in region_df.columns else 0
                toplam_10tl_tutar = region_df[kasa_tutar_col].sum() if kasa_tutar_col in region_df.columns else 0
                
                # ========== GÖRÜNÜM ==========
                st.markdown("---")
                st.subheader(f"📊 Bölge Özeti - {len(region_df)} Mağaza")
                
                # Üst metrikler
                col1, col2, col3, col4, col5 = st.columns(5)
                col1.metric("💰 Satış", f"{toplam_satis/1e6:.1f}M TL")
                col2.metric("📉 Fark", f"%{fark_oran:.2f}", f"{toplam_fark:,.0f} | Gün: {gunluk_fark:,.0f}")
                col3.metric("🔥 Fire", f"%{fire_oran:.2f}", f"{toplam_fire:,.0f} | Gün: {gunluk_fire:,.0f}")
                col4.metric("📊 Toplam", f"%{toplam_oran:.2f}", f"{toplam_acik:,.0f}")
                
                if toplam_10tl_adet != 0:
                    col5.metric("💰 10 TL", f"{toplam_10tl_adet:+.0f}", f"{toplam_10tl_tutar:,.0f}₺")
                else:
                    col5.metric("💰 10 TL", "0", "TAMAM")
                
                # Risk dağılımı
                st.markdown("### 📊 Risk Dağılımı")
                r1, r2, r3, r4 = st.columns(4)
                r1.markdown(f'<div class="risk-kritik">🔴 KRİTİK: {kritik_sayisi}</div>', unsafe_allow_html=True)
                r2.markdown(f'<div class="risk-riskli">🟠 RİSKLİ: {riskli_sayisi}</div>', unsafe_allow_html=True)
                r3.markdown(f'<div class="risk-dikkat">🟡 DİKKAT: {dikkat_sayisi}</div>', unsafe_allow_html=True)
                r4.markdown(f'<div class="risk-temiz">🟢 TEMİZ: {temiz_sayisi}</div>', unsafe_allow_html=True)
                
                # Sekmeler
                tabs = st.tabs(["👔 SM Özet", "📋 BS Özet", "🏪 Mağazalar", "📊 Top 10", "🚬 Sigara", "🔍 Mağaza Detay", "📥 İndir"])
                
                with tabs[0]:
                    st.subheader("👔 Satış Müdürü Bazlı Özet")
                    if len(sm_df) > 0:
                        # Başlık
                        cols = st.columns([2, 1.5, 1.5, 1, 1, 1, 1])
                        cols[0].markdown("**Satış Müdürü**")
                        cols[1].markdown("**Satış | Fark**")
                        cols[2].markdown("**Fire**")
                        cols[3].markdown("**Kayıp %**")
                        cols[4].markdown("**🚬 🔒**")
                        cols[5].markdown("**Risk**")
                        cols[6].markdown("**Seviye**")
                        st.markdown("---")
                        
                        for _, row in sm_df.iterrows():
                            cols = st.columns([2, 1.5, 1.5, 1, 1, 1, 1])
                            cols[0].write(f"**{row['SM']}** ({row['Mağaza Sayısı']:.0f} mağ.)")
                            cols[1].write(f"{row['Satış']/1e6:.1f}M | {row['Fark']:,.0f}")
                            cols[2].write(f"{row['Fire']:,.0f}")
                            cols[3].write(f"%{row['Toplam %']:.1f}")
                            cols[4].write(f"🚬{row['Sigara']:.0f} 🔒{row['İç Hırs.']:.0f}")
                            cols[5].write(f"**{row['Risk Puan']:.0f}**")
                            cols[6].write(row['Risk'])
                    else:
                        st.info("SM verisi bulunamadı")
                
                with tabs[1]:
                    st.subheader("📋 Bölge Sorumlusu Bazlı Özet")
                    if len(bs_df) > 0:
                        # Başlık
                        cols = st.columns([2, 1.5, 1.5, 1, 1, 1, 1])
                        cols[0].markdown("**Bölge Sorumlusu**")
                        cols[1].markdown("**Satış | Fark**")
                        cols[2].markdown("**Fire**")
                        cols[3].markdown("**Kayıp %**")
                        cols[4].markdown("**🚬 🔒**")
                        cols[5].markdown("**Risk**")
                        cols[6].markdown("**Seviye**")
                        st.markdown("---")
                        
                        for _, row in bs_df.iterrows():
                            cols = st.columns([2, 1.5, 1.5, 1, 1, 1, 1])
                            cols[0].write(f"**{row['BS']}** ({row['Mağaza Sayısı']:.0f} mağ.)")
                            cols[1].write(f"{row['Satış']/1e6:.1f}M | {row['Fark']:,.0f}")
                            cols[2].write(f"{row['Fire']:,.0f}")
                            cols[3].write(f"%{row['Toplam %']:.1f}")
                            cols[4].write(f"🚬{row['Sigara']:.0f} 🔒{row['İç Hırs.']:.0f}")
                            cols[5].write(f"**{row['Risk Puan']:.0f}**")
                            cols[6].write(row['Risk'])
                    else:
                        st.info("BS verisi bulunamadı")
                
                with tabs[2]:
                    st.subheader("🏪 Tüm Mağazalar")
                    
                    # Filtreler
                    col_f1, col_f2, col_f3 = st.columns(3)
                    with col_f1:
                        risk_filter = st.multiselect("Risk Seviyesi", ["🔴 KRİTİK", "🟠 RİSKLİ", "🟡 DİKKAT", "🟢 TEMİZ"])
                    with col_f2:
                        if 'SM' in region_df.columns:
                            sm_filter = st.multiselect("Satış Müdürü", region_df['SM'].unique().tolist())
                        else:
                            sm_filter = []
                    with col_f3:
                        bs_filter = st.multiselect("Bölge Sorumlusu", region_df['BS'].unique().tolist())
                    
                    filtered_df = region_df.copy()
                    if risk_filter:
                        filtered_df = filtered_df[filtered_df['Risk'].isin(risk_filter)]
                    if sm_filter:
                        filtered_df = filtered_df[filtered_df['SM'].isin(sm_filter)]
                    if bs_filter:
                        filtered_df = filtered_df[filtered_df['BS'].isin(bs_filter)]
                    
                    st.info(f"📊 {len(filtered_df)} mağaza gösteriliyor")
                    
                    display_cols = ['Mağaza Kodu', 'Mağaza Adı', 'SM', 'BS', 'Satış', 'Fark', 'Fark %', 
                                   'Fire', 'Fire %', 'Toplam Açık', 'Toplam %', 'Sigara', 'İç Hırs.', 'Risk Puan', 'Risk']
                    display_cols = [c for c in display_cols if c in filtered_df.columns]
                    st.dataframe(filtered_df[display_cols].sort_values('Risk Puan', ascending=False), 
                                use_container_width=True, height=500)
                
                with tabs[3]:
                    st.subheader("📊 En Riskli 10 Mağaza")
                    top10 = region_df.nlargest(10, 'Risk Puan')
                    
                    for _, row in top10.iterrows():
                        risk_text = row.get('Risk', '')
                        sm_text = row.get('SM', '') if 'SM' in row else ''
                        
                        msg = f"**{row['Mağaza Kodu']} - {row['Mağaza Adı']}** | {sm_text} | Risk: {row['Risk Puan']:.0f}\n\n"
                        msg += f"Fark: {row['Fark']:,.0f} | Fire: {row['Fire']:,.0f} | Kayıp: %{row['Toplam %']:.1f}\n\n"
                        msg += f"🚬 Sigara: {row['Sigara']:.0f} | 🔒 İç Hırs: {row['İç Hırs.']:.0f}"
                        
                        if 'KRİTİK' in str(risk_text):
                            st.error(msg)
                        elif 'RİSKLİ' in str(risk_text):
                            st.warning(msg)
                        else:
                            st.info(msg)
                
                with tabs[4]:
                    st.subheader("🚬 Sigara Açığı Olan Mağazalar")
                    sigara_df = region_df[region_df['Sigara'] > 0].sort_values('Sigara', ascending=False)
                    if len(sigara_df) > 0:
                        st.error(f"⚠️ {len(sigara_df)} mağazada sigara açığı var!")
                        for _, row in sigara_df.iterrows():
                            sm_text = row.get('SM', '') if 'SM' in row else ''
                            st.error(f"**{row['Mağaza Kodu']} - {row['Mağaza Adı']}** | {sm_text} | 🚬 {row['Sigara']:.0f} adet sigara açığı")
                    else:
                        st.success("Sigara açığı olan mağaza yok! 🎉")
                
                with tabs[5]:
                    st.subheader("🔍 Mağaza Detay Görünümü")
                    st.info("Bir mağaza seçerek İç Hırsızlık, Kronik Ürünler ve Sigara detaylarını görüntüleyebilirsiniz.")
                    
                    # Mağaza seçimi
                    mag_options_gm_detay = [f"{row['Mağaza Kodu']} - {row['Mağaza Adı']}" for _, row in region_df.iterrows()]
                    selected_mag_gm_detay = st.selectbox("📍 Mağaza Seçin", mag_options_gm_detay, key="gm_mag_detay_select")
                    
                    if st.button("🔍 Detayları Getir", key="gm_get_details"):
                        selected_mag_kod_gm_detay = selected_mag_gm_detay.split(" - ")[0]
                        
                        with st.spinner("📊 Mağaza detayları yükleniyor..."):
                            # Sadece bu mağazanın verisini çek
                            df_mag_gm_detay = get_single_store_data(selected_mag_kod_gm_detay, tuple(selected_periods) if selected_periods else None)
                            
                            if len(df_mag_gm_detay) > 0:
                                df_mag_gm_detay = analyze_inventory(df_mag_gm_detay)
                                
                                # İç Hırsızlık analizi
                                int_df_gm_detay = detect_internal_theft(df_mag_gm_detay)
                                
                                # Kamera entegrasyonu
                                if len(int_df_gm_detay) > 0:
                                    try:
                                        env_tarihi_gm_detay = df_mag_gm_detay['Envanter Tarihi'].iloc[0]
                                        int_df_gm_detay = enrich_internal_theft_with_camera(int_df_gm_detay, selected_mag_kod_gm_detay, env_tarihi_gm_detay, full_df=df_mag_gm_detay)
                                    except Exception as e:
                                        st.warning(f"Kamera entegrasyonu hatası: {e}")
                                
                                # Kronik ve Sigara
                                chr_df_gm_detay = detect_chronic_products(df_mag_gm_detay)
                                cig_df_gm_detay = detect_cigarette_shortage(df_mag_gm_detay)
                                
                                # Sonuçları göster
                                gm_detay_tabs = st.tabs(["🔒 İç Hırsızlık", "🔄 Kronik Ürünler", "🚬 Sigara"])
                                
                                with gm_detay_tabs[0]:
                                    st.markdown(f"### 🔒 İç Hırsızlık Şüphelileri ({len(int_df_gm_detay)} ürün)")
                                    if len(int_df_gm_detay) > 0:
                                        display_cols_gm = ['Malzeme Kodu', 'Malzeme Adı', 'Satış Fiyatı', 'TOPLAM', 
                                                          'İptal Satır', 'Durum', 'Risk', 'Fark Tutarı (TL)']
                                        if 'KAMERA KONTROL DETAY' in int_df_gm_detay.columns:
                                            display_cols_gm.append('KAMERA KONTROL DETAY')
                                        
                                        available_cols_gm = [c for c in display_cols_gm if c in int_df_gm_detay.columns]
                                        st.dataframe(int_df_gm_detay[available_cols_gm], use_container_width=True, hide_index=True)
                                    else:
                                        st.success("İç hırsızlık şüphelisi ürün bulunamadı! ✅")
                                
                                with gm_detay_tabs[1]:
                                    st.markdown(f"### 🔄 Kronik Açık Ürünler ({len(chr_df_gm_detay)} ürün)")
                                    if len(chr_df_gm_detay) > 0:
                                        st.dataframe(chr_df_gm_detay, use_container_width=True, hide_index=True)
                                    else:
                                        st.success("Kronik açık ürün bulunamadı! ✅")
                                
                                with gm_detay_tabs[2]:
                                    st.markdown(f"### 🚬 Sigara Analizi")
                                    if len(cig_df_gm_detay) > 0:
                                        st.dataframe(cig_df_gm_detay, use_container_width=True, hide_index=True)
                                    else:
                                        st.success("Sigara açığı bulunamadı! ✅")
                            else:
                                st.error("Mağaza verisi bulunamadı")
                
                with tabs[6]:
                    st.subheader("📥 Raporları İndir")
                    
                    # GM Excel raporu
                    excel_data = create_gm_excel_report(region_df, sm_df, bs_df, params)
                    
                    st.download_button(
                        label="📥 GM Bölge Dashboard (Excel)",
                        data=excel_data,
                        file_name=f"GM_BOLGE_DASHBOARD_{params.get('donem', '')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
                    st.markdown("---")
                    
                    # Tek mağaza raporu indirme
                    st.markdown("**📥 Mağaza Detay Raporu İndir**")
                    
                    mag_options_gm = [f"{row['Mağaza Kodu']} - {row['Mağaza Adı']}" for _, row in region_df.iterrows()]
                    selected_mag_gm = st.selectbox("Mağaza seçin", mag_options_gm, key="gm_mag_select")
                    
                    if st.button("📥 Mağaza Raporu Oluştur", key="gm_create_mag_report"):
                        selected_mag_kod_gm = selected_mag_gm.split(" - ")[0]
                        selected_row_gm = region_df[region_df['Mağaza Kodu'] == selected_mag_kod_gm].iloc[0]
                        
                        with st.spinner("📊 Mağaza verisi yükleniyor (5-10 saniye)..."):
                            # ⚡ HIZLI - Sadece bu mağaza için veri çek
                            df_mag_gm = get_single_store_data(selected_mag_kod_gm, tuple(selected_periods) if selected_periods else None)
                            
                            if len(df_mag_gm) > 0:
                                df_mag_gm = analyze_inventory(df_mag_gm)
                                mag_adi_gm = selected_row_gm['Mağaza Adı']
                                
                                # Kasa kodlarını yükle
                                kasa_kodlari_gm = load_kasa_activity_codes()
                                
                                # Analizleri yap
                                int_df_gm = detect_internal_theft(df_mag_gm)
                                
                                if len(int_df_gm) > 0:
                                    try:
                                        env_tarihi_gm = df_mag_gm['Envanter Tarihi'].iloc[0]
                                        int_df_gm = enrich_internal_theft_with_camera(int_df_gm, selected_mag_kod_gm, env_tarihi_gm, full_df=df_mag_gm)
                                    except:
                                        pass
                                
                                chr_df_gm = detect_chronic_products(df_mag_gm)
                                chr_fire_df_gm = detect_chronic_fire(df_mag_gm)
                                cig_df_gm = detect_cigarette_shortage(df_mag_gm)
                                ext_df_gm = detect_external_theft(df_mag_gm)
                                fam_df_gm = find_product_families(df_mag_gm)
                                fire_df_gm = detect_fire_manipulation(df_mag_gm)
                                kasa_df_gm, kasa_sum_gm = check_kasa_activity_products(df_mag_gm, kasa_kodlari_gm)
                                
                                int_codes_gm = set(int_df_gm['Malzeme Kodu'].astype(str).tolist()) if len(int_df_gm) > 0 else set()
                                chr_codes_gm = set(chr_df_gm['Malzeme Kodu'].astype(str).tolist()) if len(chr_df_gm) > 0 else set()
                                
                                t20_df_gm = create_top_20_risky(df_mag_gm, int_codes_gm, chr_codes_gm, set())
                                exec_c_gm, grp_s_gm = generate_executive_summary(df_mag_gm, kasa_df_gm, kasa_sum_gm)
                                
                                report_data_gm = create_excel_report(
                                    df_mag_gm, int_df_gm, chr_df_gm, chr_fire_df_gm, cig_df_gm,
                                    ext_df_gm, fam_df_gm, fire_df_gm, kasa_df_gm, t20_df_gm,
                                    exec_c_gm, grp_s_gm, selected_mag_kod_gm, mag_adi_gm, params
                                )
                                
                                mag_adi_clean_gm = mag_adi_gm.replace(' ', '_').replace('/', '_')[:30] if mag_adi_gm else ''
                                
                                st.download_button(
                                    "📥 İndir", 
                                    data=report_data_gm,
                                    file_name=f"{selected_mag_kod_gm}_{mag_adi_clean_gm}_Risk_Raporu.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    key="gm_download_mag_report"
                                )
                                st.success(f"✅ {selected_mag_kod_gm} raporu hazır!")
                            else:
                                st.error("Mağaza verisi bulunamadı")
                    
                    st.markdown("---")
                    st.markdown("""
                    **Excel İçeriği:**
                    - 📋 Bölge Özeti (Genel metrikler, risk dağılımı)
                    - 👔 SM Bazlı Analiz
                    - 👤 BS Bazlı Analiz  
                    - 🏪 Tüm Mağazalar (Risk puanına göre sıralı)
                    """)

elif uploaded_file is not None:
    try:
        xl = pd.ExcelFile(uploaded_file)
        sheet_names = xl.sheet_names
        
        best_sheet = None
        max_cols = 0
        
        for sheet in sheet_names:
            temp_df = pd.read_excel(uploaded_file, sheet_name=sheet, nrows=5)
            if len(temp_df.columns) > max_cols:
                max_cols = len(temp_df.columns)
                best_sheet = sheet
        
        df_raw = pd.read_excel(uploaded_file, sheet_name=best_sheet)
        st.success(f"✅ {len(df_raw)} satır, {len(df_raw.columns)} sütun ({best_sheet})")
        
        # ===== ARKA PLANDA SUPABASE'E KAYIT =====
        with st.spinner("Veritabanına kaydediliyor..."):
            try:
                inserted, skipped, result_info = save_to_supabase(df_raw)
                if inserted > 0:
                    st.info(f"💾 {inserted:,} kayıt eklendi | ⏭️ {skipped} envanter zaten mevcut")
                elif skipped > 0:
                    st.info(f"⏭️ Tüm envanterler zaten mevcut ({skipped} envanter)")
            except Exception as e:
                # Supabase hatası analizi engellemesin
                st.warning(f"⚠️ Veritabanı kaydı atlandı: {str(e)[:50]}")
        
        df = analyze_inventory(df_raw)
        
        # Mağaza bilgisi
        if 'Mağaza Kodu' in df.columns:
            magazalar = df['Mağaza Kodu'].dropna().unique().tolist()
            # Mağaza kod-isim eşleştirmesi
            magaza_isimleri = {}
            for mag in magazalar:
                isim = df[df['Mağaza Kodu'] == mag]['Mağaza Adı'].iloc[0] if 'Mağaza Adı' in df.columns else ''
                magaza_isimleri[mag] = f"{mag} - {isim}" if isim else str(mag)
        else:
            magazalar = ['MAGAZA']
            df['Mağaza Kodu'] = 'MAGAZA'
            magaza_isimleri = {'MAGAZA': 'MAGAZA'}
        
        params = {
            'donem': str(df['Envanter Dönemi'].iloc[0]) if 'Envanter Dönemi' in df.columns else '',
            'tarih': str(df['Envanter Tarihi'].iloc[0])[:10] if 'Envanter Tarihi' in df.columns else '',
        }
        
        # Kasa aktivitesi kodlarını yükle
        kasa_kodlari = load_kasa_activity_codes()
        
        # ========== BÖLGE ÖZETİ MODU ==========
        if analysis_mode == "🌍 Bölge Özeti":
            # Tarih aralığı filtresi (opsiyonel)
            if 'Envanter Tarihi' in df.columns:
                try:
                    df['Envanter Tarihi'] = pd.to_datetime(df['Envanter Tarihi'])
                    envanter_tarihleri = df['Envanter Tarihi'].dropna().dt.date.unique()
                    envanter_tarihleri = sorted(envanter_tarihleri)
                    
                    if len(envanter_tarihleri) > 1:
                        with st.expander("📆 Tarih Aralığı Filtresi (Opsiyonel)", expanded=False):
                            col_t1, col_t2 = st.columns(2)
                            with col_t1:
                                min_tarih = min(envanter_tarihleri)
                                max_tarih = max(envanter_tarihleri)
                                bolge_tarih_bas = st.date_input(
                                    "Başlangıç Tarihi", 
                                    value=min_tarih,
                                    min_value=min_tarih,
                                    max_value=max_tarih,
                                    key="bolge_tarih_bas"
                                )
                            with col_t2:
                                bolge_tarih_bit = st.date_input(
                                    "Bitiş Tarihi", 
                                    value=max_tarih,
                                    min_value=min_tarih,
                                    max_value=max_tarih,
                                    key="bolge_tarih_bit"
                                )
                            
                            # Tarih filtresi uygula
                            if bolge_tarih_bas != min_tarih or bolge_tarih_bit != max_tarih:
                                df = df[(df['Envanter Tarihi'].dt.date >= bolge_tarih_bas) & 
                                       (df['Envanter Tarihi'].dt.date <= bolge_tarih_bit)]
                                magazalar = df['Mağaza Kodu'].dropna().unique().tolist()
                                st.info(f"📆 Filtre: {bolge_tarih_bas.strftime('%d.%m.%Y')} - {bolge_tarih_bit.strftime('%d.%m.%Y')} | {len(magazalar)} mağaza")
                except:
                    pass
            
            st.subheader(f"🌍 Bölge Özeti - {len(magazalar)} Mağaza")
            
            with st.spinner("Tüm mağazalar analiz ediliyor..."):
                region_df = analyze_region(df, kasa_kodlari)
            
            # ⚡ Risk puanına göre sırala (yüksekten düşüğe)
            if len(region_df) > 0:
                region_df = region_df.sort_values('Risk Puan', ascending=False)
            
            if len(region_df) == 0:
                st.warning("Analiz edilecek mağaza bulunamadı!")
            else:
                # Bölge toplamları
                toplam_satis = region_df['Satış'].sum()
                toplam_fark = region_df['Fark'].sum()  # Fark + Kısmi
                toplam_fire = region_df['Fire'].sum()
                toplam_acik = region_df['Toplam Açık'].sum()  # Fark + Fire
                toplam_gun = region_df['Gün'].sum()
                
                # Oranlar
                fark_oran = abs(toplam_fark) / toplam_satis * 100 if toplam_satis > 0 else 0
                fire_oran = abs(toplam_fire) / toplam_satis * 100 if toplam_satis > 0 else 0
                toplam_oran = abs(toplam_acik) / toplam_satis * 100 if toplam_satis > 0 else 0
                gunluk_fark = toplam_fark / toplam_gun if toplam_gun > 0 else 0
                gunluk_fire = toplam_fire / toplam_gun if toplam_gun > 0 else 0
                
                # Risk dağılımı
                kritik_sayisi = len(region_df[region_df['Risk'].str.contains('KRİTİK')])
                riskli_sayisi = len(region_df[region_df['Risk'].str.contains('RİSKLİ')])
                dikkat_sayisi = len(region_df[region_df['Risk'].str.contains('DİKKAT')])
                temiz_sayisi = len(region_df[region_df['Risk'].str.contains('TEMİZ')])
                
                # Üst metrikler
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("💰 Toplam Satış", f"{toplam_satis/1_000_000:.1f}M TL")
                with col2:
                    st.metric("📉 Fark", f"{toplam_fark:,.0f} TL", f"%{fark_oran:.2f} | Günlük: {gunluk_fark:,.0f}₺")
                with col3:
                    st.metric("🔥 Fire", f"{toplam_fire:,.0f} TL", f"%{fire_oran:.2f} | Günlük: {gunluk_fire:,.0f}₺")
                with col4:
                    st.metric("📊 Toplam", f"{toplam_acik:,.0f} TL", f"%{toplam_oran:.2f}")
                
                # Risk dağılımı
                st.markdown("### 📊 Risk Dağılımı")
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    if kritik_sayisi > 0:
                        st.markdown(f'<div class="risk-kritik">🔴 KRİTİK: {kritik_sayisi}</div>', unsafe_allow_html=True)
                    else:
                        st.metric("🔴 KRİTİK", kritik_sayisi)
                with col2:
                    if riskli_sayisi > 0:
                        st.markdown(f'<div class="risk-riskli">🟠 RİSKLİ: {riskli_sayisi}</div>', unsafe_allow_html=True)
                    else:
                        st.metric("🟠 RİSKLİ", riskli_sayisi)
                with col3:
                    if dikkat_sayisi > 0:
                        st.markdown(f'<div class="risk-dikkat">🟡 DİKKAT: {dikkat_sayisi}</div>', unsafe_allow_html=True)
                    else:
                        st.metric("🟡 DİKKAT", dikkat_sayisi)
                with col4:
                    st.markdown(f'<div class="risk-temiz">🟢 TEMİZ: {temiz_sayisi}</div>', unsafe_allow_html=True)
                
                # Sekmeler
                tabs = st.tabs(["📋 Sıralama", "🔴 Kritik", "🟠 Riskli", "🚬 Sigara", "📊 Detay", "📥 İndir"])
                
                with tabs[0]:
                    st.subheader("📋 Mağaza Sıralaması (Risk Puanına Göre)")
                    
                    # Başlık satırı - 10TL eklendi
                    cols = st.columns([0.4, 0.6, 1.3, 1.1, 0.8, 0.6, 0.8, 0.6, 0.5, 0.5, 0.4, 0.6, 0.5, 0.7])
                    cols[0].markdown("**📥**")
                    cols[1].markdown("**Kod**")
                    cols[2].markdown("**Mağaza**")
                    cols[3].markdown("**BS**")
                    cols[4].markdown("**Fark**")
                    cols[5].markdown("**G.Fark**")
                    cols[6].markdown("**Fire**")
                    cols[7].markdown("**G.Fire**")
                    cols[8].markdown("**Kayıp%**")
                    cols[9].markdown("**Fire%**")
                    cols[10].markdown("**🚬**")
                    cols[11].markdown("**💰10TL**")
                    cols[12].markdown("**Risk**")
                    cols[13].markdown("**Seviye**")
                    
                    st.markdown("---")
                    
                    # Veri satırları
                    for idx, (_, row) in enumerate(region_df.iterrows()):
                        cols = st.columns([0.4, 0.6, 1.3, 1.1, 0.8, 0.6, 0.8, 0.6, 0.5, 0.5, 0.4, 0.6, 0.5, 0.7])
                        
                        # Mağaza verisini al ve tam rapor oluştur
                        mag_kod = row['Mağaza Kodu']
                        df_mag = df[df['Mağaza Kodu'] == mag_kod].copy()
                        mag_adi = row['Mağaza Adı']
                        
                        # Analizleri yap
                        int_df = detect_internal_theft(df_mag)
                        
                        # Kamera timestamp entegrasyonu (kategori araması için full_df geçir)
                        if len(int_df) > 0:
                            try:
                                env_tarihi = df_mag['Envanter Tarihi'].iloc[0]
                                int_df = enrich_internal_theft_with_camera(int_df, mag_kod, env_tarihi, full_df=df_mag)
                            except:
                                pass
                        
                        chr_df = detect_chronic_products(df_mag)
                        chr_fire_df = detect_chronic_fire(df_mag)
                        cig_df = detect_cigarette_shortage(df_mag)
                        ext_df = detect_external_theft(df_mag)
                        fam_df = find_product_families(df_mag)
                        fire_df = detect_fire_manipulation(df_mag)
                        kasa_df, kasa_sum = check_kasa_activity_products(df_mag, kasa_kodlari)
                        
                        int_codes = set(int_df['Malzeme Kodu'].astype(str).tolist()) if len(int_df) > 0 else set()
                        chr_codes = set(chr_df['Malzeme Kodu'].astype(str).tolist()) if len(chr_df) > 0 else set()
                        
                        t20_df = create_top_20_risky(df_mag, int_codes, chr_codes, set())
                        exec_c, grp_s = generate_executive_summary(df_mag, kasa_df, kasa_sum)
                        
                        # Tam rapor oluştur
                        report_data = create_excel_report(
                            df_mag, int_df, chr_df, chr_fire_df, cig_df,
                            ext_df, fam_df, fire_df, kasa_df, t20_df,
                            exec_c, grp_s, mag_kod, mag_adi, params
                        )
                        
                        mag_adi_clean = mag_adi.replace(' ', '_').replace('/', '_')[:30] if mag_adi else ''
                        
                        with cols[0]:
                            st.download_button("📥", data=report_data, 
                                file_name=f"{mag_kod}_{mag_adi_clean}_Risk_Raporu.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"dl_{idx}")
                        cols[1].write(f"{row['Mağaza Kodu']}")
                        cols[2].write(f"{row['Mağaza Adı'][:15] if row['Mağaza Adı'] else '-'}")
                        cols[3].write(f"{row['BS'][:10] if row['BS'] else '-'}")
                        cols[4].write(f"{row['Fark']/1000:.0f}K")
                        cols[5].write(f"{row['Günlük Fark']:,.0f}")
                        cols[6].write(f"{row['Fire']/1000:.0f}K")
                        cols[7].write(f"{row['Günlük Fire']:,.0f}")
                        cols[8].write(f"%{row['Toplam %']:.1f}")
                        cols[9].write(f"%{row['Fire %']:.1f}")
                        cols[10].write(f"{row['Sigara']:.0f}" if row['Sigara'] > 0 else "-")
                        cols[11].write(f"{row.get('Kasa Tutar', 0):,.0f}")
                        cols[12].write(f"{row['Risk Puan']:.0f}")
                        cols[13].write(row['Risk'])
                
                with tabs[1]:
                    st.subheader("🔴 Kritik Mağazalar")
                    kritik_df = region_df[region_df['Risk'].str.contains('KRİTİK')]
                    if len(kritik_df) > 0:
                        for _, row in kritik_df.iterrows():
                            # Risk nedenlerini hesapla
                            nedenler = []
                            if row.get('Sigara', 0) > 0:
                                nedenler.append(f"🚬 Sigara: {row['Sigara']:.0f}")
                            if row.get('İç Hırs.', 0) > 5:
                                nedenler.append(f"🔒 İç Hırs: {row['İç Hırs.']:.0f}")
                            if row.get('Toplam %', 0) >= 2:
                                nedenler.append(f"📊 Yüksek Kayıp")
                            neden_str = " | ".join(nedenler) if nedenler else "Yüksek kayıp oranı"
                            
                            st.error(f"**{row['Mağaza Kodu']} - {row['Mağaza Adı']}**\n\n"
                                    f"Kayıp: %{row['Toplam %']:.1f} | Fark: {row['Fark']:,.0f} TL\n\n"
                                    f"**Neden:** {neden_str}")
                    else:
                        st.success("Kritik mağaza yok! 🎉")
                
                with tabs[2]:
                    st.subheader("🟠 Riskli Mağazalar")
                    riskli_df = region_df[region_df['Risk'].str.contains('RİSKLİ')]
                    if len(riskli_df) > 0:
                        for _, row in riskli_df.iterrows():
                            # Risk nedenlerini hesapla
                            nedenler = []
                            if row.get('Sigara', 0) > 0:
                                nedenler.append(f"🚬 Sigara: {row['Sigara']:.0f}")
                            if row.get('İç Hırs.', 0) > 5:
                                nedenler.append(f"🔒 İç Hırs: {row['İç Hırs.']:.0f}")
                            neden_str = " | ".join(nedenler) if nedenler else "Kayıp oranı yüksek"
                            
                            st.warning(f"**{row['Mağaza Kodu']} - {row['Mağaza Adı']}**\n\n"
                                      f"Kayıp: %{row['Toplam %']:.1f} | Fark: {row['Fark']:,.0f} TL\n\n"
                                      f"**Neden:** {neden_str}")
                    else:
                        st.success("Riskli mağaza yok! 🎉")
                
                with tabs[3]:
                    st.subheader("🚬 Sigara Açığı Olan Mağazalar")
                    sigara_df = region_df[region_df['Sigara'] > 0].sort_values('Sigara', ascending=False)
                    if len(sigara_df) > 0:
                        st.error(f"⚠️ {len(sigara_df)} mağazada sigara açığı var!")
                        for _, row in sigara_df.iterrows():
                            st.error(f"**{row['Mağaza Kodu']} - {row['Mağaza Adı']}**: {row['Sigara']} ürün sigara açığı")
                    else:
                        st.success("Sigara açığı olan mağaza yok! 🎉")
                
                with tabs[4]:
                    st.subheader("📊 Tüm Detaylar")
                    st.dataframe(region_df, use_container_width=True, hide_index=True)
                
                with tabs[5]:
                    st.subheader("📥 Bölge Raporu İndir")
                    
                    excel_data = create_region_excel_report(region_df, df, kasa_kodlari, params)
                    
                    st.download_button(
                        label="📥 Bölge Özet Raporu (Excel)",
                        data=excel_data,
                        file_name=f"BOLGE_OZET_{params.get('donem', '')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
        
        # ========== TEK MAĞAZA MODU ==========
        else:
            # Mağaza seçimi
            if len(magazalar) > 1:
                # Kod + isim listesi oluştur
                magaza_options = [magaza_isimleri[m] for m in magazalar]
                selected_option = st.selectbox("🏪 Mağaza Seçin", magaza_options)
                # Seçilen option'dan kodu çıkar
                selected_str = selected_option.split(" - ")[0]
                # Orijinal tipte bul
                selected = None
                for m in magazalar:
                    if str(m) == selected_str:
                        selected = m
                        break
                if selected is None:
                    selected = magazalar[0]
                df_display = df[df['Mağaza Kodu'] == selected].copy()
                magaza_adi = df_display['Mağaza Adı'].iloc[0] if 'Mağaza Adı' in df_display.columns and len(df_display) > 0 else ''
            else:
                selected = magazalar[0]
                df_display = df.copy()
                magaza_adi = df['Mağaza Adı'].iloc[0] if 'Mağaza Adı' in df.columns and len(df) > 0 else ''
        
            # Kasa aktivitesi kodlarını yükle
            kasa_kodlari = load_kasa_activity_codes()
        
            # Analizler
            internal_df = detect_internal_theft(df_display)
            
            # Kamera timestamp entegrasyonu
            st.info(f"🔍 İç Hırsızlık: {len(internal_df)} ürün bulundu")
            if len(internal_df) > 0:
                try:
                    magaza_kodu = df_display['Mağaza Kodu'].iloc[0]
                    envanter_tarihi = df_display['Envanter Tarihi'].iloc[0]
                    st.info(f"📹 Kamera entegrasyonu başlıyor - Mağaza: {magaza_kodu}")
                    
                    # Debug: Sheets verisini kontrol et
                    df_sheets_test = get_iptal_verisi_from_sheets()
                    st.write(f"📥 Sheets satır sayısı: {len(df_sheets_test)}")
                    if not df_sheets_test.empty:
                        # 7915 mağazası için kayıt sayısı
                        mag_col = 'Mağaza - Anahtar' if 'Mağaza - Anahtar' in df_sheets_test.columns else df_sheets_test.columns[7]
                        df_sheets_test[mag_col] = df_sheets_test[mag_col].astype(str).str.replace('.0', '', regex=False)
                        mag_count = len(df_sheets_test[df_sheets_test[mag_col] == str(magaza_kodu)])
                        st.write(f"🏪 Mağaza {magaza_kodu} iptal sayısı: {mag_count}")
                    
                    internal_df = enrich_internal_theft_with_camera(internal_df, magaza_kodu, envanter_tarihi, full_df=df_display)
                    st.success(f"✅ Kamera entegrasyonu tamamlandı")
                except Exception as e:
                    st.error(f"❌ Kamera entegrasyonu hatası: {e}")
                    import traceback
                    st.code(traceback.format_exc())
            
            chronic_df = detect_chronic_products(df_display)
            chronic_fire_df = detect_chronic_fire(df_display)
            cigarette_df = detect_cigarette_shortage(df_display)
            external_df = detect_external_theft(df_display)
            family_df = find_product_families(df_display)
            fire_manip_df = detect_fire_manipulation(df_display)
            kasa_activity_df, kasa_summary = check_kasa_activity_products(df_display, kasa_kodlari)
            exec_comments, group_stats = generate_executive_summary(df_display, kasa_activity_df, kasa_summary)
        
            internal_codes = set(internal_df['Malzeme Kodu'].astype(str).tolist()) if len(internal_df) > 0 else set()
            chronic_codes = set(chronic_df['Malzeme Kodu'].astype(str).tolist()) if len(chronic_df) > 0 else set()
        
            # Aile dengelenmişlerini bul
            family_balanced_codes = set()
            if len(family_df) > 0:
                balanced_families = family_df[family_df['Sonuç'].str.contains('KARIŞIKLIK', na=False)]
                # Bu ailelerdeki ürünleri bul
        
            top20_df = create_top_20_risky(df_display, internal_codes, chronic_codes, family_balanced_codes)
        
            risk_seviyesi, risk_class = calculate_store_risk(df_display, internal_df, chronic_df, cigarette_df)
        
            st.markdown("---")
        
            # Metrikler hesapla
            toplam_satis = df_display['Satış Tutarı'].sum()
            fark_tutari = df_display['Fark Tutarı'].fillna(0).sum()
            kismi_tutari = df_display['Kısmi Envanter Tutarı'].fillna(0).sum()
            fire_tutari = df_display['Fire Tutarı'].fillna(0).sum()
            
            fark = fark_tutari + kismi_tutari
            toplam_acik = fark + fire_tutari
            
            fark_oran = abs(fark) / toplam_satis * 100 if toplam_satis > 0 else 0
            fire_oran = abs(fire_tutari) / toplam_satis * 100 if toplam_satis > 0 else 0
            toplam_oran = abs(toplam_acik) / toplam_satis * 100 if toplam_satis > 0 else 0
            
            # Gün hesabı
            gun_sayisi = 1
            try:
                if 'Envanter Tarihi' in df_display.columns and 'Envanter Başlangıç Tarihi' in df_display.columns:
                    env_tarihi = pd.to_datetime(df_display['Envanter Tarihi'].iloc[0])
                    env_baslangic = pd.to_datetime(df_display['Envanter Başlangıç Tarihi'].iloc[0])
                    gun_sayisi = (env_tarihi - env_baslangic).days
                    if gun_sayisi <= 0:
                        gun_sayisi = 1
            except:
                gun_sayisi = 1
            
            gunluk_fark = fark / gun_sayisi
            gunluk_fire = fire_tutari / gun_sayisi
        
            # Metrikler - Üst
            col1, col2, col3, col4, col5 = st.columns(5)
            with col1:
                st.markdown(f'<div class="{risk_class}"><b>RİSK</b><br/><h2>{risk_seviyesi}</h2></div>', unsafe_allow_html=True)
            with col2:
                st.metric("💰 Satış", f"{toplam_satis:,.0f} TL")
            with col3:
                st.metric("📉 Fark", f"{fark:,.0f} TL", f"%{fark_oran:.2f} | Günlük: {gunluk_fark:,.0f}₺")
            with col4:
                st.metric("🔥 Fire", f"{fire_tutari:,.0f} TL", f"%{fire_oran:.2f} | Günlük: {gunluk_fire:,.0f}₺")
            with col5:
                st.metric("📊 Toplam", f"{toplam_acik:,.0f} TL", f"%{toplam_oran:.2f}")
        
            # Metrikler - Alt
            col1, col2, col3, col4, col5 = st.columns(5)
            with col1:
                st.metric("🔒 İç Hırs.", f"{len(internal_df)}")
            with col2:
                st.metric("🔄 Kr.Açık", f"{len(chronic_df)}")
            with col3:
                st.metric("🔥 Kr.Fire", f"{len(chronic_fire_df)}")
            with col4:
                # Sigara açığı - toplam bazlı
                sigara_acik = 0
                if len(cigarette_df) > 0 and 'Ürün Toplam' in cigarette_df.columns:
                    son_satir = cigarette_df.iloc[-1]
                    if son_satir['Malzeme Kodu'] == '*** TOPLAM ***':
                        sigara_acik = abs(son_satir['Ürün Toplam'])
                
                if sigara_acik > 0:
                    st.metric("🚬 SİGARA", f"{sigara_acik:.0f}", delta="RİSK!", delta_color="inverse")
                else:
                    st.metric("🚬 Sigara", "0")
            with col5:
                if kasa_summary['toplam_adet'] > 0:
                    st.metric("💰 10 TL", f"+{kasa_summary['toplam_adet']:.0f} / {kasa_summary['toplam_tutar']:,.0f}₺", delta="FAZLA!", delta_color="inverse")
                elif kasa_summary['toplam_adet'] < 0:
                    st.metric("💰 10 TL", f"{kasa_summary['toplam_adet']:.0f} / {kasa_summary['toplam_tutar']:,.0f}₺", delta="AÇIK", delta_color="normal")
                else:
                    st.metric("💰 10 TL", "0")
        
            # Yönetici Özeti
            if exec_comments:
                with st.expander("📋 Yönetici Özeti", expanded=True):
                    for comment in exec_comments[:5]:
                        st.markdown(comment)
        
            st.markdown("---")
        
            # Sekmeler
            tabs = st.tabs(["🚨 Riskli 20", "🔒 İç Hırs.", "🔄 Kr.Açık", "🔥 Kr.Fire", "🔥 Fire Man.", "🚬 Sigara", "💰 10 TL Akt.", "📥 İndir"])
        
            with tabs[0]:
                st.subheader("🚨 En Riskli 20 Ürün")
                if len(top20_df) > 0:
                    st.dataframe(top20_df, use_container_width=True, hide_index=True)
                else:
                    st.success("Riskli ürün yok!")
        
            with tabs[1]:
                st.subheader("🔒 İç Hırsızlık (≥100TL)")
                st.caption("Fark büyüdükçe risk AZALIR, eşitse EN YÜKSEK")
                if len(internal_df) > 0:
                    st.dataframe(internal_df, use_container_width=True, hide_index=True)
                else:
                    st.success("İç hırsızlık riski yok!")
        
            with tabs[2]:
                st.subheader("🔄 Kronik Açık")
                st.caption("Her iki dönemde de Fark < 0")
                if len(chronic_df) > 0:
                    st.dataframe(chronic_df, use_container_width=True, hide_index=True)
                else:
                    st.success("Kronik açık yok!")
        
            with tabs[3]:
                st.subheader("🔥 Kronik Fire")
                st.caption("Her iki dönemde de fire kaydı var")
                if len(chronic_fire_df) > 0:
                    st.dataframe(chronic_fire_df, use_container_width=True, hide_index=True)
                else:
                    st.success("Kronik fire yok!")
        
            with tabs[4]:
                st.subheader("🔥 Fire Manipülasyonu")
                st.caption("Fire var ama Fark+Kısmi > 0")
                if len(fire_manip_df) > 0:
                    st.dataframe(fire_manip_df, use_container_width=True, hide_index=True)
                else:
                    st.success("Fire manipülasyonu yok!")
        
            with tabs[5]:
                st.subheader("🚬 Sigara Açığı")
                if len(cigarette_df) > 0:
                    st.error("⚠️ Sigarada açık = HIRSIZLIK BELİRTİSİ")
                    st.dataframe(cigarette_df, use_container_width=True, hide_index=True)
                else:
                    st.success("Sigara açığı yok!")
        
            with tabs[6]:
                st.subheader("💰 10 TL Aktivitesi Ürünleri")
            
                if kasa_summary['toplam_adet'] != 0:
                    if kasa_summary['toplam_adet'] > 0:
                        st.error(f"⚠️ NET +{kasa_summary['toplam_adet']:.0f} adet / {kasa_summary['toplam_tutar']:,.0f} TL FAZLA - Gerçek açığı gizliyor olabilir!")
                    else:
                        st.warning(f"📉 NET {kasa_summary['toplam_adet']:.0f} adet / {kasa_summary['toplam_tutar']:,.0f} TL AÇIK")
            
                if len(kasa_activity_df) > 0:
                    st.dataframe(kasa_activity_df, use_container_width=True, hide_index=True)
                else:
                    st.success("Kasa aktivitesi ürünlerinde sorun yok!")
        
            with tabs[7]:
                st.subheader("📥 Rapor İndir")
            
                excel_output = create_excel_report(
                    df_display, internal_df, chronic_df, chronic_fire_df, cigarette_df,
                    external_df, family_df, fire_manip_df, kasa_activity_df, top20_df,
                    exec_comments, group_stats, selected, magaza_adi, params
                )
                
                mag_adi_clean = magaza_adi.replace(' ', '_').replace('/', '_')[:30] if magaza_adi else ''
            
                st.download_button(
                    label=f"📥 {selected} Raporu İndir",
                    data=excel_output,
                    file_name=f"{selected}_{mag_adi_clean}_Risk_Raporu.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            
                if len(magazalar) > 1:
                    st.markdown("---")
                    if st.button("🗜️ Tüm Mağazaları Hazırla (ZIP)"):
                        with st.spinner("Raporlar hazırlanıyor..."):
                            zip_buffer = BytesIO()
                            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
                                for mag in magazalar:
                                    df_mag = df[df['Mağaza Kodu'] == mag].copy()
                                    mag_adi = df_mag['Mağaza Adı'].iloc[0] if 'Mağaza Adı' in df_mag.columns and len(df_mag) > 0 else ''
                                
                                    int_df = detect_internal_theft(df_mag)
                                    
                                    # Kamera timestamp entegrasyonu (kategori araması için full_df geçir)
                                    if len(int_df) > 0:
                                        try:
                                            env_tarihi = df_mag['Envanter Tarihi'].iloc[0]
                                            int_df = enrich_internal_theft_with_camera(int_df, mag, env_tarihi, full_df=df_mag)
                                        except:
                                            pass
                                    
                                    chr_df = detect_chronic_products(df_mag)
                                    chr_fire_df = detect_chronic_fire(df_mag)
                                    cig_df = detect_cigarette_shortage(df_mag)
                                    ext_df = detect_external_theft(df_mag)
                                    fam_df = find_product_families(df_mag)
                                    fire_df = detect_fire_manipulation(df_mag)
                                    kasa_df, kasa_sum = check_kasa_activity_products(df_mag, kasa_kodlari)
                                
                                    int_codes = set(int_df['Malzeme Kodu'].astype(str).tolist()) if len(int_df) > 0 else set()
                                    chr_codes = set(chr_df['Malzeme Kodu'].astype(str).tolist()) if len(chr_df) > 0 else set()
                                
                                    t20_df = create_top_20_risky(df_mag, int_codes, chr_codes, set())
                                    exec_c, grp_s = generate_executive_summary(df_mag, kasa_df, kasa_sum)
                                
                                    excel_data = create_excel_report(
                                        df_mag, int_df, chr_df, chr_fire_df, cig_df,
                                        ext_df, fam_df, fire_df, kasa_df, t20_df,
                                        exec_c, grp_s, mag, mag_adi, params
                                    )
                                
                                    zf.writestr(f"{mag}_Risk_Raporu.xlsx", excel_data.getvalue())
                        
                            zip_buffer.seek(0)
                            st.download_button(
                                label=f"📥 {len(magazalar)} Mağaza ZIP İndir",
                                data=zip_buffer,
                                file_name="Tum_Magazalar_Rapor.zip",
                                mime="application/zip"
                            )
    
    except Exception as e:
        st.error(f"Hata: {str(e)}")
        st.exception(e)

else:
    if analysis_mode != "👔 SM Özet":
        st.info("👆 Excel dosyası yükleyin")
